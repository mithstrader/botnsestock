#!/usr/bin/env python3
"""
Nifty Precision Bot — LOCAL VERSION (run from your PC)
═══════════════════════════════════════════════════════
Run this on your Windows/Mac/Linux machine during market hours.
Data fetched directly from Yahoo Finance (works from India).

SETUP:
  1. pip install yfinance pandas numpy openpyxl requests schedule pytz plyer
  2. Fill TELEGRAM_TOKEN and CHAT_ID below
  3. Double-click run_bot.bat  (Windows)  OR  python nifty_local_bot.py

Instruments : Nifty50 | BankNifty | Sensex
Strategy    : ORB + VWAP + EMA9/21 + SuperTrend + RSI + CPR + PDH/PDL + Fib + VIX/PCR
R:R         : 1:3
"""

import os, platform   # needed before config block below

# ═══════════════════════════════════════════════════════════════
#  ⚙️  FILL THESE IN — that's all you need to change
# ═══════════════════════════════════════════════════════════════

# ── Bot 1 : Every scan summary (score bars, RSI, LTP every 5 min) ──
SCAN_TG_TOKEN  = os.environ.get("SCAN_TG_TOKEN",  "8795392223:AAEc_4fwzYaai1acUeJC63_TECUkicX8hN0")
SCAN_TG_CHAT   = os.environ.get("SCAN_TG_CHAT",   "670433968")

# ── Bot 2 : Confirmed signals only (score ≥75, startup, EOD) ───────
SIGNAL_TG_TOKEN = os.environ.get("SIGNAL_TG_TOKEN", "8846280533:AAEkS0AMadIzTMcL7Gp6G7CHJHEStOAiQ8U")
SIGNAL_TG_CHAT  = os.environ.get("SIGNAL_TG_CHAT",  "670433968")

# ── Legacy single-bot fallback (leave blank if using the two above) ─
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
CHAT_ID        = os.environ.get("CHAT_ID",        "")

# ── Supabase (cloud backend) ──────────────────────────────────
# Get these from: Supabase → Project Settings → API
# Set as environment variables on Railway, or paste directly for local use.
SUPABASE_URL         = os.environ.get("SUPABASE_URL",         "https://ngvxrmgrapgyksvfqkva.supabase.co")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5ndnhybWdyYXBneWtzdmZxa3ZhIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MDU5MzEwNSwiZXhwIjoyMDk2MTY5MTA1fQ.IHIAtCpVMAjkFP8YnU4oqtLE7QrTIy8iYwXCbca4C9U")  # service_role key

# ── Excel (local logging, optional) ──────────────────────────
_default_excel = r"C:\NSE_Bot\logs" if platform.system() == "Windows" else "/tmp/nse_bot_logs"
EXCEL_FOLDER   = os.environ.get("EXCEL_FOLDER", _default_excel)

# ═══════════════════════════════════════════════════════════════
#  OPTIONAL SETTINGS
# ═══════════════════════════════════════════════════════════════

SCORE_ULTRA     = 85
SCORE_STRONG    = 75
SCORE_MEDIUM    = 65
SCAN_INTERVAL   = 5       # minutes between scans
MAX_TRADES_DAY  = 6
HEARTBEAT_SCANS = 12      # send alive ping every N scans (~60 min)
RESIGAL_GAP_MIN = 30      # minutes before re-alerting same instrument

# ── 1-min entry confirmation (for unattended / hands-free trading) ─
CONFIRM_TIMEOUT_MIN = 3   # max minutes to wait for 1-min confirmation
CONFIRM_CHECKS_MIN  = 3   # checks needed out of 4 (raise = stricter)
SKIP_UNCONFIRMED    = True # True = skip signal if no confirmation in time

INSTRUMENTS = {
    "NIFTY50":   {"ticker": "^NSEI",    "sl_pts": 40,  "lot_size": 25, "strike_step": 50,  "min_score": 75},
    "BANKNIFTY": {"ticker": "^NSEBANK", "sl_pts": 120, "lot_size": 15, "strike_step": 100, "min_score": 75},
    "SENSEX":    {"ticker": "^BSESN",   "sl_pts": 200, "lot_size": 10, "strike_step": 100, "min_score": 75},
}

RSI_PERIOD       = 14
RSI_CALL_LOW, RSI_CALL_HIGH = 40, 65
RSI_PUT_LOW,  RSI_PUT_HIGH  = 35, 60
CPR_NARROW_PCT   = 0.20
ORB_MINUTES      = 15
ST_PERIOD, ST_MULT = 10, 3.0
FIB_TOLERANCE_PCT = 0.15
FIB_RATIOS        = [0.236, 0.382, 0.500, 0.618, 0.786]
FIB_SCORE_MAP     = {0.618: 12, 0.500: 8, 0.382: 6, 0.786: 4, 0.236: 3}

VIX_EXTREME, VIX_NORMAL_MAX, VIX_CALM_MAX = 25.0, 20.0, 12.0
VIX_HIGH_PENALT, VIX_CALM_BONUS = 10, 5
PCR_VERY_BULL, PCR_BULL   = 1.5, 1.2
PCR_BEAR,      PCR_VERY_BEAR = 0.8, 0.6

MARKET_OPEN  = (9, 15)
MARKET_CLOSE = (15, 30)
BOT_START    = (9, 35)
LATE_SESSION = (14, 30)

# ═══════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════

import sys, time, subprocess, threading
from datetime import datetime, timedelta
from typing import Optional

import pytz, schedule, requests

try:
    from supabase import create_client as _sb_create
    _sb_available = True
except ImportError:
    _sb_available = False

IST = pytz.timezone("Asia/Kolkata")

try:
    import yfinance as yf
    import numpy as np
    import pandas as pd
    _yf = True
except ImportError:
    _yf = False
    print("❌ Missing: pip install yfinance pandas numpy")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _openpyxl = True
except ImportError:
    _openpyxl = False
    print("⚠ openpyxl not found — Excel logging disabled. pip install openpyxl")

_OS = platform.system()

# ═══════════════════════════════════════════════════════════════
#  STATE
# ═══════════════════════════════════════════════════════════════

state = {
    "scan_count":       0,
    "alerts_sent":      0,
    "trades_today":     0,
    "vix":              None,
    "pcr":              None,
    "last_vix_fetch":   None,
    "last_signal":      {},
    "data_errors":      {},
    "scores_seen":      [],
    # ── Order flow ───────────────────────────────────────────
    "fii_net":          None,   # FII cash market net (Cr)
    "dii_net":          None,   # DII cash market net (Cr)
    "fii_deriv_net":    None,   # FII index futures net (Cr)
    "last_fii_fetch":   None,
    "options_flow":     {},     # keyed by symbol: NIFTY/BANKNIFTY/SENSEX
    "dom":              {},     # keyed by symbol
}

# ═══════════════════════════════════════════════════════════════
#  INDICATORS
# ═══════════════════════════════════════════════════════════════

def calc_ema(series, period):
    return series.ewm(span=period, adjust=False).mean()

def calc_rsi(series, period=14):
    delta = series.diff()
    gain  = delta.clip(lower=0).ewm(com=period-1, adjust=False).mean()
    loss  = (-delta.clip(upper=0)).ewm(com=period-1, adjust=False).mean()
    rs    = gain / loss.replace(0, 1e-10)
    return float((100 - 100/(1+rs)).iloc[-1])

def calc_vwap(df):
    tp  = (df["High"] + df["Low"] + df["Close"]) / 3
    vol = df["Volume"]
    if vol.sum() == 0:
        return float(tp.iloc[-1])
    return float((tp * vol).cumsum().iloc[-1] / vol.cumsum().iloc[-1])

def calc_atr(df, period=14):
    hl = df["High"] - df["Low"]
    hc = (df["High"] - df["Close"].shift()).abs()
    lc = (df["Low"]  - df["Close"].shift()).abs()
    return pd.concat([hl, hc, lc], axis=1).max(axis=1).ewm(com=period-1, adjust=False).mean()

def calc_supertrend(df, period=10, mult=3.0):
    if len(df) < period + 2:
        return 1 if float(df["Close"].iloc[-1]) > float(df["Close"].iloc[0]) else -1
    hl2 = (df["High"] + df["Low"]) / 2
    atr_ = calc_atr(df, period)
    upper_r = hl2 + mult * atr_
    lower_r = hl2 - mult * atr_
    upper = upper_r.copy()
    lower = lower_r.copy()
    trend = pd.Series(1, index=df.index, dtype=float)
    for i in range(1, len(df)):
        upper.iloc[i] = upper_r.iloc[i] if upper_r.iloc[i] < upper.iloc[i-1] or df["Close"].iloc[i-1] > upper.iloc[i-1] else upper.iloc[i-1]
        lower.iloc[i] = lower_r.iloc[i] if lower_r.iloc[i] > lower.iloc[i-1] or df["Close"].iloc[i-1] < lower.iloc[i-1] else lower.iloc[i-1]
        if   trend.iloc[i-1] == -1 and df["Close"].iloc[i] > upper.iloc[i]: trend.iloc[i] = 1
        elif trend.iloc[i-1] ==  1 and df["Close"].iloc[i] < lower.iloc[i]: trend.iloc[i] = -1
        else: trend.iloc[i] = trend.iloc[i-1]
    return int(trend.iloc[-1])

def calc_cpr(ph, pl, pc):
    pivot = (ph + pl + pc) / 3
    bc    = (ph + pl) / 2
    tc    = 2 * pivot - bc
    w     = abs(tc - bc) / pivot * 100
    return {"pivot": pivot, "top": max(tc,bc), "bottom": min(tc,bc),
            "width_pct": w, "narrow": w < CPR_NARROW_PCT}

def calc_fib(df_today, direction, orb_h, orb_l):
    empty = {"level":None,"price":None,"dist_pct":None,"at_fib":False,
             "score_bonus":0,"levels":{},"swing_high":None,"swing_low":None}
    if df_today is None or len(df_today) < 3:
        return empty
    ltp = float(df_today["Close"].iloc[-1])
    if direction == "CALL":
        sl, sh = orb_l, float(df_today["High"].max())
        move = sh - sl
        if move < 1: return empty
        raw = {r: sh - move * r for r in FIB_RATIOS}
    else:
        sh, sl = orb_h, float(df_today["Low"].min())
        move = sh - sl
        if move < 1: return empty
        raw = {r: sl + move * r for r in FIB_RATIOS}
    nearest = min(raw, key=lambda r: abs(raw[r]-ltp))
    dist_pct = abs(ltp - raw[nearest]) / ltp * 100
    at_fib   = dist_pct <= FIB_TOLERANCE_PCT
    return {
        "level":       f"{nearest:.3f}",
        "price":       round(raw[nearest], 1),
        "dist_pct":    round(dist_pct, 3),
        "at_fib":      at_fib,
        "score_bonus": FIB_SCORE_MAP.get(nearest, 0) if at_fib else 0,
        "levels":      {f"{r:.3f}": round(p, 1) for r, p in raw.items()},
        "swing_high":  round(sh, 1),
        "swing_low":   round(sl, 1),
    }

# ═══════════════════════════════════════════════════════════════
#  SUPABASE CLIENT + PUSH FUNCTIONS
# ═══════════════════════════════════════════════════════════════

_sb_client = None

def get_sb():
    global _sb_client
    if _sb_client is None and _sb_available and SUPABASE_URL and SUPABASE_SERVICE_KEY:
        try:
            _sb_client = _sb_create(SUPABASE_URL, SUPABASE_SERVICE_KEY)
        except Exception as e:
            print(f"  [SUPABASE] init error: {e}")
    return _sb_client

def sb_push_signal(sig):
    """Push a triggered signal to Supabase signals table."""
    client = get_sb()
    if client is None: return
    try:
        now = datetime.now(IST)
        client.table("signals").insert({
            "signal_date":  now.strftime("%Y-%m-%d"),
            "signal_time":  sig["time"],
            "instrument":   sig["instrument"],
            "direction":    sig["direction"],
            "rating":       sig["rating"],
            "score":        sig["score"],
            "ltp":          round(float(sig["ltp"]), 2),
            "strike":       int(sig["strike"]),
            "option_type":  sig["option_type"],
            "sl_level":     round(float(sig["sl_level"]), 2),
            "target_level": round(float(sig["target_level"]), 2),
            "sl_pts":       round(float(sig["sl_pts"]), 2),
            "target_pts":   round(float(sig["target_pts"]), 2),
            "rsi":          round(float(sig["rsi"]), 2),
            "vwap":         round(float(sig["vwap"]), 2),
            "ema9":         round(float(sig["e9"]), 2),
            "ema21":        round(float(sig["e21"]), 2),
            "supertrend":   int(sig["st_dir"]),
            "vix":          float(state["vix"]) if state["vix"] else None,
            "pcr":          float(state["pcr"]) if state["pcr"] else None,
            "fib_level":    sig["fib"]["level"],
            "fib_price":    sig["fib"]["price"],
            "fib_at_level": bool(sig["fib"]["at_fib"]),
            "fib_bonus":    int(sig["fib"]["score_bonus"]),
            "orb_high":     round(float(sig["orb_high"]), 2),
            "orb_low":      round(float(sig["orb_low"]), 2),
            "pdh":          round(float(sig["pdh"]), 2),
            "pdl":          round(float(sig["pdl"]), 2),
            "cpr_top":      round(float(sig["cpr"]["top"]), 2),
            "cpr_bottom":   round(float(sig["cpr"]["bottom"]), 2),
            "cpr_narrow":   bool(sig["cpr"]["narrow"]),
            "ema_ok":       bool(sig["conditions"].get("EMA 9/21")),
            "vwap_ok":      bool(sig["conditions"].get("VWAP")),
            "orb_ok":       bool(sig["conditions"].get("ORB")),
            "cpr_ok":       bool(sig["conditions"].get("CPR")),
            "pdh_ok":       bool(sig["conditions"].get("PDH/PDL")),
            "rsi_ok":       bool(sig["conditions"].get("RSI")),
        }).execute()
        print(f"  [SUPABASE] ✅ signal pushed for {sig['instrument']}")
    except Exception as e:
        print(f"  [SUPABASE] push_signal error: {e}")

def sb_push_scan(scan_no, results):
    """Push per-scan summary to Supabase scans table."""
    client = get_sb()
    if client is None:
        if scan_no == 1:   # warn only on first scan to avoid spam
            print("  [SUPABASE] ⚠️  client is None — check URL & SERVICE_KEY")
        return
    try:
        now = datetime.now(IST)
        row = {
            "scan_date":     now.strftime("%Y-%m-%d"),
            "scan_time":     now.strftime("%H:%M"),
            "scan_number":   scan_no,
            "vix":           float(state["vix"]) if state["vix"] else None,
            "pcr":           float(state["pcr"]) if state["pcr"] else None,
            "signals_count": sum(1 for _, _, _, _, alerted in results if alerted),
        }
        footprint = {}
        for name, sig, _, _, _ in results:
            k = name.lower()
            if sig:
                row[f"{k}_score"] = int(sig["score"])
                row[f"{k}_dir"]   = sig["direction"]
                row[f"{k}_rsi"]   = round(float(sig["rsi"]), 2)
                row[f"{k}_ltp"]   = round(float(sig["ltp"]), 2)
                fp = sig.get("footprint")
                if fp:
                    footprint[name] = fp

        # Push core scan first (always safe)
        client.table("scans").insert(row).execute()
        print(f"  [SUPABASE] ✅ scan #{scan_no} pushed")

        # Push footprint separately — column may not exist yet on older DBs
        if footprint:
            try:
                client.table("scans") \
                      .update({"footprint_data": footprint}) \
                      .eq("scan_number", scan_no) \
                      .eq("scan_date", now.strftime("%Y-%m-%d")) \
                      .execute()
            except Exception as fe:
                print(f"  [SUPABASE] footprint update skipped: {fe}")

    except Exception as e:
        print(f"  [SUPABASE] push_scan error: {e}")

def sb_update_status(status="running"):
    """Update the single bot_status row."""
    client = get_sb()
    if client is None: return
    try:
        now = datetime.now(IST)
        row = {
            "id":           1,
            "updated_at":   now.isoformat(),
            "status":       status,
            "scan_count":   state["scan_count"],
            "alerts_today": state["alerts_sent"],
            "session_date": now.strftime("%Y-%m-%d"),
            "vix":          float(state["vix"]) if state["vix"] else None,
            "pcr":          float(state["pcr"]) if state["pcr"] else None,
        }
        if status == "running":
            row["last_scan_at"] = now.isoformat()
        if status == "sleeping":
            row["next_session_at"] = _next_market_open().isoformat()
        client.table("bot_status").upsert(row).execute()
        if state["scan_count"] <= 1:   # print only on startup
            print(f"  [SUPABASE] ✅ bot_status updated ({status})")
    except Exception as e:
        print(f"  [SUPABASE] update_status error: {e}")

# ═══════════════════════════════════════════════════════════════
#  DATA FETCHING
# ═══════════════════════════════════════════════════════════════

_NSE_HDR = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept":          "application/json",
    "Referer":         "https://www.nseindia.com/",
}
_nse_sess = None

def _nse_session():
    global _nse_sess
    if _nse_sess is None:
        _nse_sess = requests.Session()
        _nse_sess.headers.update(_NSE_HDR)
        try: _nse_sess.get("https://www.nseindia.com", timeout=10)
        except: pass
    return _nse_sess

def fetch_vix_pcr():
    vix_val = pcr_val = None
    try:
        r = _nse_session().get("https://www.nseindia.com/api/allIndices", timeout=10)
        if r.status_code == 200:
            for item in r.json().get("data", []):
                if item.get("index") == "INDIA VIX":
                    vix_val = float(item["last"]); break
    except Exception as e:
        print(f"  [VIX] {e}")
    try:
        r = _nse_session().get(
            "https://www.nseindia.com/api/option-chain-indices?symbol=NIFTY", timeout=15)
        if r.status_code == 200:
            data  = r.json()["records"]["data"]
            ce_oi = sum(d["CE"]["openInterest"] for d in data if "CE" in d)
            pe_oi = sum(d["PE"]["openInterest"] for d in data if "PE" in d)
            pcr_val = round(pe_oi / ce_oi, 3) if ce_oi else None
    except Exception as e:
        print(f"  [PCR] {e}")
    return vix_val, pcr_val

# ── FII / DII ─────────────────────────────────────────────────────
def fetch_fii_dii():
    """Fetch FII & DII cash market + FII derivatives data from NSE.
    Returns (fii_cash_net, dii_cash_net, fii_deriv_net) in Crores."""
    fii_cash = dii_cash = fii_deriv = None
    try:
        r = _nse_session().get(
            "https://www.nseindia.com/api/fiidiiTradeReact", timeout=12)
        if r.status_code == 200:
            for row in r.json():
                cat = row.get("category", "").upper()
                net = row.get("netValue") or row.get("net") or "0"
                try:
                    val = float(str(net).replace(",", ""))
                except Exception:
                    val = 0.0
                if "FII" in cat or "FPI" in cat:
                    fii_cash = val
                elif "DII" in cat:
                    dii_cash = val
    except Exception as e:
        print(f"  [FII/DII cash] {e}")
    try:
        r = _nse_session().get(
            "https://www.nseindia.com/api/fiiDerivStat", timeout=12)
        if r.status_code == 200:
            rows = r.json().get("data", [])
            if rows:
                latest = rows[0]
                buy  = float(str(latest.get("indexFutBuyValue",  "0")).replace(",","") or 0)
                sell = float(str(latest.get("indexFutSellValue", "0")).replace(",","") or 0)
                fii_deriv = round(buy - sell, 2)
    except Exception as e:
        print(f"  [FII deriv] {e}")
    return fii_cash, dii_cash, fii_deriv

# ── Options flow: Max Pain + ATM OI + OI walls + COI ──────────────
def _calc_max_pain(strikes: dict) -> Optional[float]:
    """Strike price where total options buyer losses are maximum (market pins)."""
    if not strikes:
        return None
    best_strike, min_loss = None, float("inf")
    for test in sorted(strikes):
        loss = sum(
            (test - s) * d["ce_oi"] if test > s else
            (s - test) * d["pe_oi"] if test < s else 0
            for s, d in strikes.items()
        )
        if loss < min_loss:
            min_loss, best_strike = loss, test
    return best_strike

def fetch_options_flow(symbol: str = "NIFTY", ltp: float = None) -> dict:
    """Full options chain analysis: max pain, ATM OI, OI walls, COI bias."""
    empty = {"max_pain": None, "atm_bias": None, "resistance": None,
             "support": None, "oi_change_bias": None, "atm_ce_oi": 0, "atm_pe_oi": 0}
    try:
        r = _nse_session().get(
            f"https://www.nseindia.com/api/option-chain-indices?symbol={symbol}",
            timeout=15)
        if r.status_code != 200:
            return empty
        js    = r.json()["records"]
        expiries = js.get("expiryDates", [])
        if not expiries:
            return empty
        nearest = expiries[0]
        strikes = {}
        for rec in js.get("data", []):
            if rec.get("expiryDate") != nearest:
                continue
            sp = rec.get("strikePrice", 0)
            strikes[sp] = {
                "ce_oi":  rec.get("CE", {}).get("openInterest",          0),
                "pe_oi":  rec.get("PE", {}).get("openInterest",          0),
                "ce_coi": rec.get("CE", {}).get("changeinOpenInterest",  0),
                "pe_coi": rec.get("PE", {}).get("changeinOpenInterest",  0),
            }
        if not strikes:
            return empty
        max_pain = _calc_max_pain(strikes)
        # ATM strike
        atm = min(strikes, key=lambda x: abs(x - ltp)) if ltp else None
        atm_d = strikes.get(atm, {}) if atm else {}
        atm_bias = ("CALL" if atm_d.get("pe_oi", 0) > atm_d.get("ce_oi", 0)
                    else "PUT") if atm_d else None
        # OI walls (highest OI outside ATM)
        above = {s: d for s, d in strikes.items() if ltp and s > ltp}
        below = {s: d for s, d in strikes.items() if ltp and s < ltp}
        resistance = max(above, key=lambda x: above[x]["ce_oi"]) if above else None
        support    = max(below, key=lambda x: below[x]["pe_oi"]) if below else None
        # COI bias
        tot_ce_coi = sum(d["ce_coi"] for d in strikes.values())
        tot_pe_coi = sum(d["pe_coi"] for d in strikes.values())
        oi_change_bias = "BULL" if tot_pe_coi > tot_ce_coi else "BEAR"
        return {
            "max_pain":       max_pain,
            "atm_bias":       atm_bias,
            "atm_ce_oi":      atm_d.get("ce_oi", 0),
            "atm_pe_oi":      atm_d.get("pe_oi", 0),
            "resistance":     resistance,
            "support":        support,
            "oi_change_bias": oi_change_bias,
            "tot_ce_coi":     tot_ce_coi,
            "tot_pe_coi":     tot_pe_coi,
        }
    except Exception as e:
        print(f"  [OPTIONS FLOW {symbol}] {e}")
        return empty

# ── Volume Delta (footprint approximation from OHLCV) ─────────────
def calc_volume_delta(df_today: pd.DataFrame) -> dict:
    """Approximate buy vs sell volume using candle body position in range.
    Positive delta = net buying pressure (footprint chart substitute)."""
    empty = {"buy_vol": 0, "sell_vol": 0, "delta": 0,
             "delta_pct": 0.0, "recent_delta": 0, "bias": "NEUTRAL"}
    if df_today is None or df_today.empty:
        return empty
    try:
        hl = (df_today["High"] - df_today["Low"]).replace(0, np.nan)
        buy_ratio  = ((df_today["Close"] - df_today["Low"]) / hl).fillna(0.5)
        vol        = df_today["Volume"].fillna(0)
        buy_vol    = int((vol * buy_ratio).sum())
        sell_vol   = int((vol * (1 - buy_ratio)).sum())
        delta      = buy_vol - sell_vol
        total      = buy_vol + sell_vol or 1
        delta_pct  = round(delta / total * 100, 1)
        # Last 30 min (6 × 5-min candles)
        recent     = df_today.tail(6)
        hl_r       = (recent["High"] - recent["Low"]).replace(0, np.nan)
        br_r       = ((recent["Close"] - recent["Low"]) / hl_r).fillna(0.5)
        recent_delta = int(((recent["Volume"].fillna(0)) * (2 * br_r - 1)).sum())
        bias = ("BULL" if recent_delta > 0 else
                "BEAR" if recent_delta < 0 else "NEUTRAL")
        return {"buy_vol": buy_vol, "sell_vol": sell_vol, "delta": delta,
                "delta_pct": delta_pct, "recent_delta": recent_delta, "bias": bias}
    except Exception as e:
        print(f"  [VOL DELTA] {e}")
        return empty

# ── Footprint: per-candle volume delta across price levels ────────
def calc_candle_footprint(df_today: pd.DataFrame, max_candles: int = 15) -> list:
    """
    Per-candle buy/sell volume approximation across 7 price levels.
    Used by the web dashboard to render the footprint chart.
    NOTE: true footprint requires tick data; this approximates using
          candle body position (buy vol ∝ close-low / high-low).
    Returns list of dicts ordered oldest → newest.
    """
    if df_today is None or df_today.empty:
        return []
    df = df_today.tail(max_candles).copy()
    rows = []
    for ts, row in df.iterrows():
        hl       = float(row["High"]) - float(row["Low"])
        if hl == 0:
            hl = 1e-9
        vol      = float(row.get("Volume", 0) or 0)
        buy_rat  = (float(row["Close"]) - float(row["Low"])) / hl
        buy_vol  = int(vol * buy_rat)
        sell_vol = int(vol - buy_vol)
        delta    = buy_vol - sell_vol
        delta_p  = round(delta / (vol or 1) * 100, 1)
        rows.append({
            "t":         ts.strftime("%H:%M"),
            "o":         round(float(row["Open"]),  1),
            "h":         round(float(row["High"]),  1),
            "l":         round(float(row["Low"]),   1),
            "c":         round(float(row["Close"]), 1),
            "v":         int(vol),
            "buy_vol":   buy_vol,
            "sell_vol":  sell_vol,
            "delta":     delta,
            "delta_pct": delta_p,
        })
    return rows

# ── DOM: Bid/Ask depth from NSE Futures ───────────────────────────
def fetch_dom_imbalance(symbol: str = "NIFTY") -> dict:
    """Get top-5 bid/ask from NSE futures market depth.
    Imbalance > 0 = more buyers, < 0 = more sellers."""
    empty = {"imbalance": None, "best_bid": None, "best_ask": None,
             "total_bid_qty": 0, "total_ask_qty": 0}
    try:
        r = _nse_session().get(
            f"https://www.nseindia.com/api/quote-derivative?symbol={symbol}",
            timeout=12)
        if r.status_code != 200:
            return empty
        stocks = r.json().get("stocks", [])
        # Find nearest futures contract
        fut = next((s for s in stocks
                    if s.get("metadata", {}).get("instrumentType") == "Index Futures"), None)
        if fut is None:
            return empty
        depth = fut.get("marketDepth", {})
        bids  = depth.get("buy",  [])
        asks  = depth.get("sell", [])
        bid_qty = sum(b.get("quantity", 0) for b in bids)
        ask_qty = sum(a.get("quantity", 0) for a in asks)
        total   = bid_qty + ask_qty or 1
        imbalance = round((bid_qty - ask_qty) / total, 3)   # +1=all bids, -1=all asks
        best_bid  = bids[0].get("price") if bids else None
        best_ask  = asks[0].get("price") if asks else None
        return {"imbalance": imbalance, "best_bid": best_bid, "best_ask": best_ask,
                "total_bid_qty": bid_qty, "total_ask_qty": ask_qty}
    except Exception as e:
        print(f"  [DOM {symbol}] {e}")
        return empty

# ── Refresh order flow (called in refresh_vix_pcr) ────────────────
def refresh_fii_dii():
    """Fetch FII/DII once per day (data is end-of-day from NSE)."""
    last = state.get("last_fii_fetch")
    now  = datetime.now(IST)
    if last and last.date() == now.date() and (now - last).total_seconds() < 3600:
        return
    fii, dii, fii_d = fetch_fii_dii()
    if fii  is not None: state["fii_net"]       = fii
    if dii  is not None: state["dii_net"]       = dii
    if fii_d is not None: state["fii_deriv_net"] = fii_d
    state["last_fii_fetch"] = now
    print(f"  [FII/DII] Cash FII={fii} DII={dii} | Deriv FII={fii_d} Cr")

def _normalise_df(df) -> Optional[pd.DataFrame]:
    """Flatten MultiIndex columns, convert index to IST, drop NaNs."""
    if df is None or df.empty:
        return None
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    # Rename 'Price' level leftover from some yf versions
    df.columns = [c.strip().capitalize() for c in df.columns]
    try:
        if df.index.tz is None:
            df.index = df.index.tz_localize("UTC").tz_convert(IST)
        else:
            df.index = df.index.tz_convert(IST)
    except Exception:
        pass
    df = df.dropna(subset=["Close", "High", "Low"])
    return df if len(df) >= 20 else None


def fetch_ohlcv(ticker: str) -> Optional[pd.DataFrame]:
    """Fetch 5-min OHLCV.  Tries Ticker.history() first (more robust),
    then falls back to yf.download()."""

    # ── Method 1: Ticker.history() ──────────────────────────────
    for period in ("5d", "7d", "10d"):
        try:
            tk  = yf.Ticker(ticker)
            df  = tk.history(period=period, interval="5m", auto_adjust=True)
            out = _normalise_df(df)
            if out is not None:
                return out
        except Exception as e:
            print(f"  [DATA] {ticker} Ticker.history period={period}: {e}")

    # ── Method 2: yf.download() fallback ────────────────────────
    for period in ("5d", "7d", "1mo"):
        try:
            df  = yf.download(ticker, period=period, interval="5m",
                              progress=False, auto_adjust=True,
                              threads=False)
            out = _normalise_df(df)
            if out is not None:
                return out
        except Exception as e:
            print(f"  [DATA] {ticker} download period={period}: {e}")

    return None

def split_sessions(df):
    today = datetime.now(IST).date()
    td = df[df.index.date == today].copy()
    prev_days = sorted(set(df.index.date))
    prev_days = [d for d in prev_days if d < today]
    prev = df[df.index.date == prev_days[-1]].copy() if prev_days else pd.DataFrame()
    return td, prev

# ── 1-Minute Data & Entry Confirmation ────────────────────────────
def _safe_to_ist(df: pd.DataFrame) -> pd.DataFrame:
    """Robust index → IST conversion that bypasses the yfinance
    'NoneType has no attribute timezone' bug on 1-min NSE data."""
    try:
        idx = pd.to_datetime(df.index)          # force proper DatetimeIndex
        if idx.tz is None:
            idx = idx.tz_localize("UTC")
        df.index = idx.tz_convert(IST)
    except Exception:
        try:
            # Last-resort: stringify and re-parse
            df.index = (pd.to_datetime([str(x) for x in df.index])
                          .tz_localize("UTC").tz_convert(IST))
        except Exception:
            pass
    return df


def fetch_1min_ohlcv(ticker: str) -> Optional[pd.DataFrame]:
    """Fetch last 30 min of 1-min OHLCV.
    Tries yf.download() first (avoids NSE timezone bug), then Ticker.history()."""
    today = datetime.now(IST).date()

    def _clean(df):
        if df is None or df.empty:
            return None
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        df.columns = [c.strip().capitalize() for c in df.columns]
        if "Close" not in df.columns:
            return None
        df = _safe_to_ist(df)
        df = df.dropna(subset=["Close", "High", "Low"])
        df = df[df.index.date == today]
        return df.tail(30) if len(df) >= 3 else None

    # Method 1: yf.download (more stable for NSE 1-min)
    try:
        df = yf.download(ticker, period="1d", interval="1m",
                         progress=False, auto_adjust=True, threads=False)
        out = _clean(df)
        if out is not None:
            return out
    except Exception as e:
        if "timezone" not in str(e).lower():
            print(f"  [1MIN download] {ticker}: {e}")

    # Method 2: Ticker.history fallback
    for period in ("1d", "2d"):
        try:
            tk  = yf.Ticker(ticker)
            df  = tk.history(period=period, interval="1m", auto_adjust=True)
            out = _clean(df)
            if out is not None:
                return out
        except Exception as e:
            if "timezone" not in str(e).lower():
                print(f"  [1MIN history {period}] {ticker}: {e}")

    return None


def confirm_1min_entry(sig: dict) -> dict:
    """
    Score the last 5 x 1-min candles across 5 checks.
    Caller decides the pass threshold (CONFIRM_CHECKS_MIN).

    Checks:
        ① Candle body direction agrees with signal
        ② 1-min Volume Delta bias agrees
        ③ Price above/below 1-min VWAP
        ④ EMA-9 slope direction agrees
        ⑤ Volume spike on last candle (>1.3× recent average)

    Returns:
        confirmed     : True if checks_met >= CONFIRM_CHECKS_MIN
        checks_met    : int 0-5
        total_checks  : 5
        entry_price   : current LTP at time of check
        tight_sl      : 1-min wick SL (always computed, never None if data ok)
        tight_sl_pts  : SL distance in pts
        tight_target  : tight_sl_pts * 3 target
        reason        : human-readable check list
        vol_delta_1m  : full volume delta dict
        timed_out     : False (set to True externally on timeout)
    """
    base = {
        "confirmed": False, "checks_met": 0, "total_checks": 5,
        "entry_label": "NO DATA", "entry_price": sig["ltp"],
        "tight_sl": None, "tight_sl_pts": None, "tight_target": None,
        "reason": "No 1-min data available",
        "vol_delta_1m": {}, "timed_out": False,
    }

    df1 = fetch_1min_ohlcv(sig["ticker"])
    if df1 is None or len(df1) < 3:
        base["reason"] = "1-min feed unavailable"
        return base

    recent    = df1.tail(5).copy()
    direction = sig["direction"]
    ltp_now   = float(recent["Close"].iloc[-1])
    last_c    = recent.iloc[-1]
    checks    = {}

    # ① Candle body in direction
    candle_bull = float(last_c["Close"]) > float(last_c["Open"])
    checks["① Candle body"] = (
        (candle_bull  and direction == "CALL") or
        (not candle_bull and direction == "PUT")
    )

    # ② 1-min Volume Delta
    vd1 = calc_volume_delta(recent)
    checks["② Vol Δ bias"] = (
        (vd1["bias"] == "BULL" and direction == "CALL") or
        (vd1["bias"] == "BEAR" and direction == "PUT")
    )

    # ③ Price vs 1-min VWAP
    vwap1 = calc_vwap(recent)
    checks["③ VWAP 1m"] = (
        (ltp_now > vwap1 and direction == "CALL") or
        (ltp_now < vwap1 and direction == "PUT")
    )

    # ④ EMA-9 slope
    ema9_1m  = calc_ema(recent["Close"], min(9, len(recent)))
    slope_up = float(ema9_1m.iloc[-1]) > float(ema9_1m.iloc[-2])
    checks["④ EMA9 slope"] = (
        (slope_up     and direction == "CALL") or
        (not slope_up and direction == "PUT")
    )

    # ⑤ Volume spike — last 1-min candle volume > 1.3× average of prior candles
    if len(recent) >= 3:
        avg_vol  = float(recent["Volume"].iloc[:-1].mean()) or 1
        last_vol = float(recent["Volume"].iloc[-1])
        checks["⑤ Vol spike"] = last_vol > avg_vol * 1.3
    else:
        checks["⑤ Vol spike"] = False

    checks_met = sum(checks.values())
    confirmed  = checks_met >= CONFIRM_CHECKS_MIN

    # ── 1-min wick SL (always compute when data available) ─────────
    buffer = round(sig["ltp"] * 0.0005, 1)   # 0.05% buffer
    if direction == "CALL":
        wick_sl  = round(float(recent["Low"].min()) - buffer, 1)
        sl_pts   = round(ltp_now - wick_sl, 1)
        tgt      = round(ltp_now + sl_pts * 3, 1)
    else:
        wick_sl  = round(float(recent["High"].max()) + buffer, 1)
        sl_pts   = round(wick_sl - ltp_now, 1)
        tgt      = round(ltp_now - sl_pts * 3, 1)
    # Sanity: must be positive and not wider than 5-min SL
    if 0 < sl_pts <= sig["sl_pts"] * 1.2:
        tight_sl, tight_sl_pts, tight_tgt = wick_sl, sl_pts, tgt
    else:
        tight_sl = tight_sl_pts = tight_tgt = None

    parts       = [f"{'✅' if v else '❌'} {k}" for k, v in checks.items()]
    entry_label = "✅ CONFIRMED — ENTER NOW" if confirmed else "⏳ NOT YET — WAITING"

    return {
        "confirmed":    confirmed,
        "checks_met":   checks_met,
        "total_checks": 5,
        "entry_label":  entry_label,
        "entry_price":  ltp_now,
        "tight_sl":     tight_sl,
        "tight_sl_pts": tight_sl_pts,
        "tight_target": tight_tgt,
        "reason":       "\n  ".join(parts),
        "vol_delta_1m": vd1,
        "timed_out":    False,
    }


def wait_for_1min_confirm(sig: dict) -> dict:
    """
    Polls confirm_1min_entry every 30 s for up to CONFIRM_TIMEOUT_MIN minutes.
    Returns as soon as CONFIRM_CHECKS_MIN/5 checks pass, or on timeout.
    Sends a 'pending' notice to the scan bot at the start.
    """
    direction = sig["direction"]
    instr     = sig["instrument"]
    score     = sig["score"]
    deadline  = datetime.now(IST) + timedelta(minutes=CONFIRM_TIMEOUT_MIN)

    send_scan_tg(
        f"⏳ *{instr}* {direction} Score {score} — "
        f"waiting up to {CONFIRM_TIMEOUT_MIN}m for 1-min confirmation\n"
        f"Need {CONFIRM_CHECKS_MIN}/5 checks to fire trade alert."
    )
    print(f"  [CONFIRM] Waiting up to {CONFIRM_TIMEOUT_MIN}m …")

    attempt        = 0
    last_ec        = None
    consec_unavail = 0   # consecutive "data unavailable" results

    while datetime.now(IST) < deadline:
        attempt += 1
        ec      = confirm_1min_entry(sig)
        last_ec = ec
        print(f"    [1MIN #{attempt}] {ec['checks_met']}/5  "
              f"{ec['reason'].replace(chr(10),' | ')}")

        # ── confirmed → fire immediately ─────────────────────────
        if ec["confirmed"]:
            ec["timed_out"] = False
            return ec

        # ── data feed broken → don't waste full timeout ──────────
        if "unavailable" in ec.get("reason", "").lower():
            consec_unavail += 1
            if consec_unavail >= 2:
                print(f"  [1MIN] Feed unavailable after 2 tries — bypassing gate")
                ec["timed_out"]        = False
                ec["data_unavailable"] = True
                ec["entry_label"]      = "⚠️ 1-MIN DATA UNAVAILABLE"
                return ec
        else:
            consec_unavail = 0

        remaining = (deadline - datetime.now(IST)).total_seconds()
        if remaining > 30:
            time.sleep(30)
        else:
            break

    # Final check at deadline
    ec = confirm_1min_entry(sig)
    if ec["confirmed"]:
        ec["timed_out"] = False
        return ec

    # Genuine timeout (checks failed, data was available)
    if last_ec is None:
        last_ec = ec
    last_ec["timed_out"]   = True
    last_ec["entry_label"] = "⛔ TIMEOUT — SIGNAL SKIPPED"
    last_ec["confirmed"]   = False
    return last_ec

def get_orb(df_today):
    if df_today.empty: return 0.0, 0.0
    end = df_today.index[0].replace(
        hour=MARKET_OPEN[0], minute=MARKET_OPEN[1]+ORB_MINUTES, second=0)
    orb = df_today[df_today.index < end]
    if orb.empty: orb = df_today.iloc[:3]
    return float(orb["High"].max()), float(orb["Low"].min())

# ═══════════════════════════════════════════════════════════════
#  SIGNAL ANALYSIS
# ═══════════════════════════════════════════════════════════════

def analyze(name, cfg):
    df = fetch_ohlcv(cfg["ticker"])
    if df is None or len(df) < 20:
        _data_error(name, f"No data from yfinance for {cfg['ticker']}")
        return None

    df_today, df_prev = split_sessions(df)
    if df_today.empty or len(df_today) < 5:
        _data_error(name, "No today's candles in data")
        return None
    if df_prev.empty:
        _data_error(name, "No previous day data (needed for CPR/PDH/PDL)")
        return None

    state["data_errors"][name] = 0

    pdh = float(df_prev["High"].max())
    pdl = float(df_prev["Low"].min())
    pdc = float(df_prev["Close"].iloc[-1])
    cpr = calc_cpr(pdh, pdl, pdc)

    close = df_today["Close"]
    ltp   = float(close.iloc[-1])
    e9    = float(calc_ema(close, 9).iloc[-1])
    e21   = float(calc_ema(close, 21).iloc[-1])
    vwap_ = calc_vwap(df_today)
    rsi_  = calc_rsi(close, RSI_PERIOD)
    st_   = calc_supertrend(df_today, ST_PERIOD, ST_MULT)
    orb_h, orb_l = get_orb(df_today)

    bull = sum([e9>e21, ltp>vwap_, st_==1, ltp>orb_h])
    bear = sum([e9<e21, ltp<vwap_, st_==-1, ltp<orb_l])
    direction   = "CALL" if bull >= bear else "PUT"
    option_type = "CE"   if direction == "CALL" else "PE"

    fib = calc_fib(df_today, direction, orb_h, orb_l)

    score, cond = 0, {}

    ema_ok  = (e9>e21 and direction=="CALL") or (e9<e21 and direction=="PUT")
    cond["EMA 9/21"] = ema_ok;  score += 15 if ema_ok else 0

    vwap_ok = (ltp>vwap_ and direction=="CALL") or (ltp<vwap_ and direction=="PUT")
    cond["VWAP"] = vwap_ok;     score += 15 if vwap_ok else 0

    st_ok  = (st_==1 and direction=="CALL") or (st_==-1 and direction=="PUT")
    cond["SuperTrend"] = st_ok; score += 15 if st_ok else 0

    orb_ok = (ltp>orb_h and direction=="CALL") or (ltp<orb_l and direction=="PUT")
    near_orb = abs(ltp-(orb_h if direction=="CALL" else orb_l))/ltp < 0.0005
    cond["ORB"] = orb_ok;       score += 20 if orb_ok else (8 if near_orb else 0)

    rsi_ok = (RSI_CALL_LOW<=rsi_<=RSI_CALL_HIGH) if direction=="CALL" else (RSI_PUT_LOW<=rsi_<=RSI_PUT_HIGH)
    cond["RSI"] = rsi_ok;       score += 10 if rsi_ok else 0

    cpr_ok = (ltp>cpr["top"] and cpr["narrow"]) if direction=="CALL" else (ltp<cpr["bottom"] and cpr["narrow"])
    cond["CPR"] = cpr_ok;       score += 10 if cpr_ok else (4 if cpr["narrow"] else 0)

    pdhl_ok = (ltp>pdh and direction=="CALL") or (ltp<pdl and direction=="PUT")
    cond["PDH/PDL"] = pdhl_ok;  score += 10 if pdhl_ok else 0

    if fib["at_fib"]:
        cond["Fib Pullback"] = True
        fib_label = f"0.{fib['level'].split('.')[1]} ({fib['price']:,.1f}) ✓ +{fib['score_bonus']}pts"
        score += fib["score_bonus"]
    else:
        cond["Fib Pullback"] = False
        fib_label = f"nearest {fib['level']} ({fib['price']:,.1f}) dist {fib['dist_pct']:.2f}%" if fib["level"] else "N/A"

    vix_v = state["vix"]
    if vix_v:
        if   vix_v > VIX_EXTREME:    score -= 30; cond["VIX"] = f"EXTREME {vix_v:.1f} ⛔"
        elif vix_v > VIX_NORMAL_MAX: score -= VIX_HIGH_PENALT; cond["VIX"] = f"HIGH {vix_v:.1f} ⚠"
        elif vix_v < VIX_CALM_MAX:   score += VIX_CALM_BONUS;  cond["VIX"] = f"CALM {vix_v:.1f} ✓"
        else:                          cond["VIX"] = f"NORMAL {vix_v:.1f}"

    pcr_v = state["pcr"]
    if pcr_v:
        if direction == "CALL":
            bonus = 5 if pcr_v>=PCR_VERY_BULL else (3 if pcr_v>=PCR_BULL else (-5 if pcr_v<=PCR_VERY_BEAR else 0))
        else:
            bonus = 5 if pcr_v<=PCR_VERY_BEAR else (3 if pcr_v<=PCR_BEAR else (-5 if pcr_v>=PCR_VERY_BULL else 0))
        score += bonus
        cond["PCR"] = f"{pcr_v:.2f}"

    # ── Order Flow ────────────────────────────────────────────────
    # 1. Volume Delta  (footprint approx — no extra API call)
    vd = calc_volume_delta(df_today)
    vd_ok = (vd["bias"] == "BULL" and direction == "CALL") or \
            (vd["bias"] == "BEAR" and direction == "PUT")
    if vd_ok: score += 5
    cond["Vol Delta"] = f"{vd['bias']} Δ{vd['delta_pct']:+.1f}% (30m:{vd['recent_delta']:+,})"

    # 2. Options chain flow  (NIFTY / BANKNIFTY only)
    _opt_map = {"NIFTY50": "NIFTY", "BANKNIFTY": "BANKNIFTY"}
    opt_sym  = _opt_map.get(name)
    of       = {}
    if opt_sym:
        of = fetch_options_flow(opt_sym, ltp)
        oi_ok  = (of.get("oi_change_bias") == "BULL" and direction == "CALL") or \
                 (of.get("oi_change_bias") == "BEAR" and direction == "PUT")
        atm_ok = (of.get("atm_bias") == "CALL" and direction == "CALL") or \
                 (of.get("atm_bias") == "PUT"  and direction == "PUT")
        score += (5 if oi_ok else 0) + (5 if atm_ok else 0)
        mp = of.get("max_pain")
        mp_str = f"{mp:,.0f}" if mp else "—"
        cond["Options Flow"] = (
            f"MaxPain {mp_str} | COI {of.get('oi_change_bias','—')} | "
            f"ATM {of.get('atm_bias','—')} | R {of.get('resistance','—')} / S {of.get('support','—')}"
        )

    # 3. FII / DII macro flow  (state refreshed hourly)
    fii_cash  = state.get("fii_net")
    fii_deriv = state.get("fii_deriv_net")
    dii_cash  = state.get("dii_net")
    fii_ok = (fii_cash is not None and fii_cash  > 0 and direction == "CALL") or \
             (fii_cash is not None and fii_cash  < 0 and direction == "PUT")
    if fii_ok: score += 4
    if fii_cash is not None:
        fd_str = f"{fii_deriv:+,.0f}" if fii_deriv is not None else "—"
        cond["FII/DII"] = (
            f"Cash FII {fii_cash:+,.0f} / DII {dii_cash:+,.0f} Cr | "
            f"Deriv FII {fd_str} Cr"
        )

    # 4. DOM bid/ask imbalance  (NSE futures depth)
    _dom_map = {"NIFTY50": "NIFTY", "BANKNIFTY": "BANKNIFTY"}
    dom_sym = _dom_map.get(name)
    dom = {}
    if dom_sym:
        dom = fetch_dom_imbalance(dom_sym)
        imb = dom.get("imbalance")
        dom_ok = (imb is not None and imb >  0.05 and direction == "CALL") or \
                 (imb is not None and imb < -0.05 and direction == "PUT")
        if dom_ok: score += 4
        if imb is not None:
            cond["DOM"] = (
                f"Imb {imb:+.2f} | "
                f"Bid {dom['total_bid_qty']:,} vs Ask {dom['total_ask_qty']:,} | "
                f"Best {dom['best_bid']}/{dom['best_ask']}"
            )

    score = max(0, min(100, score))

    now_ist = datetime.now(IST)
    late = now_ist.hour > LATE_SESSION[0] or (now_ist.hour == LATE_SESSION[0] and now_ist.minute >= LATE_SESSION[1])
    if late and score < SCORE_ULTRA:
        score = max(0, score - 10)
        cond["Session"] = "LATE ⚠ (-10)"

    conds_met = sum(1 for v in cond.values() if isinstance(v, bool) and v)

    if   score >= SCORE_ULTRA:  rating, rr = "ULTRA STRONG", "1:3"
    elif score >= SCORE_STRONG: rating, rr = "STRONG",       "1:3"
    elif score >= SCORE_MEDIUM: rating, rr = "MEDIUM",       "1:2"
    else:                       rating, rr = "WEAK",         "SKIP"

    sl_pts     = cfg["sl_pts"]
    target_pts = sl_pts * 3
    strike     = round(ltp / cfg["strike_step"]) * cfg["strike_step"]
    if direction == "CALL":
        sl_level, target_level = round(ltp-sl_pts,1), round(ltp+target_pts,1)
    else:
        sl_level, target_level = round(ltp+sl_pts,1), round(ltp-target_pts,1)

    return {
        "instrument": name, "ticker": cfg["ticker"],
        "direction": direction, "option_type": option_type,
        "ltp": ltp, "strike": strike, "sl_level": sl_level, "sl_pts": sl_pts,
        "target_level": target_level, "target_pts": target_pts,
        "score": score, "rating": rating, "rr": rr, "conds_met": conds_met,
        "conditions": cond, "e9": e9, "e21": e21, "vwap": vwap_, "rsi": rsi_,
        "st_dir": st_, "orb_high": orb_h, "orb_low": orb_l,
        "pdh": pdh, "pdl": pdl, "cpr": cpr, "fib": fib, "fib_label": fib_label,
        "time": now_ist.strftime("%H:%M"),
        # Order flow
        "vol_delta":    vd,
        "options_flow": of,
        "dom":          dom,
        # Footprint chart data (per-candle, for web dashboard)
        "footprint":    calc_candle_footprint(df_today),
    }

# ═══════════════════════════════════════════════════════════════
#  NOTIFICATIONS
# ═══════════════════════════════════════════════════════════════

TELEGRAM_ENABLED = True   # set to False at runtime if user skips Telegram

_PLACEHOLDER = {"YOUR_SCAN_BOT_TOKEN", "YOUR_SIGNAL_BOT_TOKEN",
                "YOUR_BOT_TOKEN_HERE", "", None}

def _tg_post(token, chat_id, msg):
    """Raw Telegram send. Returns True on success."""
    if not TELEGRAM_ENABLED:
        return False
    if token in _PLACEHOLDER or not chat_id or chat_id in _PLACEHOLDER:
        return False
    try:
        r = requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data={"chat_id": chat_id, "text": msg,
                  "parse_mode": "Markdown", "disable_web_page_preview": True},
            timeout=10)
        if r.status_code != 200:
            print(f"  [TG] Error {r.status_code}: {r.text[:120]}")
            return False
        return True
    except Exception as e:
        print(f"  [TG] {e}")
        return False

def send_scan_tg(msg: str):
    """Send to Bot 1 (scan summaries). Falls back to signal bot or legacy."""
    sent = _tg_post(SCAN_TG_TOKEN, SCAN_TG_CHAT, msg)
    if not sent:
        _tg_post(SIGNAL_TG_TOKEN, SIGNAL_TG_CHAT, msg) or \
        _tg_post(TELEGRAM_TOKEN, CHAT_ID, msg)

def send_signal_tg(msg: str):
    """Send to Bot 2 (confirmed signals, startup, EOD). Falls back to scan bot or legacy."""
    sent = _tg_post(SIGNAL_TG_TOKEN, SIGNAL_TG_CHAT, msg)
    if not sent:
        _tg_post(SCAN_TG_TOKEN, SCAN_TG_CHAT, msg) or \
        _tg_post(TELEGRAM_TOKEN, CHAT_ID, msg)

# Backward-compat alias used by any leftover direct calls
def send_telegram(msg: str):
    send_signal_tg(msg)

def _beep():
    try:
        if _OS == "Windows":
            import winsound; winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        elif _OS == "Darwin":
            subprocess.Popen(["afplay", "/System/Library/Sounds/Glass.aiff"],
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except: pass

def _popup(title, msg, urgent=False):
    import tkinter as tk
    bg, accent = "#1e1e2e", "#e05c5c" if urgent else "#1D9E75"
    def _show():
        root = tk.Tk(); root.overrideredirect(True)
        root.attributes("-topmost", True); root.configure(bg=bg)
        w, h = 430, 115
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        root.geometry(f"{w}x{h}+{sw-w-20}+{sh-h-60}")
        tk.Frame(root, bg=accent, height=4).pack(fill="x")
        tk.Label(root, text=title, font=("Segoe UI",11,"bold"),
                 fg="#fff", bg=bg, anchor="w", padx=12).pack(fill="x", pady=(6,0))
        tk.Label(root, text=msg, font=("Segoe UI",9), fg="#aaa", bg=bg,
                 anchor="w", padx=12, wraplength=410).pack(fill="x")
        tk.Button(root, text="✕", font=("Segoe UI",8), fg="#666", bg=bg,
                  bd=0, command=root.destroy).place(x=w-24, y=6)
        root.after(12000, root.destroy); root.mainloop()
    threading.Thread(target=_show, daemon=True).start()

def send_desktop(title, msg, urgent=False):
    _beep()
    if _OS == "Windows":
        try: _popup(title, msg, urgent); return
        except: pass
    print(f"\n{'='*55}\n  {title}\n  {msg}\n{'='*55}\n")

def startup_ping():
    now_str = datetime.now(IST).strftime("%d %b %Y %H:%M IST")
    send_signal_tg(
        f"🚀 *Nifty Precision Bot — Started*\n"
        f"⏰ {now_str}\n"
        f"🎯 Scanning: {', '.join(INSTRUMENTS)}\n"
        f"🔔 Signal threshold: score ≥ {SCORE_STRONG}\n"
        f"⏱ Every {SCAN_INTERVAL} min | 9:35 AM – 3:30 PM IST"
    )
    send_scan_tg(
        f"📡 *Scan bot active* | {now_str}\n"
        f"Every {SCAN_INTERVAL}-min scan summaries will appear here."
    )
    sb_update_status("running")

def heartbeat_ping():
    now_str = datetime.now(IST).strftime("%H:%M IST")
    recent  = state["scores_seen"][-6:]
    avg_sc  = round(sum(recent)/len(recent)) if recent else 0
    send_scan_tg(
        f"💓 *Bot alive* | {now_str}\n"
        f"🔍 Scans: {state['scan_count']}  |  Alerts: {state['alerts_sent']}\n"
        f"📊 Avg score (recent): {avg_sc}/100\n"
        f"📈 VIX: {state['vix'] or '—'}  |  PCR: {state['pcr'] or '—'}"
    )

def _data_error(name, reason):
    cnt = state["data_errors"].get(name, 0) + 1
    state["data_errors"][name] = cnt
    print(f"  [{name}] data error #{cnt}: {reason}")
    if cnt == 1 or cnt % 6 == 0:
        send_scan_tg(f"⚠️ *Data issue — {name}*\n{reason}\nAttempt #{cnt}")

# ═══════════════════════════════════════════════════════════════
#  EXCEL LOGGING
# ═══════════════════════════════════════════════════════════════

try:
    os.makedirs(EXCEL_FOLDER, exist_ok=True)
except Exception:
    EXCEL_FOLDER = "/tmp"
EXCEL_FILE = os.path.join(EXCEL_FOLDER, "Nifty_Precision_Bot.xlsx")

_HEADERS = [
    "Date","Time","Instrument","Direction","Strike","Option",
    "LTP","Score /100","Rating","R:R","Conds Met /8",
    "EMA9","EMA21","VWAP","RSI(14)","SuperTrend",
    "ORB High","ORB Low","PDH","PDL",
    "CPR Top","CPR Bottom","CPR Narrow?",
    "Fib Level","Fib Price","Fib Dist %","Fib At Level?",
    "SL Level","SL Pts","Target Level","Target Pts",
    "VIX","PCR",
    "Entry Premium","Exit Premium","P&L (Rs)","Result","Notes",
]

def _init_excel():
    if not _openpyxl or os.path.exists(EXCEL_FILE): return
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Signals"; ws.freeze_panes = "A2"
    for i, h in enumerate(_HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center")
    wb.save(EXCEL_FILE)

def log_excel(sig):
    if not _openpyxl: return
    try:
        _init_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Signals"]
        now = datetime.now(IST)
        cpr = sig["cpr"]
        ws.append([
            now.strftime("%Y-%m-%d"), now.strftime("%H:%M"),
            sig["instrument"], sig["direction"], sig["strike"], sig["option_type"],
            round(sig["ltp"],1), sig["score"], sig["rating"], sig["rr"], sig["conds_met"],
            round(sig["e9"],1), round(sig["e21"],1), round(sig["vwap"],1),
            round(sig["rsi"],1), "Bull" if sig["st_dir"]==1 else "Bear",
            round(sig["orb_high"],1), round(sig["orb_low"],1),
            round(sig["pdh"],1), round(sig["pdl"],1),
            round(cpr["top"],1), round(cpr["bottom"],1), "Yes" if cpr["narrow"] else "No",
            sig["fib"].get("level") or "", sig["fib"].get("price") or "",
            sig["fib"].get("dist_pct") or "", "Yes" if sig["fib"].get("at_fib") else "No",
            round(sig["sl_level"],1), sig["sl_pts"],
            round(sig["target_level"],1), sig["target_pts"],
            state.get("vix") or "", state.get("pcr") or "",
            "","","","","",
        ])
        last = ws.max_row
        ws.cell(row=last, column=4).font = Font(
            bold=True, color="276221" if sig["direction"]=="CALL" else "9C0006")
        wb.save(EXCEL_FILE)
        print(f"  [EXCEL] Saved → {EXCEL_FILE}")
    except Exception as e:
        print(f"  [EXCEL] {e}")

# ═══════════════════════════════════════════════════════════════
#  SIGNAL FORMAT
# ═══════════════════════════════════════════════════════════════

def format_signal(sig):
    d_icon = "▲ CALL" if sig["direction"] == "CALL" else "▼ PUT"
    icons  = {"ULTRA STRONG":"🔥","STRONG":"💪","MEDIUM":"⚡"}
    ri     = icons.get(sig["rating"],"")
    cpr    = sig["cpr"]
    clines = []
    for k, v in sig["conditions"].items():
        if   isinstance(v, bool): clines.append(f"  {'✅' if v else '❌'} {k}")
        else:                      clines.append(f"  ℹ️ {k}: {v}")

    # ── Order Flow section ──────────────────────────────────────
    vd  = sig.get("vol_delta", {})
    of  = sig.get("options_flow", {})
    dom = sig.get("dom", {})

    vd_bar = ""
    if vd.get("buy_vol", 0) + vd.get("sell_vol", 0) > 0:
        tot = vd["buy_vol"] + vd["sell_vol"]
        bull_pct = round(vd["buy_vol"] / tot * 10)
        bear_pct = 10 - bull_pct
        delta_icon = "🟢" if vd["delta"] > 0 else "🔴" if vd["delta"] < 0 else "⚪"
        vd_bar = (
            f"\n🔀 *Volume Delta (Footprint):*\n"
            f"  {'🟩'*bull_pct}{'🟥'*bear_pct}  {vd['delta_pct']:+.1f}% {delta_icon}\n"
            f"  Buy {vd['buy_vol']:,} | Sell {vd['sell_vol']:,} | 30m {vd['recent_delta']:+,}"
        )

    of_lines = ""
    if of.get("max_pain") or of.get("atm_bias"):
        mp = of.get("max_pain"); mp_s = f"{mp:,.0f}" if mp else "—"
        r  = of.get("resistance"); r_s = f"{r:,.0f}" if r else "—"
        s  = of.get("support");    s_s = f"{s:,.0f}" if s else "—"
        of_lines = (
            f"\n📐 *Options Flow:*\n"
            f"  MaxPain : {mp_s}\n"
            f"  COI Bias: {of.get('oi_change_bias','—')} | ATM bias: {of.get('atm_bias','—')}\n"
            f"  CE wall : {r_s} | PE wall : {s_s}"
        )

    dom_line = ""
    imb = dom.get("imbalance")
    if imb is not None:
        dom_icon = "🟢" if imb > 0.05 else "🔴" if imb < -0.05 else "⚪"
        dom_line = (
            f"\n📊 *DOM (Futures Depth):*\n"
            f"  Imbalance: {imb:+.2f} {dom_icon}\n"
            f"  Bid qty: {dom.get('total_bid_qty',0):,} | Ask qty: {dom.get('total_ask_qty',0):,}"
        )

    fii_c = sig["conditions"].get("FII/DII", "")
    fii_line = f"\n🏦 *FII/DII:* {fii_c}" if fii_c else ""

    order_flow_block = (vd_bar or of_lines or dom_line or fii_line) and (
        "\n" + vd_bar + of_lines + dom_line + fii_line
    )

    # ── 1-min Entry Confirmation block ──────────────────────────────
    ec = sig.get("entry_confirm")
    if ec and not ec.get("timed_out"):
        # Confirmed entry — show exact levels at top
        tsl   = ec.get("tight_sl")
        tspts = ec.get("tight_sl_pts")
        ttgt  = ec.get("tight_target")
        ep    = ec.get("entry_price", sig["ltp"])

        ec_block = f"\n\n{'═'*32}\n"
        ec_block += f"🚀 *ACTION: {sig['direction']} NOW*\n"
        ec_block += f"{'═'*32}\n"
        ec_block += f"  📥 Entry  : *{ep:,.1f}*\n"
        if tsl and tspts and ttgt:
            ec_block += f"  🛑 SL     : *{tsl:,.1f}*  (−{tspts} pts)\n"
            ec_block += f"  🎯 Target : *{ttgt:,.1f}*  (+{round(tspts*3,1)} pts)\n"
            ec_block += f"  ⚖️ R:R    : 1:3\n"
        else:
            ec_block += f"  🛑 SL     : *{sig['sl_level']:,.1f}*  (−{sig['sl_pts']} pts)\n"
            ec_block += f"  🎯 Target : *{sig['target_level']:,.1f}*  (+{sig['target_pts']} pts)\n"
            ec_block += f"  ⚖️ R:R    : 1:3\n"
        ec_block += f"{'─'*32}\n"
        ec_block += f"⏱ *1-Min Gate:* ✅ {ec['checks_met']}/5 checks passed\n"
        ec_block += f"  {ec['reason']}\n"
        vd1 = ec.get("vol_delta_1m", {})
        if vd1.get("buy_vol", 0) + vd1.get("sell_vol", 0) > 0:
            ec_block += (
                f"  📊 1m Δ: {vd1['bias']} {vd1['delta_pct']:+.1f}% "
                f"(B:{vd1['buy_vol']:,} / S:{vd1['sell_vol']:,})\n"
            )
    elif ec and ec.get("data_unavailable"):
        # yfinance 1-min feed broken — signal sent on 5-min basis
        ec_block = (
            f"\n\n⚠️ *1-Min Gate: Feed Unavailable*\n"
            f"  Signal sent on 5-min indicators only.\n"
            f"  Use 5-min SL: {sig['sl_level']:,.1f} (−{sig['sl_pts']} pts)\n"
        )
    elif ec and ec.get("timed_out"):
        # Should not reach format_signal on timeout+SKIP, but handle gracefully
        ec_block = (
            f"\n\n⛔ *1-Min Gate: TIMED OUT* ({ec['checks_met']}/5 checks)\n"
            f"  {ec['reason']}\n"
        )
    else:
        ec_block = ""   # MEDIUM signal — no 1-min gate

    return (
        f"{ri} *{sig['instrument']} — {d_icon} | {sig['rating']}*\n"
        f"📅 {datetime.now(IST).strftime('%d %b %Y %H:%M IST')}\n"
        # ── ACTION block first — visible without scrolling ──────────
        + ec_block +
        f"\n💰 *LTP:* {sig['ltp']:,.1f}\n"
        f"🎫 *Strike:* {sig['strike']} {sig['option_type']}\n"
        f"📊 *Score:* {sig['score']}/100 ({sig['conds_met']}/8 conditions)\n\n"
        f"📈 *Indicators:*\n"
        f"  EMA9/21 : {sig['e9']:,.0f} / {sig['e21']:,.0f}\n"
        f"  VWAP    : {sig['vwap']:,.1f}\n"
        f"  RSI(14) : {sig['rsi']:.1f}\n"
        f"  ST      : {'↑ Bullish' if sig['st_dir']==1 else '↓ Bearish'}\n"
        f"  ORB     : {sig['orb_high']:,.1f} H / {sig['orb_low']:,.1f} L\n"
        f"  PDH/PDL : {sig['pdh']:,.1f} / {sig['pdl']:,.1f}\n"
        f"  CPR     : {cpr['bottom']:,.1f}–{cpr['top']:,.1f} "
        f"({'Narrow 📈' if cpr['narrow'] else 'Wide 📉'})\n\n"
        f"🔢 *Fib:* {sig['fib_label']}\n"
        + (order_flow_block if order_flow_block else "") +
        f"\n\n📋 *Conditions:*\n" + "\n".join(clines) +
        f"\n\n⚠️ _Trade at your own risk._"
    )

# ═══════════════════════════════════════════════════════════════
#  SCAN SUMMARY (sent after every scan)
# ═══════════════════════════════════════════════════════════════

def send_scan_summary(scan_no, time_str, results):
    lines = []
    for name, sig, rating, score, alerted in results:
        if sig is None:
            lines.append(f"`{name:<12}` ❓ NO DATA")
            continue
        di   = "▲" if sig["direction"] == "CALL" else "▼"
        fill = round(score / 10)
        bar  = "█" * fill + "░" * (10 - fill)
        icon = "🔥" if score>=SCORE_ULTRA else "💪" if score>=SCORE_STRONG else "⚡" if score>=SCORE_MEDIUM else "·"
        tag  = " ✅ *SIGNAL!*" if alerted else ""
        # Volume delta footprint mini-bar
        vd = sig.get("vol_delta", {})
        if vd.get("buy_vol", 0) + vd.get("sell_vol", 0) > 0:
            tot = vd["buy_vol"] + vd["sell_vol"] or 1
            bp  = round(vd["buy_vol"] / tot * 5)   # 5-char mini bar
            vd_mini = f" Δ{'▲'*bp+'▼'*(5-bp)} {vd['delta_pct']:+.0f}%"
        else:
            vd_mini = ""
        lines.append(
            f"`{name:<12}` {di} {sig['direction']} {icon} "
            f"`{score:3d}/100` `[{bar}]` RSI {sig['rsi']:.0f} | {sig['ltp']:,.0f}"
            f"{vd_mini}{tag}"
        )

    vix_s = f"{state['vix']:.1f}" if state["vix"] else "—"
    pcr_s = f"{state['pcr']:.2f}" if state["pcr"] else "—"
    fii_s = f"{state['fii_net']:+,.0f}" if state.get("fii_net") is not None else "—"
    dii_s = f"{state['dii_net']:+,.0f}" if state.get("dii_net") is not None else "—"
    alerted_any = any(r[4] for r in results)
    hdr = f"🔍 *Scan #{scan_no}* | {time_str} | {'🚨 Signal!' if alerted_any else 'No signal'}"
    send_scan_tg(
        f"{hdr}\n\n" + "\n".join(lines) +
        f"\n\n📈 VIX `{vix_s}` | PCR `{pcr_s}` | "
        f"FII `{fii_s}` DII `{dii_s}` Cr | Alerts: {state['alerts_sent']}"
    )

# ═══════════════════════════════════════════════════════════════
#  MAIN SCAN LOOP
# ═══════════════════════════════════════════════════════════════

def in_market_hours():
    now  = datetime.now(IST)
    if now.weekday() >= 5: return False
    mins = now.hour * 60 + now.minute
    return (BOT_START[0]*60+BOT_START[1]) <= mins <= (MARKET_CLOSE[0]*60+MARKET_CLOSE[1])

def refresh_vix_pcr():
    last = state.get("last_vix_fetch")
    now  = datetime.now(IST)
    if last is None or (now - last).total_seconds() > 900:
        vix_v, pcr_v = fetch_vix_pcr()
        if vix_v is not None: state["vix"] = vix_v
        if pcr_v is not None: state["pcr"] = pcr_v
        state["last_vix_fetch"] = now
        print(f"  [VIX] {state['vix']}   [PCR] {state['pcr']}")
    refresh_fii_dii()   # FII/DII once-per-hour refresh

def scan():
    if not in_market_hours(): return
    if state["trades_today"] >= MAX_TRADES_DAY:
        print("[LIMIT] Max daily trades reached"); return

    state["scan_count"] += 1
    now_str = datetime.now(IST).strftime("%H:%M IST")
    print(f"\n{'─'*55}\n[SCAN #{state['scan_count']}] {now_str}")

    if state["scan_count"] % HEARTBEAT_SCANS == 0:
        heartbeat_ping()

    refresh_vix_pcr()

    if state["vix"] and state["vix"] > VIX_EXTREME:
        msg = f"⛔ VIX={state['vix']:.1f} EXTREME — signals blocked"
        print(f"  [BLOCK] {msg}")
        send_scan_tg(f"🔍 *Scan #{state['scan_count']}* | {now_str}\n\n{msg}")
        return

    results = []
    for name, cfg in INSTRUMENTS.items():
        try:
            sig = analyze(name, cfg)
            if sig is None:
                results.append((name, None, "NO DATA", 0, False)); continue

            state["scores_seen"].append(sig["score"])
            state["scores_seen"] = state["scores_seen"][-30:]

            print(f"  [{name}] {sig['direction']} Score={sig['score']} {sig['rating']} "
                  f"RSI={sig['rsi']:.1f} LTP={sig['ltp']:,.0f}")

            alerted = False
            below_thresh = sig["rating"] in ("WEAK","MEDIUM") and sig["score"] < cfg["min_score"]

            if not below_thresh:
                now = datetime.now(IST)
                last = state["last_signal"].get(name, {})
                suppress = (last and
                            (now-last["time"]).total_seconds()/60 < RESIGAL_GAP_MIN and
                            last["direction"] == sig["direction"] and
                            sig["score"] < last["score"] + 10)
                if not suppress:
                    # ── 1-min entry confirmation gate ─────────────────
                    if sig["score"] >= SCORE_STRONG:
                        ec = wait_for_1min_confirm(sig)
                        sig["entry_confirm"] = ec

                        if ec.get("data_unavailable"):
                            # yfinance 1-min feed broken → send on 5-min basis
                            print(f"      ⚠️ 1-min feed unavailable — sending on 5-min basis")
                            send_scan_tg(
                                f"⚠️ *{name}* {sig['direction']} Score {sig['score']} "
                                f"— 1-min feed unavailable. Sending on 5-min basis."
                            )
                            # fall through to fire signal below

                        elif ec["timed_out"] and SKIP_UNCONFIRMED:
                            # Checks ran but market didn't confirm — skip
                            send_scan_tg(
                                f"⛔ *{name}* {sig['direction']} Score {sig['score']} "
                                f"— 1-min never confirmed in {CONFIRM_TIMEOUT_MIN}m "
                                f"({ec['checks_met']}/5 checks). Signal skipped."
                            )
                            print(f"      ⛔ SKIPPED (no 1-min confirm)")
                            results.append((name, sig, sig["rating"], sig["score"], False))
                            continue   # don't fire trade alert

                    else:
                        sig["entry_confirm"] = None   # MEDIUM — no 1-min gate

                    # ── Signal confirmed (or MEDIUM) — fire trade alert ──
                    send_signal_tg(format_signal(sig))
                    log_excel(sig)
                    sb_push_signal(sig)
                    ec = sig.get("entry_confirm") or {}
                    tsl  = ec.get("tight_sl_pts")
                    tgt  = ec.get("tight_target")
                    desk_body = (
                        f"LTP {sig['ltp']:,.0f} | Score {sig['score']} | "
                        f"SL {ec.get('tight_sl', sig['sl_level']):,.0f} "
                        f"→ T {tgt if tgt else sig['target_level']:,.0f}"
                    )
                    send_desktop(f"{name} {sig['direction']} {sig['rating']}",
                                 desk_body,
                                 urgent=(sig["score"] >= SCORE_ULTRA))
                    state["last_signal"][name] = {
                        "direction": sig["direction"], "score": sig["score"], "time": now}
                    state["alerts_sent"]  += 1
                    state["trades_today"] += 1
                    alerted = True
                    print(f"      ✅ ALERTED")

            results.append((name, sig, sig["rating"], sig["score"], alerted))

        except Exception as e:
            import traceback; traceback.print_exc()
            results.append((name, None, "ERROR", 0, False))

    send_scan_summary(state["scan_count"], now_str, results)
    sb_push_scan(state["scan_count"], results)
    sb_update_status("running")

def eod_summary():
    now_str = datetime.now(IST).strftime("%d %b %Y")
    eod_msg = (
        f"📊 *Nifty Precision Bot — EOD {now_str}*\n\n"
        f"🔍 Scans      : {state['scan_count']}\n"
        f"📢 Signals    : {state['alerts_sent']}\n"
        f"📈 VIX (last) : {state['vix']}\n"
        f"📊 PCR (last) : {state['pcr']}\n"
    )
    send_signal_tg(eod_msg)
    send_scan_tg(eod_msg)
    send_desktop("Market Closed", f"Scans: {state['scan_count']} | Alerts: {state['alerts_sent']}")
    print(f"\n{'='*55}\nEOD | Scans: {state['scan_count']} | Alerts: {state['alerts_sent']}\n{'='*55}")
    sb_update_status("sleeping")


def _reset_daily_state():
    """Reset per-day counters when a new trading session begins."""
    state["scan_count"]     = 0
    state["alerts_sent"]    = 0
    state["trades_today"]   = 0
    state["last_signal"]    = {}
    state["scores_seen"]    = []
    state["vix"]            = None
    state["pcr"]            = None
    state["last_vix_fetch"] = None
    state["data_errors"]    = {}


def _next_market_open():
    """Return IST datetime of the next weekday 9:35 AM."""
    now  = datetime.now(IST)
    cand = now.replace(hour=BOT_START[0], minute=BOT_START[1], second=0, microsecond=0)
    if cand <= now:
        cand += timedelta(days=1)
    while cand.weekday() >= 5:          # skip Saturday / Sunday
        cand += timedelta(days=1)
    return cand

# ═══════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def main():
    print("=" * 55)
    print("  Nifty Precision Bot — LOCAL  (runs 24×5, auto-restart)")
    print(f"  Instruments : {', '.join(INSTRUMENTS)}")
    print(f"  Excel log   : {EXCEL_FILE}")
    print(f"  Alert score : ≥{SCORE_STRONG} (Strong) | ≥{SCORE_ULTRA} (Ultra)")
    print("=" * 55)

    # Show which bots are configured
    scan_ok   = SCAN_TG_TOKEN   not in _PLACEHOLDER
    signal_ok = SIGNAL_TG_TOKEN not in _PLACEHOLDER
    legacy_ok = TELEGRAM_TOKEN  not in _PLACEHOLDER
    print(f"  Bot 1 (scan summaries) : {'✅ configured' if scan_ok   else '⚠  not set'}")
    print(f"  Bot 2 (signals/alerts) : {'✅ configured' if signal_ok else '⚠  not set'}")
    if legacy_ok and not (scan_ok or signal_ok):
        print(f"  Legacy single bot      : ✅ configured (used for both)")
    print()

    _init_excel()

    # ── Supabase status check ─────────────────────────────────
    sb_ok = bool(SUPABASE_URL and SUPABASE_SERVICE_KEY
                 and "YOUR_PROJECT" not in SUPABASE_URL)
    if sb_ok:
        print(f"  📡 Supabase  : ✅ {SUPABASE_URL[:40]}…")
    else:
        print("  📡 Supabase  : ⚠️  NOT configured — dashboard will show no data")
        print("                 Set SUPABASE_URL and SUPABASE_SERVICE_KEY")
        print("                 in nifty_local_bot.py (lines ~39-40)")
    print()

    # ── Ask whether to use Telegram ──────────────────────────
    global TELEGRAM_ENABLED
    try:
        ans = input("Enable Telegram notifications? (y/n) [y]: ").strip().lower()
        TELEGRAM_ENABLED = (ans != 'n')
    except (EOFError, OSError):
        TELEGRAM_ENABLED = True   # non-interactive (systemd/cloud) → always on
    if TELEGRAM_ENABLED:
        print("  ✅ Telegram ON")
    else:
        print("  🔕 Telegram OFF — running silently")

    startup_ping()

    # Schedule recurring scans; EOD detection is done inside the while loop
    schedule.every(SCAN_INTERVAL).minutes.do(scan)

    # First scan immediately if already in market hours
    if in_market_hours():
        scan()

    eod_sent_date  = None   # date on which EOD summary was last sent
    active_date    = None   # date of the current / most recent trading session

    while True:
        try:
            schedule.run_pending()
            now    = datetime.now(IST)
            today  = now.date()
            in_mkt = in_market_hours()

            # ── New trading session started ──────────────────────
            if in_mkt and active_date != today:
                if active_date is not None:     # not the very first boot
                    print(f"\n🌅  New session: {today} — resetting daily state")
                    _reset_daily_state()
                    startup_ping()
                    scan()                      # immediate first scan of the day
                active_date = today

            # ── End of trading day ───────────────────────────────
            after_close = now.hour > 15 or (now.hour == 15 and now.minute >= 35)
            if after_close and now.weekday() < 5 and eod_sent_date != today:
                eod_summary()
                eod_sent_date = today
                nxt  = _next_market_open()
                diff = nxt - now
                hrs, leftover = divmod(int(diff.total_seconds()), 3600)
                mins = leftover // 60
                sleep_msg = (
                    f"💤 *Market Closed — Bot Sleeping*\n"
                    f"Next session: *{nxt.strftime('%a %d %b, %H:%M IST')}*\n"
                    f"Auto-resumes in {hrs}h {mins}m — no restart needed 🔄"
                )
                print(f"\n💤  Sleeping until {nxt.strftime('%a %d %b %H:%M IST')} "
                      f"({hrs}h {mins}m away)")
                send_signal_tg(sleep_msg)
                send_scan_tg(f"💤 *Scan bot sleeping*\nNext: {nxt.strftime('%a %d %b, %H:%M IST')}")
                send_desktop("Market Closed",
                             f"Next: {nxt.strftime('%a %d %b %H:%M')} | "
                             f"{hrs}h {mins}m")

            # Sleep longer when outside market hours (saves CPU)
            time.sleep(20 if in_mkt else 60)

        except KeyboardInterrupt:
            print("\n\n[STOP] Ctrl-C received — goodbye!")
            break
        except Exception as exc:
            print(f"\n[MAIN-LOOP ERROR] {exc}")
            time.sleep(30)          # brief pause, then keep running

    # ── Graceful shutdown: mark bot as stopped in Supabase ──────
    try:
        sb_update_status("stopped")
        print("  [SUPABASE] ✅ bot_status → stopped")
    except Exception:
        pass
    print("[STOP] Bot exited cleanly.")


if __name__ == "__main__":
    main()
