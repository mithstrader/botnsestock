#!/usr/bin/env python3
"""
NSE Options Smart Signal Bot
────────────────────────────
• Scans every 5 min during full market hours (9:15 AM – 3:30 PM IST)
• Detects "active movement periods" dynamically
• Alerts ONLY when a strong NEW signal appears — no spam
• Desktop pop-up notifications (Windows / Mac / Linux)
• Telegram alerts in parallel
• End-of-day summary at 3:35 PM IST

Setup guide at the bottom of this file.
"""

import os
import requests
import calendar
import time
import schedule
import platform
import subprocess
from datetime import datetime, date, timedelta
import pytz

# ── Excel logging ────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter
    _openpyxl = True
except ImportError:
    _openpyxl = False

# ── No browser cookies needed — expiry fetched from API ──────────

# ═══════════════════════════════════════════════════════════════════
#  ⚙️  CONFIGURATION — Fill these in before running
# ═══════════════════════════════════════════════════════════════════

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")   # Set in GitHub Secrets
CHAT_ID        = os.environ.get("CHAT_ID", "")           # Set in GitHub Secrets

# ─── Notification channels ─────────────────────────────────────────
ENABLE_TELEGRAM         = True    # Send to Telegram
# Desktop notifications are disabled when running in CI (GitHub Actions)
_CI = os.environ.get("CI", "false").lower() == "true"
ENABLE_DESKTOP_NOTIFY   = not _CI  # Pop-up on your PC / laptop
DESKTOP_NOTIFY_SOUND    = not _CI  # Play a beep with the notification

# ─── Your Trading Rules ────────────────────────────────────────────
MIN_DAY_CHANGE_LONG  =  1.5    # % min rise  → CALL signal
MIN_DAY_CHANGE_SHORT = -1.5    # % min fall  → PUT  signal
MIN_VOL_CHANGE       = -999.0  # Volume check removed as hard gate — used only in scoring
                               # Trendlyne volume % compares to FULL previous day,
                               # so morning scans always show negative values
MIN_OI_CHANGE        =   2.0   # % min OI build to qualify
MIN_SCORE_TO_ALERT   =  30.0   # Minimum score to trigger an alert
                               # Note: vol_change is always ~0 in morning scans (Trendlyne
                               # compares intraday volume to previous full day), so max morning
                               # score is 65/100 (day=35 + OI=30). 30 ≈ 46% of that cap,
                               # equivalent to the 40% bar used on full-day scores.
SCORE_JUMP_TO_ALERT  =  15.0   # Re-alert if score jumps by this much
MAX_RISK_PER_TRADE   =  1500   # Rs max risk per trade
DAILY_LOSS_LIMIT     =  3000   # Rs
MAX_TRADES_PER_DAY   =  10
MONTHLY_TARGET       = (8000, 12000)

# ─── SL / Target calculation (stock price levels) ─────────────────
SL_PCT     = 1.5   # % move against signal → stop loss (e.g. CALL: LTP × 0.985)
TARGET_PCT = 3.0   # % move with signal   → target    (e.g. CALL: LTP × 1.030)
                   # Risk : Reward = 1 : 2  (SL_PCT : TARGET_PCT)

# ─── Excel logging ────────────────────────────────────────────────
ENABLE_EXCEL_LOG   = True          # Save results to Excel
EXCEL_FOLDER       = os.environ.get("EXCEL_FOLDER", "./logs")  # Folder to save files (created auto)

# ─── Reversal detection ───────────────────────────────────────────
# Long Build Up with negative day change = smart money accumulation
# Short Build Up with positive day change = smart money distribution
ENABLE_REVERSAL_SIGNALS  = True
REVERSAL_MAX_NEG_CHANGE  = -2.0   # Only flag if day drop is within this (not a freefall)
REVERSAL_MIN_OI_CHANGE   =  5.0   # Higher OI threshold — needs strong conviction
REVERSAL_MIN_SCORE       = 50.0   # Higher score required than normal signals

# ─── Pre-market scan ──────────────────────────────────────────────
PREMARKET_SCAN_HOUR    = 9
PREMARKET_SCAN_MIN     = 10    # Scan at 9:10 AM, 5 min before open

# ─── Move exhaustion filter ────────────────────────────────────────
# If stock already moved a LOT, the signal is late — lower its score
EXHAUSTION_THRESHOLD   = 3.0   # If day_change > this%, flag as exhausted
EXHAUSTION_SCORE_PENALTY = 20  # Deduct this from score if exhausted
# If exhausted AND OI building = likely reversal/consolidation trap
EXHAUSTION_REVERSAL_OI = 4.0   # Min OI for exhausted stock to flip to REVERSAL

# ─── Price velocity tracking ──────────────────────────────────────
# Compare price 2 scans ago vs now to detect fast-moving stocks early
VELOCITY_THRESHOLD     = 0.4   # % price move per scan = fast momentum

# ─── Early detection ──────────────────────────────────────────────
# Layer 1: WATCHLIST — stocks building up before threshold
WATCHLIST_MIN_SCORE     = 25.0   # Start watching below alert threshold
WATCHLIST_ALERT_SCORE   = 38.0   # Alert when watchlist stock crosses this

# Layer 2: OI ACCELERATION — OI building faster scan-over-scan
OI_ACCEL_THRESHOLD      =  2.0   # OI grew by this % MORE than last scan

# Layer 3: EARLY SESSION — lower thresholds at market open
EARLY_SESSION_END_MINS  = 30     # First N minutes after open = early session
EARLY_DAY_CHANGE_LONG   =  0.8   # Lower threshold (vs 1.5 normally)
EARLY_DAY_CHANGE_SHORT  = -0.8
EARLY_OI_CHANGE         =  1.5   # Lower OI threshold

# ─── Scan settings ─────────────────────────────────────────────────
SCAN_INTERVAL_MIN           = 5       # Scan every N minutes
MARKET_OPEN_H,  MARKET_OPEN_M  = 9,  15   # 9:15 AM IST
MARKET_CLOSE_H, MARKET_CLOSE_M = 15, 30   # 3:30 PM IST
QUIET_SCORE_BOOST            = 20.0   # Extra score needed in quiet market

# ─── India VIX thresholds ──────────────────────────────────────────
VIX_CALM_MAX        = 12.0   # VIX < 12  → calm, ease score by 5
VIX_NORMAL_MAX      = 20.0   # VIX 12-20 → normal
VIX_HIGH_BOOST      = 15.0   # VIX > 20  → need 15 extra score
VIX_EXTREME         = 25.0   # VIX > 25  → block all intraday signals

# ─── Nifty PCR thresholds ─────────────────────────────────────────
PCR_VERY_BULL       = 1.5    # PCR > 1.5 → +15 bonus for CALL
PCR_BULL            = 1.2    # PCR > 1.2 → +8  bonus for CALL
PCR_BEAR            = 0.8    # PCR < 0.8 → +8  bonus for PUT
PCR_VERY_BEAR       = 0.6    # PCR < 0.6 → +15 bonus for PUT

# ═══════════════════════════════════════════════════════════════════
#  DESKTOP NOTIFICATION ENGINE  (auto-detects your OS)
# ═══════════════════════════════════════════════════════════════════

_OS = platform.system()   # 'Windows', 'Darwin' (Mac), 'Linux'

# Try plyer first — works on all platforms if installed
_plyer = None
try:
    from plyer import notification as _plyer
except ImportError:
    pass

# Determine which backend is active (for status display)
if _plyer:
    _notifier = 'plyer'
elif _OS == 'Darwin':
    _notifier = 'macos (osascript)'
elif _OS == 'Linux':
    _notifier = 'linux (notify-send)'
elif _OS == 'Windows':
    _notifier = 'windows (powershell)'
else:
    _notifier = 'terminal fallback'


def _beep():
    """Play a short alert sound."""
    if not DESKTOP_NOTIFY_SOUND:
        return
    try:
        if _OS == 'Windows':
            import winsound
            winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        elif _OS == 'Darwin':
            subprocess.Popen(['afplay', '/System/Library/Sounds/Glass.aiff'],
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        else:
            subprocess.Popen(['paplay', '/usr/share/sounds/freedesktop/stereo/bell.oga'],
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass


def _tkinter_popup(title, message, urgent=False, duration=10):
    """
    Custom dark-themed pop-up window — bottom-right corner, auto-closes.
    Works on Windows without any extra install.
    """
    import tkinter as tk
    import threading

    bg      = "#1e1e2e"
    fg_head = "#ffffff"
    fg_body = "#aaaaaa"
    accent  = "#e05c5c" if urgent else "#1D9E75"

    def _show():
        root = tk.Tk()
        root.overrideredirect(True)          # No title bar
        root.attributes("-topmost", True)
        root.configure(bg=bg)

        # Size & position — bottom-right corner
        w, h   = 400, 110
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        root.geometry(f"{w}x{h}+{sw - w - 20}+{sh - h - 60}")

        # Coloured top bar (accent stripe)
        tk.Frame(root, bg=accent, height=4).pack(fill="x")

        # Title
        tk.Label(root, text=title, font=("Segoe UI", 11, "bold"),
                 fg=fg_head, bg=bg, anchor="w", padx=12).pack(fill="x", pady=(8, 0))

        # Body
        tk.Label(root, text=message, font=("Segoe UI", 9),
                 fg=fg_body, bg=bg, anchor="w", padx=12, wraplength=380,
                 justify="left").pack(fill="x")

        # Close button
        tk.Button(root, text="✕", font=("Segoe UI", 8), fg="#666", bg=bg,
                  bd=0, activebackground=bg, cursor="hand2",
                  command=root.destroy).place(x=w-24, y=6)

        root.after(duration * 1000, root.destroy)
        root.mainloop()

    threading.Thread(target=_show, daemon=True).start()


def send_desktop(title, message, urgent=False):
    """Send desktop pop-up. Tries tkinter → plyer → OS-native."""
    if not ENABLE_DESKTOP_NOTIFY:
        return

    _beep()

    # ── tkinter (primary on Windows — confirmed working) ──────────
    if _OS == 'Windows':
        try:
            _tkinter_popup(title, message, urgent=urgent)
            return
        except Exception as e:
            print(f"[NOTIFY] tkinter failed: {e}")

    # ── plyer (cross-platform if installed) ───────────────────────
    if _plyer:
        try:
            _plyer.notify(title=title, message=message,
                          app_name="NSE Options Bot", timeout=10)
            return
        except Exception:
            pass

    # ── macOS native ──────────────────────────────────────────────
    if _OS == 'Darwin':
        try:
            safe = message.replace('"', "'")
            stitle = title.replace('"', "'")
            subprocess.Popen(
                ['osascript', '-e',
                 f'display notification "{safe}" with title "{stitle}" sound name "Glass"'],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            return
        except Exception:
            pass

    # ── Linux native ──────────────────────────────────────────────
    if _OS == 'Linux':
        try:
            urgency = 'critical' if urgent else 'normal'
            subprocess.Popen(
                ['notify-send', '-u', urgency, '-t', '10000', title, message],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            return
        except Exception:
            pass

    # ── Terminal fallback ─────────────────────────────────────────
    print(f"\n{'='*50}\n  🔔 {title}\n  {message}\n{'='*50}\n")


def notify_signal(picks, activity, session):
    """Format and send desktop notification for trade signals."""
    if not picks:
        return

    for s in picks:
        sym   = s['symbol']
        d     = s['direction']
        ltp   = s['ltp']
        score = s['score']
        dc    = s['day_change']
        sign  = "+" if dc > 0 else ""
        tag   = "NEW" if s['is_new'] else f"STRONGER (+{s['score_jump']:.0f}pts)"
        icon  = "▲ CALL" if d == 'CALL' else "▼ PUT"

        title = f"NSE Signal: {sym} {icon}  [{tag}]"
        body  = (
            f"LTP: Rs {ltp:,.1f}  |  Day: {sign}{dc:.1f}%\n"
            f"Score: {score}/100  |  Activity: {activity:.0f}/100\n"
            f"Session: {session}\n"
            f"SL: Rs {MAX_RISK_PER_TRADE:,}  |  Target: Rs {MAX_RISK_PER_TRADE*2:,}"
        )
        send_desktop(title, body, urgent=(score >= 70))


def notify_daily_cap():
    send_desktop(
        "NSE Bot — Daily Limit Reached",
        f"All {MAX_TRADES_PER_DAY} trade slots used today. No more alerts.",
        urgent=False
    )


def notify_eod():
    send_desktop(
        "NSE Bot — Market Closed",
        f"Day complete. Scans: {state['scan_count']} | Alerts: {state['alerts_sent']}",
        urgent=False
    )



# ═══════════════════════════════════════════════════════════════════
#  EXCEL LOGGING
# ═══════════════════════════════════════════════════════════════════

import os

# ── Single persistent file ───────────────────────────────────────
EXCEL_FILE = os.path.join(EXCEL_FOLDER, "NSE_Bot_Master.xlsx")

# Column headers (Date column added as first column)
_SIG_HEADERS = [
    "Date", "Time", "Symbol", "Direction", "LTP (Rs)",
    "Day Chg %", "Vol Surge %", "OI Build %", "Score /100",
    "Buildup", "Signal Type", "Tag", "Suggested Strike",
    "SL Price",        # col 14 — stock price at stop loss  (LTP ± SL_PCT%)
    "Target Price",    # col 15 — stock price at target     (LTP ± TARGET_PCT%)
    "Session", "Market Activity",
    "Entry Premium", "Exit Premium", "P&L (Rs)",
    "Trade Outcome",   # col 21 — TP / SL / SKIP / PENDING
    "Notes",
    "India VIX", "Nifty PCR"
]
_SCAN_HEADERS = [
    "Date", "Time", "Scan #", "Total Stocks",
    "Activity /100", "Session", "Signals Sent", "Stocks Alerted"
]

_C = {
    "call_fg"  : "276221",
    "put_fg"   : "9C0006",
    "header_bg": "1F3864",
    "header_fg": "FFFFFF",
    "new_bg"   : "FFEB9C",   # yellow  — new signal
    "strong_bg": "DDEBF7",   # blue    — stronger
    "alt_bg"   : "F2F2F2",
    "border"   : "BFBFBF",
    "eod_bg"   : "E2EFDA",   # light green — EOD divider
}

def _thin_border():
    s = Side(style="thin", color=_C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def _header_style(cell, text):
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, color=_C["header_fg"], size=10)
    cell.fill      = PatternFill("solid", fgColor=_C["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()

def _apply_cell(cell, value, bold=False, fg=None, bg=None, center=True):
    cell.value     = value
    cell.font      = Font(name="Arial", size=10, bold=bold,
                          color=fg if fg else "000000")
    cell.fill      = PatternFill("solid", fgColor=bg) if bg else PatternFill()
    cell.border    = _thin_border()
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center")

def _load_or_create_wb():
    """Load the single master workbook, creating it fresh if missing."""
    os.makedirs(EXCEL_FOLDER, exist_ok=True)

    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Signals" in wb.sheetnames:
            ws_sig = wb["Signals"]
            last_col = ws_sig.max_column
            headers_present = [ws_sig.cell(1, c).value for c in range(1, last_col + 1)]
            migrated = False

            # ── Migrate: rename SL/Target/Result columns ──────────
            _renames = {
                "SL (Rs)"     : "SL Price",
                "Target (Rs)" : "Target Price",
                "Result"      : "Trade Outcome",
            }
            for col_idx in range(1, last_col + 1):
                old_hdr = ws_sig.cell(1, col_idx).value
                if old_hdr in _renames:
                    ws_sig.cell(1, col_idx).value = _renames[old_hdr]
                    migrated = True
                    print(f"[EXCEL] Migrated: '{old_hdr}' → '{_renames[old_hdr]}'")

            # ── Migrate: add VIX / PCR header columns if not present ──
            if "India VIX" not in headers_present:
                for col, title in [(last_col + 1, "India VIX"),
                                   (last_col + 2, "Nifty PCR")]:
                    _header_style(ws_sig.cell(1, col), title)
                    ws_sig.column_dimensions[get_column_letter(col)].width = 10
                migrated = True
                print("[EXCEL] Migrated: added India VIX + Nifty PCR columns.")

            if migrated:
                wb.save(EXCEL_FILE)
        return wb

    wb  = openpyxl.Workbook()

    # ── Sheet 1: Signals ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Signals"
    ws1.row_dimensions[1].height = 32
    widths = [12, 10, 14, 10, 10, 10, 12, 10, 10, 16, 12, 14, 18, 10, 12, 18, 14, 14, 12, 10, 20, 10, 10]
    for col, (h, w) in enumerate(zip(_SIG_HEADERS, widths), 1):
        _header_style(ws1.cell(row=1, column=col), h)
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(_SIG_HEADERS))}1"

    # ── Migrate existing workbooks: add VIX/PCR headers if missing ──

    # ── Sheet 2: All Scans ────────────────────────────────────────
    ws2 = wb.create_sheet("All Scans")
    ws2.row_dimensions[1].height = 32
    widths2 = [12, 10, 8, 12, 12, 18, 12, 35]
    for col, (h, w) in enumerate(zip(_SCAN_HEADERS, widths2), 1):
        _header_style(ws2.cell(row=1, column=col), h)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(_SCAN_HEADERS))}1"

    # ── Sheet 3: Summary ──────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    summary_headers = ["Date", "Total Scans", "Stocks w/ Signal",
                       "CALLs", "PUTs", "Avg Score", "Best Score", "Best Stock"]
    widths3 = [12, 12, 16, 8, 8, 10, 10, 14]
    ws3.row_dimensions[1].height = 32
    for col, (h, w) in enumerate(zip(summary_headers, widths3), 1):
        _header_style(ws3.cell(row=1, column=col), h)
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = "A2"

    wb.save(EXCEL_FILE)
    print(f"[EXCEL] Master file created → {EXCEL_FILE}")
    return wb


def excel_log_signals(picks, activity, session, expiry):
    """Append qualified signals to the Signals sheet."""
    if not ENABLE_EXCEL_LOG or not _openpyxl or not picks:
        return
    try:
        wb  = _load_or_create_wb()
        ws  = wb["Signals"]
        IST = pytz.timezone("Asia/Kolkata")
        now = datetime.now(IST)
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")

        for s in picks:
            d      = s["direction"]
            is_new = s.get("is_new", True)
            jump   = s.get("score_jump", 0) or 0
            tag    = "NEW" if is_new else f"+{jump:.0f} pts"
            bg     = _C["new_bg"] if is_new else _C["strong_bg"]
            fg     = _C["call_fg"] if d == "CALL" else _C["put_fg"]

            ltp = s["ltp"]
            # Calculate actual stock price levels for SL and Target
            if d == "CALL":
                sl_price     = round(ltp * (1 - SL_PCT     / 100), 2)
                target_price = round(ltp * (1 + TARGET_PCT  / 100), 2)
            else:   # PUT
                sl_price     = round(ltp * (1 + SL_PCT     / 100), 2)
                target_price = round(ltp * (1 - TARGET_PCT  / 100), 2)

            row_vals = [
                date_str, time_str, s["symbol"], d,
                ltp,
                round(s["day_change"], 2),
                round(s["vol_change"], 1),
                round(s["oi_change"],  2),
                s["score"], s.get("buildup",""), s.get("signal_type","NORMAL"), tag,
                f"ATM or 1 OTM {d}",
                sl_price,     # actual stop-loss price level
                target_price, # actual target price level
                session, round(activity, 1),
                "", "", "", "PENDING", "",
                _market_context.get("vix") or "",
                _market_context.get("pcr") or "",
            ]

            r = ws.max_row + 1
            for col, val in enumerate(row_vals, 1):
                _apply_cell(ws.cell(r, col), val,
                            bold=(col <= 4), fg=fg, bg=bg)

        wb.save(EXCEL_FILE)
        print(f"[EXCEL] Signal logged → {EXCEL_FILE}")
    except Exception as e:
        print(f"[EXCEL ERROR] {e}")


def excel_log_scan(scan_num, stock_count, activity, session, signals_sent, alerted):
    """Append every scan to the All Scans sheet."""
    if not ENABLE_EXCEL_LOG or not _openpyxl:
        return
    try:
        wb  = _load_or_create_wb()
        ws  = wb["All Scans"]
        IST = pytz.timezone("Asia/Kolkata")
        now = datetime.now(IST)
        r   = ws.max_row + 1
        alt = (r % 2 == 0)
        bg  = _C["alt_bg"] if alt else "FFFFFF"

        row_vals = [
            now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"),
            scan_num, stock_count,
            round(activity, 1), session,
            signals_sent, ", ".join(alerted) or "—"
        ]
        for col, val in enumerate(row_vals, 1):
            _apply_cell(ws.cell(r, col), val, bg=bg)

        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[EXCEL SCAN ERROR] {e}")


def excel_log_eod():
    """Write EOD divider + daily summary row to both sheets."""
    if not ENABLE_EXCEL_LOG or not _openpyxl:
        return
    try:
        wb  = _load_or_create_wb()
        IST = pytz.timezone("Asia/Kolkata")
        now = datetime.now(IST)
        date_str = now.strftime("%Y-%m-%d")

        # ── Signals sheet: divider row ────────────────────────────
        ws1 = wb["Signals"]
        r1  = ws1.max_row + 1
        ncols = len(_SIG_HEADERS)
        ws1.merge_cells(start_row=r1, start_column=1,
                        end_row=r1, end_column=ncols)
        cell = ws1.cell(r1, 1)
        sent = state["alerts_sent"]
        cell.value     = f"  ── {date_str}  End of Day  |  Signals: {sent}  |  Stocks: {', '.join(state['alerted_today']) or 'None'}  ──"
        cell.font      = Font(name="Arial", bold=True, italic=True, size=10, color="276221")
        cell.fill      = PatternFill("solid", fgColor=_C["eod_bg"])
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # ── Summary sheet: one row per day ────────────────────────
        ws3    = wb["Summary"]
        ws_sig = wb["Signals"]
        calls  = sum(1 for row in ws_sig.iter_rows(min_row=2, values_only=True)
                     if row[0] == date_str and row[3] == "CALL")
        puts   = sum(1 for row in ws_sig.iter_rows(min_row=2, values_only=True)
                     if row[0] == date_str and row[3] == "PUT")
        scores = [row[8] for row in ws_sig.iter_rows(min_row=2, values_only=True)
                  if row[0] == date_str and isinstance(row[8], (int, float))]
        avg_sc  = round(sum(scores)/len(scores), 1) if scores else 0
        best_sc = max(scores) if scores else 0
        best_st = next((row[2] for row in ws_sig.iter_rows(min_row=2, values_only=True)
                        if row[0] == date_str and row[8] == best_sc), "—")

        r3 = ws3.max_row + 1
        day_vals = [date_str, state["scan_count"], sent,
                    calls, puts, avg_sc, best_sc, best_st]
        for col, val in enumerate(day_vals, 1):
            _apply_cell(ws3.cell(r3, col), val, bold=(col == 1))

        wb.save(EXCEL_FILE)
        print(f"[EXCEL] EOD summary written → {EXCEL_FILE}")
    except Exception as e:
        print(f"[EXCEL EOD ERROR] {e}")


# ═══════════════════════════════════════════════════════════════════
#  STATE  (resets every trading day)
# ═══════════════════════════════════════════════════════════════════

state = {
    'alerted_today'   : set(),
    'watchlist'       : {},      # symbol -> {score, first_seen, scans}
    'prev_oi'         : {},      # symbol -> oi_change from last scan
    'prev_price'      : {},      # symbol -> ltp from last scan
    'prev2_price'     : {},      # symbol -> ltp from 2 scans ago
    'last_scores'     : {},
    'activity_history': [],
    'alerts_sent'     : 0,
    'scan_count'      : 0,
    'date'            : None,
}


# ── Market context (VIX + PCR) — refreshed every 15 min ─────────
_market_context = {
    "vix":         None,   # India VIX float
    "pcr":         None,   # Nifty PCR float
    "last_refresh": None,  # datetime of last fetch
}


def reset_daily_state():
    state['alerted_today']    = set()
    state['last_scores']      = {}
    state['activity_history'] = []
    state['alerts_sent']      = 0
    state['scan_count']       = 0
    state['prev_oi']          = {}
    state['prev_price']       = {}
    state['prev2_price']      = {}
    state['watchlist']        = {}
    state['date']             = ist_now().date()
    print("[RESET] New trading day — state cleared.")
    send_telegram(
            f"🛑 *Bot Started*\n"
            f"Wait For Signals.\n"
            f"....... 🙏....."
        )



# ═══════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════

def ist_now():
    return datetime.now(pytz.timezone('Asia/Kolkata'))


def is_market_open():
    now = ist_now()
    if now.weekday() >= 5:
        return False
    mins      = now.hour * 60 + now.minute
    open_min  = MARKET_OPEN_H  * 60 + MARKET_OPEN_M
    close_min = MARKET_CLOSE_H * 60 + MARKET_CLOSE_M
    return open_min <= mins <= close_min


def is_early_session():
    """True during first 30 minutes after market open."""
    now  = ist_now()
    open_time = now.replace(hour=MARKET_OPEN_H, minute=MARKET_OPEN_M, second=0)
    mins_since_open = (now - open_time).total_seconds() / 60
    return 0 <= mins_since_open <= EARLY_SESSION_END_MINS


def get_oi_acceleration(symbol, current_oic):
    """How much faster is OI building vs last scan? Returns acceleration."""
    prev = state['prev_oi'].get(symbol, 0)
    return round(current_oic - prev, 2)


def classify_session():
    mins = ist_now().hour * 60 + ist_now().minute
    if   mins < 9*60+45:   return "Opening Volatility"
    elif mins < 11*60:     return "Morning Momentum"
    elif mins < 13*60:     return "Midday Lull"
    elif mins < 14*60+30:  return "Afternoon Pickup"
    else:                  return "Power Hour"


def get_near_expiry():
    today = ist_now().date()

    def last_thursday(year, month):
        last_day = calendar.monthrange(year, month)[1]
        for day in range(last_day, 0, -1):
            if datetime(year, month, day).weekday() == 3:
                return datetime(year, month, day).date()

    expiry = last_thursday(today.year, today.month)
    if today > expiry:
        nm     = today.month + 1 if today.month < 12 else 1
        ny     = today.year  + 1 if today.month == 12 else today.year
        expiry = last_thursday(ny, nm)
    return expiry.strftime('%Y-%m-%d')


# ═══════════════════════════════════════════════════════════════════
#  DATA
# ═══════════════════════════════════════════════════════════════════

# ── Shared session with browser-like headers ─────────────────────
_session = requests.Session()
_session.headers.update({
    "User-Agent"     : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept"         : "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer"        : "https://smartoptions.trendlyne.com/",
    "Origin"         : "https://smartoptions.trendlyne.com",
})

# Cached expiry (refreshed daily)
_expiry_cache = {"expiry": None, "day": None}

EXPIRY_URL = "https://smartoptions.trendlyne.com/phoenix/api/fno/get-expiry-dates/?mtype=futures"

# All screener types to combine — covers all 4 buildup patterns
SCREENER_TYPES = ["oi-gainers", "short-covering", "long-unwinding"]


def _expiry_calendar_fallback():
    """Calculate near expiry as last Thursday of current (or next) month."""
    today = ist_now().date()
    def _last_thu(year, month):
        last_day = calendar.monthrange(year, month)[1]
        for d in range(last_day, 0, -1):
            if date(year, month, d).weekday() == 3:   # Thursday
                return date(year, month, d)
    expiry = _last_thu(today.year, today.month)
    if today > expiry:
        nm = today.month + 1 if today.month < 12 else 1
        ny = today.year + 1 if today.month == 12 else today.year
        expiry = _last_thu(ny, nm)
    return expiry.strftime('%Y-%m-%d')


def _warm_trendlyne_session():
    """
    Visit Trendlyne pages to get session cookies, then extract the Django
    CSRF token and set it as X-CSRFToken header for all subsequent API calls.
    Without this, the API returns an auth-error JSON (head = string) instead
    of real data.
    """
    pages = [
        "https://smartoptions.trendlyne.com/",
        "https://smartoptions.trendlyne.com/fno-market-filter/",
        "https://smartoptions.trendlyne.com/oi-gainers/oi-gainers/",
    ]
    try:
        for url in pages:
            _session.get(url, timeout=10)

        # Django requires csrftoken cookie to be echoed back as X-CSRFToken header
        csrf = _session.cookies.get("csrftoken", "")
        if csrf:
            _session.headers.update({"X-CSRFToken": csrf})
            print(f"[TRENDLYNE] Session warmed up. CSRF: {csrf[:10]}...")
        else:
            print("[TRENDLYNE] Session warmed up (no CSRF cookie found).")
    except Exception as e:
        print(f"[TRENDLYNE] Warm-up failed (continuing): {e}")


def get_near_expiry():
    """
    Fetch near expiry from Trendlyne API, with calendar fallback.
    Caches the result for the day to avoid repeated calls.
    """
    IST   = pytz.timezone("Asia/Kolkata")
    today = datetime.now(IST).date()

    if _expiry_cache["day"] == today and _expiry_cache["expiry"]:
        return _expiry_cache["expiry"]

    # ── Try Trendlyne API ────────────────────────────────────────
    try:
        r = _session.get(EXPIRY_URL, timeout=10)
        r.raise_for_status()
        data = r.json()
        # Safely navigate — API can return unexpected types when session is cold
        if isinstance(data, dict):
            body = data.get("body")
            if isinstance(body, dict):
                expiry = body.get("default_expiry_date")
                if expiry and isinstance(expiry, str):
                    _expiry_cache.update({"expiry": expiry, "day": today})
                    print(f"[EXPIRY] {expiry}  (from API, available: {body.get('expiryDates', [])})")
                    return expiry
        print(f"[EXPIRY] Unexpected response format ({type(data).__name__}) — using calendar fallback")
    except Exception as e:
        print(f"[EXPIRY] API error ({e}) — using calendar fallback")

    # ── Calendar fallback ────────────────────────────────────────
    expiry = _expiry_calendar_fallback()
    _expiry_cache.update({"expiry": expiry, "day": today})
    print(f"[EXPIRY] Fallback (calendar): {expiry}")
    return expiry


def _try_fetch_oi(expiry):
    """Single attempt to fetch OI gainers. Returns data dict or None."""
    url = (
        "https://smartoptions.trendlyne.com/phoenix/api/fno/market/filter/"
        f"?mtype=futures&expDate={expiry}&screenType=oi-gainers"
    )
    r = _session.get(url, timeout=20)
    r.raise_for_status()
    d = r.json()
    if not isinstance(d, dict):
        print(f"[FETCH] Unexpected type ({type(d).__name__}) — session cold")
        return None
    head   = d.get("head")
    status = head.get("status") if isinstance(head, dict) else None
    if str(status) in ("0", "0.0") or status == 0:
        rows = len((d.get("body") or {}).get("tableData") or [])
        print(f"[FETCH] OK — {rows} stocks for expiry {expiry}")
        return d
    print(f"[FETCH] status={status!r} head={type(head).__name__} — auth/CSRF problem")
    return None


def fetch_oi_gainers(expiry_override=None):
    """Fetch OI gainers — auto-retries with re-warm on first failure."""
    expiry = expiry_override or get_near_expiry()
    if not expiry:
        print("[FETCH] No valid expiry date available.")
        return None

    try:
        data = _try_fetch_oi(expiry)
        if data is not None:
            return data
    except Exception as e:
        print(f"[FETCH] Attempt 1 failed: {e}")

    # ── Retry: re-warm Trendlyne session and try again ────────────
    print("[FETCH] Re-warming Trendlyne session and retrying...")
    _warm_trendlyne_session()
    try:
        return _try_fetch_oi(expiry)
    except Exception as e:
        print(f"[FETCH ERROR] Attempt 2 failed: {e}")
        return None


def parse_rows(data):
    stocks = []
    for row in data['body']['tableData']:
        try:
            stocks.append({
                'symbol'    : row[0]['name'],
                'lot_size'  : row[0]['callbackinfo']['lotSize'],
                'ltp'       : float(row[1]),
                'day_change': float(row[2]),
                'vol_change': float(row[4]),
                'oi_change' : float(row[7]),
                'buildup'   : str(row[11]),
            })
        except Exception:
            continue
    return stocks


# ═══════════════════════════════════════════════════════════════════
#  SCORING & ACTIVITY DETECTION
# ═══════════════════════════════════════════════════════════════════

def compute_score(s):
    dc  = abs(s['day_change'])
    vc  = max(s['vol_change'], 0)
    oic = s['oi_change']
    score  = min(dc  * 7,  35)
    score += min(vc  / 10, 35)
    score += min(oic * 3,  30)
    prev = state['last_scores'].get(s['symbol'], 0)
    if prev >= 30 and score >= 30:
        score = min(score + 10, 100)
    return round(score, 1)


def market_activity_score(stocks):
    if not stocks:
        return 0
    strong  = sum(1 for s in stocks if abs(s['day_change']) >= 1.5 and s['vol_change'] >= 20)
    avg_oic = sum(s['oi_change'] for s in stocks) / len(stocks)
    avg_vc  = sum(max(s['vol_change'], 0) for s in stocks) / len(stocks)
    act  = min(strong * 5,   40)
    act += min(avg_oic * 4,  30)
    act += min(avg_vc  / 10, 30)
    return round(act, 1)


def is_active_market():
    recent = state['activity_history'][-4:]
    avg    = sum(recent) / len(recent) if recent else 0
    return avg >= 35


# ═══════════════════════════════════════════════════════════════════
#  SIGNAL FILTERING
# ═══════════════════════════════════════════════════════════════════

def filter_signals(stocks):
    active    = is_active_market()
    threshold = MIN_SCORE_TO_ALERT if active else MIN_SCORE_TO_ALERT + QUIET_SCORE_BOOST
    remaining = MAX_TRADES_PER_DAY - state['alerts_sent']
    if remaining <= 0:
        return []

    # ── VIX override ─────────────────────────────────────────────
    vix_delta, vix_block = get_vix_threshold_delta()
    if vix_block:
        vix = _market_context.get("vix", 0)
        print(f"[VIX] India VIX {vix:.1f} > {VIX_EXTREME} — all signals blocked.")
        send_telegram(
            f"⚠️ *VIX ALERT*\n"
            f"India VIX is extremely high at *{vix:.1f}*\n"
            f"All intraday signals blocked for safety.\n"
            f"_Options premiums are inflated — risk is very high._"
        )
        return []
    threshold += vix_delta

    calls, puts = [], []

    for s in stocks:
        dc  = s['day_change']
        vc  = s['vol_change']
        oic = s['oi_change']
        b   = s['buildup']
        sym = s['symbol']

        if oic < MIN_OI_CHANGE or vc < MIN_VOL_CHANGE:
            continue

        if b == 'Long Build Up' and dc >= MIN_DAY_CHANGE_LONG:
            s['direction'] = 'CALL'
        elif b == 'Short Build Up' and dc <= MIN_DAY_CHANGE_SHORT:
            s['direction'] = 'PUT'
        else:
            continue

        s['score']      = compute_score(s)

        # ── PCR bonus: confirm or penalise based on market sentiment
        pcr_bonus       = get_pcr_bonus(s['direction'])
        s['score']      = min(round(s['score'] + pcr_bonus, 1), 100)
        s['pcr_bonus']  = pcr_bonus

        prev_score      = state['last_scores'].get(sym, 0)
        s['is_new']     = sym not in state['alerted_today']
        s['score_jump'] = round(s['score'] - prev_score, 1)

        if s['score'] < threshold:
            continue
        if not s['is_new'] and s['score_jump'] < SCORE_JUMP_TO_ALERT:
            continue

        (calls if s['direction'] == 'CALL' else puts).append(s)

    calls.sort(key=lambda x: x['score'], reverse=True)
    puts.sort(key=lambda x: x['score'],  reverse=True)

    picks = []
    if calls:
        picks.append(calls[0])
    if puts and len(picks) < remaining:
        picks.append(puts[0])
    elif len(calls) > 1 and len(picks) < remaining:
        picks.append(calls[1])

    return picks[:remaining]


# ═══════════════════════════════════════════════════════════════════
#  TELEGRAM MESSAGE BUILDER
# ═══════════════════════════════════════════════════════════════════

def signal_message(picks, expiry, activity, updated_at):
    now     = ist_now().strftime('%d %b %Y  %I:%M %p')
    session = classify_session()
    bar     = "▓" * int(activity / 10) + "░" * (10 - int(activity / 10))
    line    = "─" * 28

    msg = (
        f"⚡ *NSE Options Signal*\n"
        f"🕐 {now} IST\n"
        f"📅 Expiry: `{expiry}`\n"
        f"📊 Market: `{bar}` {activity:.0f}/100\n"
        f"🕰 Session: {session}\n"
        f"🌡 VIX: `{vix_label()}`\n"
        f"📈 PCR: `{pcr_label()}`\n"
        f"`{line}`\n\n"
    )

    for s in picks:
        icon = "🟢" if s['direction'] == 'CALL' else "🔴"
        sign = "+" if s['day_change'] > 0 else ""
        tag  = "🆕 NEW" if s['is_new'] else f"📈 +{s['score_jump']:.0f} pts"

        pcr_b = s.get("pcr_bonus", 0)
        pcr_b_str = (f"   PCR Boost: `{pcr_b:+.0f} pts` ({'↑ confirms' if pcr_b > 0 else '↓ conflicts'})\n"
                     if pcr_b != 0 else "")
        # Calculate actual SL / Target stock price levels
        ltp = s['ltp']
        if s['direction'] == 'CALL':
            sl_price  = round(ltp * (1 - SL_PCT    / 100), 1)
            tgt_price = round(ltp * (1 + TARGET_PCT / 100), 1)
        else:
            sl_price  = round(ltp * (1 + SL_PCT    / 100), 1)
            tgt_price = round(ltp * (1 - TARGET_PCT / 100), 1)
        move_pts  = round(abs(tgt_price - ltp), 1)

        msg += (
            f"{icon} *{s['symbol']} — {s['direction']}*  {tag}\n"
            f"   LTP      : Rs {ltp:,.1f}\n"
            f"   Lot Size : {s['lot_size']:,}\n"
            f"   Day Chg  : {sign}{s['day_change']:.2f}%\n"
            f"   Vol Surge: {s['vol_change']:+.1f}%\n"
            f"   OI Build : +{s['oi_change']:.2f}%\n"
            f"   Score    : *{s['score']}/100*\n"
            f"{pcr_b_str}"
            f"   📌 Strike : ATM or 1 OTM {s['direction']}\n"
            f"   🛑 SL     : Rs {sl_price:,.1f}  ({SL_PCT:.1f}% away)\n"
            f"   🎯 Target : Rs {tgt_price:,.1f}  (+{move_pts:,.1f} pts, {TARGET_PCT:.1f}%)\n"
            f"   💰 Max Risk: Rs {MAX_RISK_PER_TRADE:,}\n"
            f"`{line}`\n\n"
        )

    remaining = MAX_TRADES_PER_DAY - state['alerts_sent'] - len(picks)
    msg += (
        f"🔢 Trade slots left: *{max(remaining, 0)}*\n"
        f"🛡 Daily limit: Rs {DAILY_LOSS_LIMIT:,}\n"
        f"_Data: Trendlyne · {updated_at}_"
    )
    return msg


def eod_summary_message():
    now  = ist_now().strftime('%d %b %Y')
    sent = state['alerts_sent']
    msg  = (
        f"📋 *End of Day — {now}*\n\n"
        f"   Scans run   : {state['scan_count']}\n"
        f"   Alerts sent : {sent}\n"
        f"   Stocks      : {', '.join(state['alerted_today']) or 'None'}\n\n"
    )
    msg += "✅ No trades today — capital preserved.\n" if sent == 0 else f"✅ {sent} signal(s) delivered.\n"
    msg += (
        f"\n*Rules for tomorrow:*\n"
        f"  Max risk/trade : Rs {MAX_RISK_PER_TRADE:,}\n"
        f"  Daily limit    : Rs {DAILY_LOSS_LIMIT:,}\n"
        f"  Monthly target : Rs {MONTHLY_TARGET[0]:,}–{MONTHLY_TARGET[1]:,}\n\n"
        f"_Rest well. Trade fresh tomorrow. 🙏_"
    )
    return msg


# ═══════════════════════════════════════════════════════════════════
#  TELEGRAM SENDER
# ═══════════════════════════════════════════════════════════════════

def send_telegram(text):
    if not ENABLE_TELEGRAM:
        return True
    url     = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {'chat_id': CHAT_ID, 'text': text, 'parse_mode': 'Markdown'}
    try:
        r  = requests.post(url, json=payload, timeout=15)
        ok = r.json().get('ok', False)
        print("[TG]", "✅ Sent" if ok else f"❌ {r.json()}")
        return ok
    except Exception as e:
        print(f"[TG ERROR] {e}")
        return False


# ═══════════════════════════════════════════════════════════════════
#  MAIN SCAN JOB
# ═══════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════
#  PRE-OPEN NSE DATA  (available at 9:07 AM, 8 min before open)
# ═══════════════════════════════════════════════════════════════════

PREOPEN_URL = "https://www.nseindia.com/api/market-data-pre-open?key=FO"

# ── Buildup direction map  (used by both scan_job and premarket) ──
BUILDUP_MAP = {
    "Long Build Up"    : ("CALL",  10),
    "Short Build Up"   : ("PUT",   10),
    "Short Covering"   : ("CALL",   5),
    "Long Unwinding"   : ("PUT",    5),
}

# ── Excel colour constants for pre-open sheet ─────────────────────
CALL_FG = "276221"
PUT_FG  = "9C0006"
CALL_BG = "C6EFCE"
PUT_BG  = "FFC7CE"
REV_BG  = "FFEB9C"

def hdr_cell(cell, text, width, ws):
    """Style a header cell in the pre-open sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    s = Side(style="thin", color="BFBFBF")
    cell.value     = text
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor="1F3864")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(left=s, right=s, top=s, bottom=s)


def data_cell(cell, value, bold=False, fg=None, bg=None):
    """Style a data cell in the pre-open sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    s = Side(style="thin", color="BFBFBF")
    cell.value     = value
    cell.font      = Font(name="Arial", size=10, bold=bold,
                          color=fg if fg else "000000")
    cell.fill      = PatternFill("solid", fgColor=bg) if bg else PatternFill()
    cell.border    = Border(left=s, right=s, top=s, bottom=s)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ── Telegram offset for polling ───────────────────────────────────
_tg_offset = 0

# Separate NSE session (different domain from Trendlyne)
_nse_session = requests.Session()
_nse_session.headers.update({
    "User-Agent"      : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept"          : "application/json, text/plain, */*",
    "Accept-Language" : "en-US,en;q=0.9",
    "Accept-Encoding" : "gzip, deflate, br",
    "Connection"      : "keep-alive",
    "Referer"         : "https://www.nseindia.com/option-chain",
    "X-Requested-With": "XMLHttpRequest",
})


def _warm_nse_session():
    """
    NSE requires a browser-like session visit before API calls.
    Visit homepage + pre-open page + option-chain page to get all cookies.
    """
    pages = [
        "https://www.nseindia.com/",
        "https://www.nseindia.com/market-data/pre-open-market-cm-and-emerge-market",
        "https://www.nseindia.com/option-chain",   # needed for PCR API cookie
    ]
    for url in pages:
        try:
            _nse_session.get(url, timeout=10)
        except Exception as e:
            print(f"[NSE] Warm-up page failed ({url}): {e}")
    print("[NSE] Session warmed up.")


# ═══════════════════════════════════════════════════════════════════
#  INDIA VIX + NIFTY PCR
# ═══════════════════════════════════════════════════════════════════

def fetch_india_vix():
    """Fetch India VIX from NSE allIndices API."""
    try:
        r = _nse_session.get(
            "https://www.nseindia.com/api/allIndices", timeout=10
        )
        r.raise_for_status()
        for idx in r.json().get("data", []):
            if idx.get("index") == "INDIA VIX":
                vix = round(float(idx.get("last", 0)), 2)
                _market_context["vix"] = vix
                chg = idx.get("percentChange", 0)
                print(f"[VIX] India VIX: {vix:.2f}  ({chg:+.2f}%)")
                return vix
    except Exception as e:
        print(f"[VIX ERROR] {e}")
    return None


def fetch_nifty_pcr():
    """
    Calculate Nifty PCR from live option chain (PE OI / CE OI).

    NSE changed their API — option-chain-indices now needs an expiryDate param.
    Strategy:
      1. Get near expiry from option-chain-contract-info (confirmed working)
      2. Try option-chain-indices with that expiry date
      3. Fallback: try without expiry (old style)
    """
    # ── Step 1: get near expiry date from working NSE endpoint ────
    near_expiry = None
    try:
        ci = _nse_session.get(
            "https://www.nseindia.com/api/option-chain-contract-info?symbol=NIFTY",
            timeout=10
        )
        ci.raise_for_status()
        expiry_dates = ci.json().get("expiryDates", [])
        near_expiry  = expiry_dates[0] if expiry_dates else None
        if near_expiry:
            print(f"[PCR] Near expiry from NSE: {near_expiry}")
    except Exception as e:
        print(f"[PCR] Could not fetch expiry dates: {e}")

    # ── Step 2: try option-chain-indices with and without expiry ──
    urls = []
    if near_expiry:
        urls.append(
            f"https://www.nseindia.com/api/option-chain-indices"
            f"?symbol=NIFTY&expiryDate={near_expiry}"
        )
    urls.append("https://www.nseindia.com/api/option-chain-indices?symbol=NIFTY")

    for url in urls:
        try:
            r = _nse_session.get(url, timeout=15)
            if r.status_code == 404:
                continue
            r.raise_for_status()
            data     = r.json()
            filtered = data.get("filtered", {})
            ce_oi    = filtered.get("CE", {}).get("totOI", 0)
            pe_oi    = filtered.get("PE", {}).get("totOI", 0)
            if ce_oi > 0:
                pcr = round(pe_oi / ce_oi, 3)
                _market_context["pcr"] = pcr
                print(f"[PCR] Nifty PCR: {pcr:.3f}  (PE {pe_oi:,} / CE {ce_oi:,})")
                return pcr
        except Exception as e:
            print(f"[PCR] {url} failed: {e}")
            continue

    print("[PCR] Could not fetch PCR — continuing without it")
    return None


def refresh_market_context(force=False):
    """
    Refresh VIX + PCR every 15 min (or immediately if force=True).
    Called at start of each scan — only hits NSE when stale.
    """
    now  = ist_now()
    last = _market_context["last_refresh"]
    if not force and last and (now - last).total_seconds() < 900:
        return   # Still fresh (< 15 min old)

    fetch_india_vix()
    fetch_nifty_pcr()
    _market_context["last_refresh"] = now


def vix_label():
    v = _market_context.get("vix")
    if v is None:  return "—"
    if v > VIX_EXTREME:  return f"{v:.1f} 🔴 EXTREME"
    if v > VIX_NORMAL_MAX: return f"{v:.1f} 🟠 HIGH"
    if v < VIX_CALM_MAX:   return f"{v:.1f} 🟢 CALM"
    return f"{v:.1f} 🟡 NORMAL"


def pcr_label():
    p = _market_context.get("pcr")
    if p is None:  return "—"
    if p >= PCR_VERY_BULL: return f"{p:.2f} 🟢 VERY BULLISH"
    if p >= PCR_BULL:      return f"{p:.2f} 🟢 BULLISH"
    if p <= PCR_VERY_BEAR: return f"{p:.2f} 🔴 VERY BEARISH"
    if p <= PCR_BEAR:      return f"{p:.2f} 🔴 BEARISH"
    return f"{p:.2f} 🟡 NEUTRAL"


def get_pcr_bonus(direction):
    """Return score bonus based on PCR alignment with signal direction."""
    pcr = _market_context.get("pcr")
    if pcr is None:
        return 0
    if direction == "CALL":
        if pcr >= PCR_VERY_BULL: return 15
        if pcr >= PCR_BULL:      return 8
        if pcr <= PCR_VERY_BEAR: return -10  # PCR says PUT, penalise CALL
        if pcr <= PCR_BEAR:      return -5
    elif direction == "PUT":
        if pcr <= PCR_VERY_BEAR: return 15
        if pcr <= PCR_BEAR:      return 8
        if pcr >= PCR_VERY_BULL: return -10
        if pcr >= PCR_BULL:      return -5
    return 0


def get_vix_threshold_delta():
    """
    Return how much to ADD to MIN_SCORE_TO_ALERT based on VIX.
    Positive = raise bar (risky market).  Negative = lower bar (calm).
    Returns (delta, block_signals).
    """
    vix = _market_context.get("vix")
    if vix is None:
        return 0, False
    if vix > VIX_EXTREME:
        return 0, True          # Block all signals
    if vix > VIX_NORMAL_MAX:
        return VIX_HIGH_BOOST, False
    if vix < VIX_CALM_MAX:
        return -5, False        # Ease slightly in calm market
    return 0, False


def fetch_preopen_data():
    """
    Fetch NSE F&O pre-open data.
    Enhanced: extracts order-book depth, ATO quantities, 52-week range,
    and final traded quantity alongside the existing BSR/IEP fields.
    Returns list of stock dicts.
    """
    try:
        r = _nse_session.get(PREOPEN_URL, timeout=15)
        if r.status_code == 401 or r.status_code == 403:
            print("[PREOPEN] Session expired — re-warming NSE session...")
            _warm_nse_session()
            r = _nse_session.get(PREOPEN_URL, timeout=15)

        r.raise_for_status()
        data = r.json()
        stocks = []

        for item in data.get("data", []):
            meta = item.get("metadata", {})
            pre  = item.get("detail", {}).get("preOpenMarket", {})
            sym  = meta.get("symbol", "")
            if not sym:
                continue

            # ── Basic fields (existing) ───────────────────────────
            buy_qty  = pre.get("totalBuyQuantity",  0) or 0
            sell_qty = pre.get("totalSellQuantity", 0) or 0
            iep      = pre.get("IEP", meta.get("lastPrice", 0)) or 0
            prev_close = meta.get("previousClose", 0) or 0

            # ── NEW 1: ATO (At-The-Open) market orders ────────────
            ato_buy  = pre.get("atoBuyQty",  0) or 0
            ato_sell = pre.get("atoSellQty", 0) or 0
            ato_net  = ato_buy - ato_sell   # positive = net buy pressure at open

            # ── NEW 2: Final traded quantity (IEP credibility) ────
            # High finalQuantity = real price discovery; low = ghost IEP
            final_qty = meta.get("finalQuantity", 0) or 0

            # ── NEW 3: 52-week range position ─────────────────────
            year_high = meta.get("yearHigh", 0) or 0
            year_low  = meta.get("yearLow",  0) or 0
            if year_high > year_low > 0 and iep > 0:
                # 0% = at 52-week low, 100% = at 52-week high
                yr_pos = round((iep - year_low) / (year_high - year_low) * 100, 1)
            else:
                yr_pos = 50.0   # unknown — neutral

            # ── NEW 4: Order book depth imbalance ─────────────────
            # Weighted bid/ask from all 10 price levels
            book_levels = pre.get("preopen", [])
            depth_buy  = 0
            depth_sell = 0
            for level in book_levels:
                bq = level.get("buyQty",  0) or 0
                sq = level.get("sellQty", 0) or 0
                depth_buy  += bq
                depth_sell += sq

            # Depth imbalance: > 0 means buyers, < 0 means sellers
            # Normalised to [-1, +1]
            depth_total = depth_buy + depth_sell
            depth_imbalance = round(
                (depth_buy - depth_sell) / depth_total, 3
            ) if depth_total > 0 else 0.0

            # Count how many buy levels have qty (order concentration)
            buy_levels_filled  = sum(1 for lv in book_levels if (lv.get("buyQty")  or 0) > 0)
            sell_levels_filled = sum(1 for lv in book_levels if (lv.get("sellQty") or 0) > 0)

            stocks.append({
                "symbol"          : sym,
                "iep"             : iep,
                "prev_close"      : prev_close,
                "pchange"         : meta.get("pChange", 0),
                # Existing
                "buy_qty"         : buy_qty,
                "sell_qty"        : sell_qty,
                "bsr"             : round(buy_qty / sell_qty, 3) if sell_qty > 0 else 99,
                "total_qty"       : buy_qty + sell_qty,
                "timestamp"       : pre.get("lastUpdateTime", ""),
                # New fields
                "ato_buy"         : ato_buy,
                "ato_sell"        : ato_sell,
                "ato_net"         : ato_net,
                "final_qty"       : final_qty,
                "year_high"       : year_high,
                "year_low"        : year_low,
                "yr_pos"          : yr_pos,          # % position in 52-wk range
                "depth_imbalance" : depth_imbalance, # -1 (sell) to +1 (buy)
                "buy_levels"      : buy_levels_filled,
                "sell_levels"     : sell_levels_filled,
            })

        print(f"[PREOPEN] {len(stocks)} F&O stocks fetched.")
        return stocks
    except Exception as e:
        print(f"[PREOPEN ERROR] {e}")
        return []


def score_preopen(s):
    """
    Pre-open signal score 0-100.

    Component breakdown (max pts):
      Gap magnitude          → 30 pts  (was 40)
      BSR conviction         → 20 pts  (was 40, now complemented by depth)
      Order-book depth       → 20 pts  NEW — weighted bid/ask imbalance
      ATO conviction         → 10 pts  NEW — committed market-order flow
      Direction bonus        →  8 pts  (was 10)
      IEP credibility        →  7 pts  NEW — penalise ghost prices
      52-week exhaustion     →  5 pts  NEW — room-to-run or exhaustion flag
    Total max               = 100 pts

    52-week range bonus/penalty:
      yr_pos <  30%  → +5 pts  (stock has room to run, gap more meaningful)
      yr_pos 30-70%  →  0 pts  (neutral)
      yr_pos 70-85%  → -5 pts  (approaching yearly resistance)
      yr_pos > 85%   → -12 pts (near 52-wk high = exhaustion risk)

      Same logic inverted for bearish gaps.

    IEP credibility (finalQuantity filter):
      Stocks with very few pre-open trades have an unreliable IEP.
      finalQuantity < 100   → -7 pts (near-zero credibility)
      finalQuantity 100-500 → -3 pts (low credibility)
      finalQuantity > 500   →  0 pts (credible price discovery)
    """
    pchg         = s.get("pchange", 0)
    bsr          = s.get("bsr", 1)
    depth_imb    = s.get("depth_imbalance", 0)   # -1 to +1
    ato_buy      = s.get("ato_buy",  0)
    ato_sell     = s.get("ato_sell", 0)
    ato_total    = ato_buy + ato_sell
    yr_pos       = s.get("yr_pos",  50.0)
    final_qty    = s.get("final_qty", 0)

    # ── 1. Gap size score (max 30 pts) ────────────────────────────
    gap_score = min(abs(pchg) * 8, 30)

    # ── 2. BSR conviction score (max 20 pts) ─────────────────────
    if pchg >= 0:
        bsr_score = min((bsr - 1) * 8, 20) if bsr > 1 else 0
    else:
        bsr_score = min((1 / bsr - 1) * 8, 20) if bsr < 1 else 0

    # ── 3. Order-book depth imbalance (max 20 pts) ────────────────
    # depth_imbalance is already direction-aware relative to the gap
    # For bullish gap: positive imbalance (buyers) = good
    # For bearish gap: negative imbalance (sellers) = good
    if pchg >= 0:
        depth_score = min(max(depth_imb, 0) * 20, 20)
    else:
        depth_score = min(max(-depth_imb, 0) * 20, 20)

    # ── 4. ATO conviction (max 10 pts) ───────────────────────────
    # Large net ATO in the gap direction = committed participants
    ato_score = 0
    if ato_total > 0:
        ato_imb = (ato_buy - ato_sell) / ato_total   # -1 to +1
        if pchg >= 0:
            ato_score = min(max(ato_imb, 0) * 15, 10)
        else:
            ato_score = min(max(-ato_imb, 0) * 15, 10)

    # ── 5. Direction bonus (max 8 pts) ───────────────────────────
    dir_bonus = 8 if abs(pchg) >= 1.0 else 4 if abs(pchg) >= 0.5 else 0

    # ── 6. IEP credibility penalty (max -7 pts) ──────────────────
    if final_qty < 100:
        credibility_adj = -7
    elif final_qty < 500:
        credibility_adj = -3
    else:
        credibility_adj = 0

    # ── 7. 52-week position adjustment ───────────────────────────
    # For bullish gaps: near high = exhaustion; near low = room to run
    if pchg >= 0:
        if yr_pos > 85:
            yr_adj = -12   # Near 52-week high — strong exhaustion risk
        elif yr_pos > 70:
            yr_adj = -5    # Approaching resistance
        elif yr_pos < 30:
            yr_adj = +5    # Plenty of room — gap more credible
        else:
            yr_adj = 0
    else:
        # For bearish gaps: near low = exhaustion; near high = room to fall
        if yr_pos < 15:
            yr_adj = -12   # Near 52-week low — bounce risk
        elif yr_pos < 30:
            yr_adj = -5
        elif yr_pos > 70:
            yr_adj = +5    # Lots of room to fall
        else:
            yr_adj = 0

    # ── Reversal flag ─────────────────────────────────────────────
    # Gap down but buyers dominating in depth AND ATO = reversal setup
    s["is_reversal"] = (
        pchg < -0.3
        and bsr >= 2.0
        and depth_imb > 0.2
    )

    total = gap_score + bsr_score + depth_score + ato_score + dir_bonus
    total += credibility_adj + yr_adj
    total = min(max(round(total, 1), 0), 100)

    # Store sub-scores for display in messages
    s["_score_breakdown"] = {
        "gap"        : round(gap_score, 1),
        "bsr"        : round(bsr_score, 1),
        "depth"      : round(depth_score, 1),
        "ato"        : round(ato_score, 1),
        "dir_bonus"  : dir_bonus,
        "credibility": credibility_adj,
        "yr_adj"     : yr_adj,
    }
    return total


def get_preopen_direction(s):
    """
    Determine trade direction from pre-open data.
    Enhanced: uses depth imbalance and ATO alongside BSR/pchange.

    Conviction tiers:
      HIGH   — gap + BSR + depth + ATO all aligned
      MEDIUM — gap + BSR + at least one of depth/ATO aligned
      LOW    — gap + BSR only (depth/ATO neutral or conflicting)
      NONE   — no clear signal
    """
    pchg      = s.get("pchange", 0)
    bsr       = s.get("bsr", 1)
    depth_imb = s.get("depth_imbalance", 0)
    ato_buy   = s.get("ato_buy",  0)
    ato_sell  = s.get("ato_sell", 0)
    ato_total = ato_buy + ato_sell
    ato_imb   = (ato_buy - ato_sell) / ato_total if ato_total > 0 else 0

    # Bullish conditions
    gap_up    = pchg >= 0.5
    bsr_bull  = bsr >= 1.5
    depth_bull = depth_imb > 0.15    # Buyers dominating the book
    ato_bull   = ato_imb > 0.2       # More market buy orders at open

    # Bearish conditions
    gap_down   = pchg <= -0.5
    bsr_bear   = bsr <= 0.67
    depth_bear = depth_imb < -0.15
    ato_bear   = ato_imb < -0.2

    # Reversal: gap down but order book filled with buyers
    reversal_bull = s.get("is_reversal", False)

    if gap_up and bsr_bull:
        s["conviction"] = "HIGH" if (depth_bull or ato_bull) else "MEDIUM"
        return "CALL"
    elif gap_down and bsr_bear:
        s["conviction"] = "HIGH" if (depth_bear or ato_bear) else "MEDIUM"
        return "PUT"
    elif reversal_bull:
        s["conviction"] = "REVERSAL"
        return "CALL"
    else:
        s["conviction"] = "NONE"
        return None


def analyze_preopen(preopen_stocks, oi_stocks):
    """
    Cross-reference pre-open data with OI data.
    Returns combined signals scored on both dimensions.
    """
    # Build OI lookup: symbol -> oi stock dict
    oi_lookup = {s["symbol"]: s for s in oi_stocks}

    combined = []
    for s in preopen_stocks:
        sym   = s["symbol"]
        score = score_preopen(s)
        if score < 30:
            continue

        direction = get_preopen_direction(s)
        if not direction:
            continue

        s["preopen_score"] = score
        s["direction"]     = direction

        # Bonus if OI data confirms same direction
        oi = oi_lookup.get(sym)
        s["oi_confirmed"] = False
        if oi:
            oi_dir = BUILDUP_MAP.get(oi.get("buildup", ""), (None, 0))[0]
            if oi_dir == direction:
                s["oi_confirmed"] = True
                s["preopen_score"] = min(score + 15, 100)  # Confirmation bonus
            s["oi_change"]  = oi.get("oi_change", 0)
            s["vol_change"] = oi.get("vol_change", 0)
            s["ltp"]        = oi.get("ltp", s["iep"])
            s["buildup"]    = oi.get("buildup", "")
        else:
            s["oi_change"]  = 0
            s["vol_change"] = 0
            s["ltp"]        = s["iep"]
            s["buildup"]    = ""

        combined.append(s)

    combined.sort(key=lambda x: x["preopen_score"], reverse=True)
    return combined[:8]   # Top 8 pre-open setups



def premarket_scan():
    """
    Runs at 9:10 AM IST.
    Combines NSE pre-open order book (BSR) + Trendlyne OI data
    for the strongest possible pre-market signal.
    """
    IST = pytz.timezone("Asia/Kolkata")
    now = datetime.now(IST)
    if now.weekday() >= 5:
        return

    print(f"[{now.strftime('%H:%M:%S')}] Pre-market combined scan running...")

    # ── Notify on Telegram that scan has started ──────────────────
    send_telegram(
        f"🔔 *PRE-MARKET SCAN STARTED*\n"
        f"━━━━━━━━━━━━━━━━━━━━━\n"
        f"🕐 {now.strftime('%d %b %Y  %I:%M %p')} IST\n"
        f"📋 Fetching NSE pre-open data + OI gainers...\n"
        f"_Full setup alert will follow in ~30 seconds_"
    )

    # ── Fetch VIX + PCR first (force refresh at market open) ─────
    refresh_market_context(force=True)

    # Fetch both data sources in parallel context
    oi_data     = fetch_oi_gainers()
    oi_stocks   = parse_rows(oi_data) if oi_data else []
    pre_stocks  = fetch_preopen_data()
    expiry      = _expiry_cache.get("expiry", "?")

    if not pre_stocks and not oi_stocks:
        print("[PRE-MARKET] No data available.")
        return

    # ── Combined analysis ─────────────────────────────────────────
    combined = analyze_preopen(pre_stocks, oi_stocks) if pre_stocks else []

    # ── OI-only setups (not in pre-open but strong OI) ───────────
    oi_only = []
    combined_syms = {s["symbol"] for s in combined}
    for s in oi_stocks:
        if s["symbol"] in combined_syms:
            continue
        oic = s["oi_change"]
        vc  = s["vol_change"]
        b   = s["buildup"]
        if b not in BUILDUP_MAP:
            continue
        direction, bonus = BUILDUP_MAP[b]
        if oic >= 5.0 and vc >= 40:
            oi_only.append({
                "symbol"   : s["symbol"],
                "direction": direction,
                "buildup"  : b,
                "ltp"      : s["ltp"],
                "oi_change": oic,
                "vol_change": vc,
                "preopen_score": min(oic * 4 + bonus, 80),
                "oi_confirmed" : False,
                "bsr"      : 0,
                "pchange"  : 0,
            })
    oi_only.sort(key=lambda x: x["preopen_score"], reverse=True)

    # ── Build Telegram message ────────────────────────────────────
    timestamp = pre_stocks[0]["timestamp"] if pre_stocks else now.strftime("%H:%M")

    # ── Market breadth from top-level API fields ──────────────────
    # These come from data.get("advances") etc. — fetch raw response
    # We already have pre_stocks; compute breadth from pchange
    adv = sum(1 for s in pre_stocks if s.get("pchange", 0) > 0)
    dec = sum(1 for s in pre_stocks if s.get("pchange", 0) < 0)
    unch = len(pre_stocks) - adv - dec
    breadth_bias = "BULLISH" if adv > dec * 1.5 else "BEARISH" if dec > adv * 1.5 else "MIXED"

    breadth_icon = "🟢" if breadth_bias == "BULLISH" else "🔴" if breadth_bias == "BEARISH" else "🟡"
    vix_d, vix_blk = get_vix_threshold_delta()

    msg  = "📊 *PRE-MARKET SETUP ALERT*\n"
    msg += "━━━━━━━━━━━━━━━━━━━━━\n"
    msg += "🕐 " + now.strftime("%d %b %Y  %I:%M %p") + " IST\n"
    msg += "📅 Expiry    : " + expiry + "\n"
    msg += "━━━━━━━━━━━━━━━━━━━━━\n"
    msg += breadth_icon + " Breadth    : ADV " + str(adv) + "  DEC " + str(dec) + "  UNCH " + str(unch)
    msg += "  (" + breadth_bias + ")\n"
    msg += "📈 India VIX : " + vix_label() + "\n"
    msg += "⚖️ Nifty PCR : " + pcr_label() + "\n"
    # VIX warning
    if vix_blk:
        msg += "⚠️ *VIX EXTREME* — wait for market to settle before trading\n"
    elif vix_d > 0:
        msg += f"⚠️ High VIX — score bar raised by {vix_d:.0f} pts today\n"
    elif vix_d < 0:
        msg += "✅ Low VIX — calm market, thresholds slightly eased\n"
    msg += "━━━━━━━━━━━━━━━━━━━━━\n\n"

    total_signals = 0

    if combined:
        msg += "COMBINED (Pre-Open + OI):\n\n"
        for i, s in enumerate(combined[:5], 1):
            sym   = s["symbol"]
            d     = s["direction"]
            sc    = s["preopen_score"]
            pchg  = s.get("pchange", 0)
            bsr   = s.get("bsr", 0)
            oic   = s.get("oi_change", 0)
            conf  = " + OI CONFIRMED" if s.get("oi_confirmed") else ""
            rev   = " [REVERSAL]" if s.get("is_reversal") else ""
            conv  = s.get("conviction", "")
            conv_tag = " [" + conv + "]" if conv and conv not in ("NONE", "") else ""

            # New metrics
            depth_imb  = s.get("depth_imbalance", 0)
            yr_pos     = s.get("yr_pos", 50.0)
            final_qty  = s.get("final_qty", 0)
            ato_net    = s.get("ato_net", 0)
            breakdown  = s.get("_score_breakdown", {})

            # IEP credibility label
            if final_qty < 100:
                cred_tag = " [LOW CREDIBILITY]"
            elif final_qty < 500:
                cred_tag = " [MED CREDIBILITY]"
            else:
                cred_tag = ""

            # 52-week range label
            if yr_pos > 85:
                yr_tag = " [NEAR 52W HIGH]"
            elif yr_pos > 70:
                yr_tag = " [UPPER RANGE]"
            elif yr_pos < 15:
                yr_tag = " [NEAR 52W LOW]"
            elif yr_pos < 30:
                yr_tag = " [LOWER RANGE]"
            else:
                yr_tag = ""

            msg += str(i) + ". " + sym + " - " + d + rev + conv_tag + conf + "\n"
            msg += "   IEP Gap : " + "{:+.2f}".format(pchg) + "%" + cred_tag + "\n"
            if bsr > 0:
                msg += "   BSR     : " + "{:.2f}".format(bsr) + "x"
            msg += "  Depth: " + ("{:+.0%}".format(depth_imb)) + "\n"
            if ato_net != 0:
                ato_dir = "buy" if ato_net > 0 else "sell"
                msg += "   ATO     : " + "{:+,}".format(ato_net) + " net " + ato_dir + "\n"
            msg += "   52W Pos : " + "{:.0f}%".format(yr_pos) + yr_tag + "\n"
            if oic:
                msg += "   OI Bld  : +" + "{:.1f}".format(oic) + "%\n"
            msg += "   Score   : " + str(sc) + "/100"
            if breakdown:
                msg += "  (gap " + str(breakdown.get("gap", 0))
                msg += " + bsr " + str(breakdown.get("bsr", 0))
                msg += " + depth " + str(breakdown.get("depth", 0))
                msg += " + ato " + str(breakdown.get("ato", 0))
                msg += " + yr " + str(breakdown.get("yr_adj", 0)) + ")"
            msg += "\n"
            msg += "   LTP     : Rs " + "{:,.1f}".format(s.get("ltp", s.get("iep", 0))) + "\n\n"
            total_signals += 1

    if oi_only:
        msg += "OI ONLY (strong pre-open OI):\n\n"
        for s in oi_only[:3]:
            msg += "- " + s["symbol"] + " " + s["direction"]
            msg += "  OI+" + "{:.1f}".format(s["oi_change"]) + "%"
            msg += "  Vol+" + "{:.0f}".format(s["vol_change"]) + "%\n"

    msg += "\n----------------------------\n"
    msg += "BSR    = Buy/Sell Ratio (order book top-level)\n"
    msg += "Depth  = Weighted bid/ask imbalance (all 10 book levels)\n"
    msg += "ATO    = At-The-Open market orders (committed flow)\n"
    msg += "52W Pos= IEP % position in 52-week range\n"
    msg += "BSR >2 = strong buying  |  BSR <0.5 = strong selling\n"
    msg += "Depth +50%+ = dominant buyers  |  -50%- = dominant sellers\n"
    msg += "These are WATCH signals - wait for 9:15 confirmation"

    send_telegram(msg)
    send_desktop(
        "Pre-Market: " + str(total_signals) + " combined setups",
        ", ".join([s["symbol"] + " " + s["direction"]
                   for s in combined[:3]]) or "OI only signals",
        urgent=False
    )
    excel_log_preopen(combined + oi_only[:3], expiry, now.strftime("%H:%M"))
    print(f"[PRE-MARKET] {total_signals} combined + {len(oi_only)} OI-only setups sent.")



def excel_log_preopen(setups, expiry, scan_time):
    """Log pre-open combined setups to a dedicated sheet — enhanced columns."""
    if not ENABLE_EXCEL_LOG or not _openpyxl or not setups:
        return
    try:
        wb  = _load_or_create_wb()

        # Create Pre-Open sheet if missing
        if "Pre-Open" not in wb.sheetnames:
            ws = wb.create_sheet("Pre-Open")
            ws.row_dimensions[1].height = 32
            headers = [
                "Date", "Time", "Expiry", "Symbol", "Direction",
                "IEP Gap %", "BSR", "Depth Imb", "ATO Net",
                "52W Pos %", "Final Qty", "Pre-Open Score",
                "Conviction", "OI Confirmed",
                "OI Change %", "Vol Change %", "LTP", "Buildup", "Notes"
            ]
            widths  = [12, 8, 12, 14, 10,  11, 8, 11, 10,  10, 10, 14,  12, 13,  12, 12, 10, 18, 20]
            for col, (h, w) in enumerate(zip(headers, widths), 1):
                hdr_cell(ws.cell(1, col), h, w, ws)
                ws.column_dimensions[get_column_letter(col)].width = w
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = "A1:S1"

        ws  = wb["Pre-Open"]
        IST = pytz.timezone("Asia/Kolkata")
        today = datetime.now(IST).strftime("%Y-%m-%d")

        for s in setups:
            d   = s.get("direction", "")
            fg  = CALL_FG if d == "CALL" else PUT_FG
            bg  = REV_BG  if s.get("is_reversal") else (CALL_BG if d == "CALL" else PUT_BG)
            r   = ws.max_row + 1

            # 52-week credibility label
            fq = s.get("final_qty", 0)
            cred = "LOW" if fq < 100 else "MED" if fq < 500 else "OK"

            vals = [
                today,
                scan_time,
                expiry,
                s.get("symbol", ""),
                d,
                round(s.get("pchange", 0), 2),
                round(s.get("bsr", 0), 2),
                round(s.get("depth_imbalance", 0), 3),
                s.get("ato_net", 0),
                round(s.get("yr_pos", 50.0), 1),
                fq,
                s.get("preopen_score", 0),
                s.get("conviction", ""),
                "YES" if s.get("oi_confirmed") else "NO",
                round(s.get("oi_change", 0), 2),
                round(s.get("vol_change", 0), 1),
                s.get("ltp", s.get("iep", 0)),
                s.get("buildup", ""),
                ("REVERSAL" if s.get("is_reversal") else "") + (" IEP-CRED:" + cred if cred != "OK" else ""),
            ]
            for col, val in enumerate(vals, 1):
                data_cell(ws.cell(r, col), val, bold=(col <= 5), fg=fg, bg=bg)

        wb.save(EXCEL_FILE)
        print(f"[EXCEL] Pre-open setups logged.")
    except Exception as e:
        print(f"[EXCEL PREOPEN ERROR] {e}")


def scan_job():
    now = ist_now()

    if state['date'] != now.date():
        reset_daily_state()

    if not is_market_open():
        print(f"[{now.strftime('%H:%M')}] Market closed.")
        return

    if state['alerts_sent'] >= MAX_TRADES_PER_DAY:
        print(f"[{now.strftime('%H:%M')}] Daily cap reached.")
        return

    state['scan_count'] += 1
    print(f"\n[{now.strftime('%H:%M:%S')}] Scan #{state['scan_count']}")

    # ── Refresh VIX + PCR every 15 min ───────────────────────────
    refresh_market_context()

    data   = fetch_oi_gainers()
    expiry = _expiry_cache.get('expiry', '?')
    if not data:
        print("[ERROR] Fetch failed.")
        return

    stocks   = parse_rows(data)
    activity = market_activity_score(stocks)
    state['activity_history'].append(activity)
    if len(state['activity_history']) > 12:
        state['activity_history'].pop(0)

    session = classify_session()
    print(f"[INFO] {len(stocks)} stocks | Activity: {activity}/100 | {session}")

    # ── Excel: log every scan ─────────────────────────────────────
    excel_log_scan(state['scan_count'], len(stocks), activity,
                   session, state['alerts_sent'], state['alerted_today'])

    picks = filter_signals(stocks)

    for s in stocks:
        sc = compute_score(s)
        if sc > 0:
            state['last_scores'][s['symbol']] = sc

    # ── Debug: show top 5 candidates every scan ───────────────────
    _candidates = []
    for s in stocks:
        b  = s.get('buildup', '')
        dc = s.get('day_change', 0)
        if (b == 'Long Build Up' and dc >= MIN_DAY_CHANGE_LONG) or \
           (b == 'Short Build Up' and dc <= MIN_DAY_CHANGE_SHORT):
            sc = compute_score(s)
            _candidates.append((s['symbol'], dc, s.get('oi_change',0), s.get('vol_change',0), sc, b))
    _candidates.sort(key=lambda x: x[4], reverse=True)
    if _candidates:
        print(f"[DEBUG] Top candidates (threshold={MIN_SCORE_TO_ALERT}):")
        for sym, dc, oic, vc, sc, bu in _candidates[:5]:
            flag = "✓ SIGNAL" if sc >= MIN_SCORE_TO_ALERT else f"✗ score<{MIN_SCORE_TO_ALERT}"
            print(f"  {sym}: dc={dc:+.2f}% oic={oic:.1f}% vc={vc:.0f}% score={sc} [{bu[:3]}] {flag}")
    else:
        print(f"[DEBUG] No qualifying buildup/direction stocks this scan.")

    # ── Telegram scan heartbeat (every scan) ──────────────────────
    _time_str = now.strftime('%I:%M %p')
    _act_bar  = "▓" * int(activity / 20) + "░" * (5 - int(activity / 20))
    _vix_d, _vix_blk = get_vix_threshold_delta()
    _eff_thresh = int(MIN_SCORE_TO_ALERT + _vix_d)

    if picks:
        _result_line = f"✅ *SIGNAL FIRED* — {', '.join(p['symbol'] for p in picks)}"
    elif _candidates:
        top = _candidates[0]
        _gap = _eff_thresh - top[4]  # how many pts away from threshold
        _dir = "🟢" if top[5].startswith("Long") else "🔴"
        _result_line = (
            f"⏭ No signal  _(closest: {_dir}{top[0]}  score `{top[4]:.0f}`"
            f"  needs `{_eff_thresh}`)_"
        )
    else:
        _result_line = f"⏭ No qualifying stocks  _(need dc≥{MIN_DAY_CHANGE_LONG}% + OI≥{MIN_OI_CHANGE}%)_"

    _top_lines = ""
    for sym, dc, oic, vc, sc, bu in _candidates[:3]:
        _dir_ic = "🟢" if bu.startswith("Long") else "🔴"
        _bar = "█" * int(sc / 10) + "░" * (10 - int(sc / 10))
        _top_lines += f"  {_dir_ic} `{sym:<12}` dc`{dc:+.1f}%` OI`+{oic:.1f}%` score`{sc:.0f}/{_eff_thresh}`\n"

    _tg_scan_msg = (
        f"🔍 *Scan #{state['scan_count']}*  ·  {_time_str} IST  ·  {session}\n"
        f"📦 {len(stocks)} stocks  ·  Activity `{_act_bar}` `{activity:.0f}/100`\n"
    )
    if _vix_blk:
        _tg_scan_msg += f"⛔ VIX EXTREME — signals blocked\n"
    elif _vix_d != 0:
        _tg_scan_msg += f"⚠️ VIX adj: threshold raised to `{_eff_thresh}`\n"
    if _top_lines:
        _tg_scan_msg += f"📊 Top candidates:\n{_top_lines}"
    _tg_scan_msg += _result_line
    send_telegram(_tg_scan_msg)

    if not picks:
        print(f"[INFO] No new signals. Alerted: {state['alerted_today']}")
        return

    # ── Fire both channels ────────────────────────────────────────
    notify_signal(picks, activity, session)           # Desktop pop-up
    msg = signal_message(picks, expiry, activity, data.get('contractLastUpdated', ''))
    ok  = send_telegram(msg)                          # Telegram

    if ok or ENABLE_DESKTOP_NOTIFY:
        for p in picks:
            state['alerted_today'].add(p['symbol'])
        state['alerts_sent'] += len(picks)

        # ── Excel: log signals ────────────────────────────────────
        excel_log_signals(picks, activity, session, expiry)

        if state['alerts_sent'] >= MAX_TRADES_PER_DAY:
            notify_daily_cap()
            send_telegram(
                f"🛑 *Daily limit reached*\n"
                f"All {MAX_TRADES_PER_DAY} trade slots used.\n"
                f"Bot resumes tomorrow at 9:15 AM IST. 🙏"
            )



# ═══════════════════════════════════════════════════════════════════
#  P&L TRACKER  — update trade results from Telegram commands
# ═══════════════════════════════════════════════════════════════════

# Tracks last signal row per symbol for Telegram updates
# { symbol: excel_row_number }
_signal_rows = {}


def excel_update_result(symbol, result, pnl=None, entry=None, exit_p=None, notes=None):
    """Update P&L columns for a signal row by symbol."""
    if not ENABLE_EXCEL_LOG or not _openpyxl:
        return False
    try:
        wb  = _load_or_create_wb()
        ws  = wb["Signals"]
        IST = pytz.timezone("Asia/Kolkata")
        today = datetime.now(IST).strftime("%Y-%m-%d")

        # Find last row for this symbol today
        target_row = None
        for row in ws.iter_rows(min_row=2):
            if row[0].value == today and row[2].value == symbol:
                target_row = row[0].row

        if not target_row:
            return False

        # Col layout: 18=Entry Premium, 19=Exit Premium, 20=P&L, 21=Trade Outcome, 22=Notes
        # Normalise legacy WIN/LOSS → TP/SL
        outcome = result
        if outcome == "WIN":  outcome = "TP"
        if outcome == "LOSS": outcome = "SL"

        # Color scheme: TP=green, SL=red, SKIP=grey, PENDING=white
        _fg = {"TP": "276221", "SL": "9C0006", "SKIP": "595959"}.get(outcome, "000000")
        _bg = {"TP": "C6EFCE", "SL": "FFC7CE", "SKIP": "F2F2F2"}.get(outcome, "FFFFFF")

        updates = {18: entry, 19: exit_p, 20: pnl, 21: outcome, 22: notes}
        for col, val in updates.items():
            if val is not None:
                cell = ws.cell(target_row, col, value=val)
                cell.font      = Font(name="Arial", size=10, bold=(col == 21), color=_fg)
                cell.fill      = PatternFill("solid", fgColor=_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _thin_border()

        wb.save(EXCEL_FILE)
        print(f"[EXCEL] {symbol} → {result} | P&L: {pnl}")
        return True
    except Exception as e:
        print(f"[EXCEL UPDATE ERROR] {e}")
        return False


# ═══════════════════════════════════════════════════════════════════
#  WEEKLY REPORT
# ═══════════════════════════════════════════════════════════════════

def weekly_report():
    """Generate weekly performance report from Excel data."""
    if not _openpyxl or not os.path.exists(EXCEL_FILE):
        return "No data yet."
    try:
        from datetime import timedelta
        wb    = openpyxl.load_workbook(EXCEL_FILE)
        ws    = wb["Signals"]
        IST   = pytz.timezone("Asia/Kolkata")
        today = datetime.now(IST).date()

        week_ago = (today - timedelta(days=7)).strftime("%Y-%m-%d")
        today_s  = today.strftime("%Y-%m-%d")

        # Col indices (0-based): Trade Outcome=20, P&L=19, Session=15
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0] and isinstance(r[0], str)
                and week_ago <= r[0] <= today_s
                and r[20] is not None and r[20] != "PENDING"]

        if not rows:
            return "Weekly Report\n\nNo completed trades this week.\nTip: Reply TP/SL/SKIP to signals to track results."

        win_ct    = sum(1 for r in rows if r[20] in ("TP", "WIN"))
        loss_ct   = sum(1 for r in rows if r[20] in ("SL", "LOSS"))
        skip_ct   = sum(1 for r in rows if r[20] == "SKIP")
        total     = len(rows)
        total_pnl = sum(r[19] for r in rows if isinstance(r[19], (int, float)))
        win_rate  = round(win_ct / (win_ct + loss_ct) * 100) if (win_ct + loss_ct) > 0 else 0

        pnl_by_stock = {}
        for r in rows:
            sym = r[2]; p = r[19] if isinstance(r[19], (int,float)) else 0
            pnl_by_stock[sym] = pnl_by_stock.get(sym, 0) + p
        best_sym  = max(pnl_by_stock, key=pnl_by_stock.get) if pnl_by_stock else "-"

        sess_pnl  = {}
        for r in rows:
            s = r[15] or "?"; p = r[19] if isinstance(r[19], (int,float)) else 0
            sess_pnl[s] = sess_pnl.get(s, 0) + p
        best_sess = max(sess_pnl, key=sess_pnl.get) if sess_pnl else "-"

        pnl_icon  = "UP" if total_pnl >= 0 else "DOWN"
        from_date = (today - timedelta(days=7)).strftime("%d %b")
        to_date   = today.strftime("%d %b %Y")

        msg  = "Weekly Report\n"
        msg += from_date + " to " + to_date + "\n"
        msg += "----------------------------\n\n"
        msg += "Signals  : " + str(total) + "\n"
        msg += "Wins: " + str(win_ct) + "  Losses: " + str(loss_ct) + "  Skip: " + str(skip_ct) + "\n"
        msg += "Win Rate : " + str(win_rate) + "%\n"
        msg += pnl_icon + " Total P&L : Rs " + "{:+,.0f}".format(total_pnl) + "\n\n"
        msg += "Best Stock  : " + best_sym + "  (Rs " + "{:+,.0f}".format(pnl_by_stock.get(best_sym, 0)) + ")\n"
        msg += "Best Session: " + best_sess + "  (Rs " + "{:+,.0f}".format(sess_pnl.get(best_sess, 0)) + ")\n"

        month_start = today.strftime("%Y-%m-01")
        month_rows  = [r for r in ws.iter_rows(min_row=2, values_only=True)
                       if r[0] and isinstance(r[0], str) and r[0] >= month_start
                       and isinstance(r[19], (int, float))]
        month_pnl = sum(r[19] for r in month_rows)
        progress  = min(int(month_pnl / MONTHLY_TARGET[1] * 10), 10) if MONTHLY_TARGET[1] else 0
        bar       = "#" * progress + "." * (10 - progress)

        msg += "\n----------------------------\n"
        msg += "Monthly P&L:\n"
        msg += "[" + bar + "] Rs " + "{:+,.0f}".format(month_pnl) + "\n"
        msg += "Target: Rs " + "{:,}".format(MONTHLY_TARGET[0]) + " to Rs " + "{:,}".format(MONTHLY_TARGET[1])

        return msg

    except Exception as e:
        return "Weekly report error: " + str(e)


def auto_tune():
    """Analyse historical data and suggest optimal MIN_SCORE_TO_ALERT."""
    if not _openpyxl or not os.path.exists(EXCEL_FILE):
        return "No data for tuning yet. Need at least 10 completed trades."
    try:
        wb   = openpyxl.load_workbook(EXCEL_FILE)
        ws   = wb["Signals"]
        # Col indices (0-based): Trade Outcome=20, P&L=19, Score=8
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r[20] in ("TP", "SL", "WIN", "LOSS") and isinstance(r[8], (int, float))]

        if len(rows) < 10:
            return ("Auto-Tune\n\nNeed at least 10 TP/SL trades.\n"
                    "Currently have: " + str(len(rows)) + "\nKeep logging results.")

        buckets = {}
        for r in rows:
            b = (int(r[8]) // 10) * 10
            if b not in buckets:
                buckets[b] = {"win": 0, "loss": 0, "pnl": 0}
            if r[20] in ("TP", "WIN"):
                buckets[b]["win"]  += 1
            else:
                buckets[b]["loss"] += 1
            if isinstance(r[19], (int, float)):
                buckets[b]["pnl"] += r[19]

        msg  = "Auto-Tune Analysis\n----------------------------\n\n"
        msg += "{:<10} {:<8} {:<8} {}\n".format("Score", "Trades", "Win%", "P&L")

        best_bucket = None
        best_wr     = 0
        for b in sorted(buckets):
            d  = buckets[b]
            t  = d["win"] + d["loss"]
            wr = round(d["win"] / t * 100) if t else 0
            msg += "{}-{:<6} {:<8} {:<8} Rs {:+,.0f}\n".format(b, b+9, t, str(wr)+"%", d["pnl"])
            if wr > best_wr and t >= 3:
                best_wr = wr; best_bucket = b

        current   = int(MIN_SCORE_TO_ALERT)
        suggested = best_bucket if best_bucket else current
        msg += "\n----------------------------\n"
        msg += "Current threshold : " + str(current) + "\n"
        msg += "Suggested         : " + str(suggested)
        if suggested != current:
            msg += "  <- change MIN_SCORE_TO_ALERT to " + str(suggested)
        msg += "\n\nBased on " + str(len(rows)) + " completed trades"
        return msg

    except Exception as e:
        return "Auto-tune error: " + str(e)


def process_command(text, username=""):
    """Parse and act on an incoming Telegram message."""
    text  = text.strip()
    parts = text.upper().split()
    if not parts:
        return None
    cmd = parts[0]

    # ── TP / WIN [SYMBOL] AMOUNT ──────────────────────────────────
    if cmd in ("TP", "WIN"):
        sym = parts[1] if len(parts) == 3 else ((list(state["alerted_today"]) or ["?"])[-1])
        pnl = float(parts[-1]) if parts[-1].replace(".","").isdigit() else None
        excel_update_result(sym, "TP", pnl=pnl)
        reply = "✅ TP logged for " + sym
        if pnl: reply += "  |  P&L: Rs +" + "{:,.0f}".format(pnl)
        return reply

    # ── SL / LOSS [SYMBOL] AMOUNT ─────────────────────────────────
    if cmd in ("SL", "LOSS"):
        sym = parts[1] if len(parts) == 3 else ((list(state["alerted_today"]) or ["?"])[-1])
        pnl = -abs(float(parts[-1])) if parts[-1].replace(".","").isdigit() else None
        excel_update_result(sym, "SL", pnl=pnl)
        reply = "🔴 SL logged for " + sym
        if pnl: reply += "  |  P&L: Rs " + "{:,.0f}".format(pnl)
        return reply

    # ── SKIP [SYMBOL] ─────────────────────────────────────────────
    if cmd == "SKIP":
        sym = parts[1] if len(parts) > 1 else ((list(state["alerted_today"]) or ["?"])[-1])
        excel_update_result(sym, "SKIP")
        return "⏭ SKIP logged for " + sym

    # ── ENTRY AMOUNT [SYMBOL] ─────────────────────────────────────
    if cmd == "ENTRY" and len(parts) >= 2:
        amount = float(parts[1]) if parts[1].replace(".","").isdigit() else None
        sym    = parts[2] if len(parts) > 2 else ((list(state["alerted_today"]) or ["?"])[-1])
        if amount:
            excel_update_result(sym, "OPEN", entry=amount)
            return "Entry Rs " + "{:,.1f}".format(amount) + " logged for " + sym

    # ── EXIT AMOUNT [SYMBOL] ──────────────────────────────────────
    if cmd == "EXIT" and len(parts) >= 2:
        amount = float(parts[1]) if parts[1].replace(".","").isdigit() else None
        sym    = parts[2] if len(parts) > 2 else ((list(state["alerted_today"]) or ["?"])[-1])
        if amount:
            excel_update_result(sym, "CLOSED", exit_p=amount)
            return "Exit Rs " + "{:,.1f}".format(amount) + " logged for " + sym

    # ── WEEKLY ────────────────────────────────────────────────────
    if cmd == "WEEKLY":
        return weekly_report()

    # ── TUNE ──────────────────────────────────────────────────────
    if cmd == "TUNE":
        return auto_tune()

    # ── STATUS ────────────────────────────────────────────────────
    if cmd == "STATUS":
        now = ist_now().strftime("%H:%M")
        vix_d, vix_blk = get_vix_threshold_delta()
        eff_threshold  = MIN_SCORE_TO_ALERT + vix_d
        msg  = "Bot Status " + now + " IST\n\n"
        msg += "  Scans today  : " + str(state["scan_count"]) + "\n"
        msg += "  Alerts today : " + str(state["alerts_sent"]) + "\n"
        msg += "  Alerted      : " + (", ".join(state["alerted_today"]) or "None") + "\n"
        msg += "  Market open  : " + ("Yes" if is_market_open() else "No") + "\n"
        msg += "  Score min    : " + str(MIN_SCORE_TO_ALERT)
        if vix_d != 0:
            msg += f" → {eff_threshold:.0f} (VIX adj {vix_d:+.0f})"
        msg += "\n"
        msg += "  Slots left   : " + str(MAX_TRADES_PER_DAY - state["alerts_sent"]) + "\n"
        msg += "  India VIX    : " + vix_label() + "\n"
        msg += "  Nifty PCR    : " + pcr_label()
        if vix_blk:
            msg += "\n  *** SIGNALS BLOCKED — VIX EXTREME ***"
        return msg

    # ── HELP ──────────────────────────────────────────────────────
    if cmd == "HELP":
        msg  = "Commands\n\n"
        msg += "WIN 3200           - log win, Rs 3200 profit\n"
        msg += "WIN UNITDSPR 3200  - specify stock\n"
        msg += "LOSS 1500          - log loss\n"
        msg += "SKIP               - signal not taken\n"
        msg += "ENTRY 45           - log entry premium\n"
        msg += "EXIT 85            - log exit premium\n"
        msg += "WEEKLY             - weekly performance report\n"
        msg += "TUNE               - analyse score threshold\n"
        msg += "STATUS             - bot status + VIX + PCR\n"
        msg += "HELP               - this message\n\n"
        msg += "VIX rules:\n"
        msg += f"  < {VIX_CALM_MAX}   = CALM  (score bar -5)\n"
        msg += f"  {VIX_CALM_MAX}-{VIX_NORMAL_MAX} = NORMAL\n"
        msg += f"  > {VIX_NORMAL_MAX}  = HIGH  (score bar +{VIX_HIGH_BOOST:.0f})\n"
        msg += f"  > {VIX_EXTREME}  = EXTREME (signals blocked)\n\n"
        msg += "PCR rules:\n"
        msg += f"  > {PCR_VERY_BULL} = CALL +15 pts\n"
        msg += f"  > {PCR_BULL} = CALL  +8 pts\n"
        msg += f"  < {PCR_BEAR} = PUT   +8 pts\n"
        msg += f"  < {PCR_VERY_BEAR} = PUT  +15 pts"
        return msg

    return None


def telegram_poll_loop():
    """
    Background thread: long-polls Telegram for incoming messages.
    Runs continuously — handles trade result commands from user.
    """
    global _tg_offset
    print("[TG POLL] Listener started.")

    while True:
        try:
            url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/getUpdates"
            r   = requests.get(url,
                               params={"offset": _tg_offset + 1, "timeout": 25},
                               timeout=30)
            updates = r.json().get("result", [])

            for upd in updates:
                _tg_offset = upd["update_id"]
                msg  = upd.get("message") or upd.get("edited_message")
                if not msg:
                    continue
                text = msg.get("text", "").strip()
                if not text:
                    continue
                cid  = str(msg["chat"]["id"])
                if cid != str(CHAT_ID):
                    continue   # Ignore messages from other chats

                reply = process_command(text)
                if reply:
                    send_telegram(reply)

        except Exception as e:
            pass   # Network hiccup — retry silently

        time.sleep(2)


def start_telegram_listener():
    """Launch the Telegram command listener in a daemon thread."""
    import threading
    t = threading.Thread(target=telegram_poll_loop, daemon=True)
    t.start()



def _update_watchlist(symbol, s, direction):
    """Add/update a stock in the watchlist with current metrics."""
    existing = state["watchlist"].get(symbol, {})
    state["watchlist"][symbol] = {
        "direction"  : direction,
        "score"      : s.get("score", 0),
        "ltp"        : s.get("ltp", 0),
        "day_change" : s.get("day_change", 0),
        "oi_change"  : s.get("oi_change", 0),
        "oi_accel"   : s.get("oi_accel", 0),
        "buildup"    : s.get("buildup", ""),
        "first_seen" : existing.get("first_seen", ist_now().strftime("%H:%M")),
        "scans"      : existing.get("scans", 0) + 1,
    }


def check_watchlist_promotions():
    """
    Check if any watchlist stock now crosses WATCHLIST_ALERT_SCORE.
    Returns list of stocks ready to alert.
    """
    promotions = []
    for sym, w in list(state["watchlist"].items()):
        if sym in state["alerted_today"]:
            state["watchlist"].pop(sym, None)
            continue
        if w["score"] >= WATCHLIST_ALERT_SCORE and w["scans"] >= 2:
            promotions.append(w | {"symbol": sym, "signal_type": "WATCHLIST"})
    return promotions


def send_watchlist_alert(promotions):
    """Send Telegram + desktop alert for watchlist promotions."""
    if not promotions:
        return
    line = "----------------------------"
    msg  = "WATCHLIST ALERT - Building Up\n" + line + "\n\n"
    for w in promotions:
        msg += w["symbol"] + " - " + w["direction"] + " (WATCH)\n"
        msg += "  Buildup : " + w["buildup"] + "\n"
        msg += "  Score   : " + str(w["score"]) + "/100  (since " + w["first_seen"] + ")\n"
        msg += "  Day Chg : " + "{:+.2f}".format(w["day_change"]) + "%  OI: +" + "{:.2f}".format(w["oi_change"]) + "%\n"
        msg += "  OI Accel: " + "{:+.2f}".format(w["oi_accel"]) + "% this scan\n"
        msg += "  Not a trade yet - monitor closely\n"
        msg += line + "\n\n"
    msg += "Score building - signal may fire next scan"
    send_telegram(msg)
    for w in promotions:
        send_desktop(
            "WATCH: " + w["symbol"] + " " + w["direction"] + " (Score " + str(w["score"]) + ")",
            "Building since " + w["first_seen"] + " | OI accel: " + "{:+.1f}".format(w["oi_accel"]) + "%",
            urgent=False
        )


def eod_job():
    if state['scan_count'] > 0:
        notify_eod()
        send_telegram(eod_summary_message())
        excel_log_eod()
        # Friday weekly report
        if ist_now().weekday() == 4:   # 4 = Friday
            send_telegram(weekly_report())


# ═══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys

    IST_NOW = ist_now()

    # ── Test mode — runs regardless of market hours / day ───────────
    TEST_MODE = (os.environ.get("TEST_MODE", "false").lower() == "true"
                 or (len(sys.argv) > 1 and sys.argv[1] == 'test'))

    if TEST_MODE:
        if len(sys.argv) <= 1:
            sys.argv.append('test')   # reuse existing test block below

    if len(sys.argv) > 1 and sys.argv[1] == 'test':
        print("🧪 Test mode...")
        reset_daily_state()
        data   = fetch_oi_gainers()
        expiry = _expiry_cache.get('expiry', '?')
        if data:
            stocks   = parse_rows(data)
            activity = market_activity_score(stocks)
            state['activity_history'] = [50, 50, 50, 50]   # Force active
            picks    = filter_signals(stocks)
            updated  = data.get('contractLastUpdated', '')

            # Test desktop notification
            send_desktop(
                "NSE Bot — Test Notification",
                "If you see this, desktop notifications are working! ✅",
                urgent=False
            )

            if picks:
                notify_signal(picks, activity, classify_session())
                msg = signal_message(picks, expiry, activity, updated)
            else:
                msg = (
                    f"🧪 *Test OK*\n"
                    f"Activity: {activity}/100\n"
                    f"No signals above threshold right now.\n"
                    f"Desktop notification sent. Check your screen. ✅"
                )
            send_telegram(msg)
        else:
            send_telegram("⚠️ Test failed — data fetch error.")
        sys.exit(0)

    # ── Guard: exit immediately if started outside market window ────
    _mins_now   = IST_NOW.hour * 60 + IST_NOW.minute
    _is_weekday = IST_NOW.weekday() < 5
    _after_close    = _mins_now > 15 * 60 + 35
    _before_preopen = _mins_now < 9 * 60 + 5

    if not _is_weekday:
        print(f"[{IST_NOW.strftime('%A %H:%M')} IST] Weekend — bot exiting.")
        sys.exit(0)
    if _after_close or _before_preopen:
        print(f"[{IST_NOW.strftime('%H:%M')} IST] Outside market window — bot exiting.")
        sys.exit(0)

    # ── Pre-market only mode ─────────────────────────────────────────
    PREMARKET_ONLY = os.environ.get("PREMARKET_ONLY", "false").lower() == "true"
    if PREMARKET_ONLY:
        print(f"[{IST_NOW.strftime('%H:%M')} IST] Running pre-market scan only...")
        _warm_nse_session()
        _warm_trendlyne_session()
        refresh_market_context(force=True)
        premarket_scan()
        print("[PREMARKET] Done. Exiting.")
        sys.exit(0)

    reset_daily_state()

    # Print startup info
    desktop_status = f"✅ ({_notifier})" if _notifier else "❌ (install plyer)"
    tg_status      = "✅" if ENABLE_TELEGRAM else "⏸ disabled"
    print(f"""
┌─────────────────────────────────────────┐
│     NSE Smart Options Bot — Started     │
├─────────────────────────────────────────┤
│ Market hours : 09:15 – 15:30 IST        │
│ Scan every   : {SCAN_INTERVAL_MIN} minutes                    │
│ Min score    : {MIN_SCORE_TO_ALERT} / 100                    │
│ Telegram     : {tg_status:<30} │
│ Desktop notif: {desktop_status:<30} │
└─────────────────────────────────────────┘
💡 Test : python nse_options_bot.py test
📩 Commands: WIN / LOSS / SKIP / ENTRY / EXIT / WEEKLY / TUNE / STATUS
""")

    schedule.every(SCAN_INTERVAL_MIN).minutes.do(scan_job)
    schedule.every().day.at("10:05").do(eod_job)   # 3:35 PM IST = 10:05 UTC
    schedule.every().day.at("03:40").do(premarket_scan)  # 9:10 AM IST

    # Warm NSE + Trendlyne sessions before first API call
    _warm_nse_session()
    _warm_trendlyne_session()

    # ── Startup Telegram notification ────────────────────────────
    refresh_market_context(force=True)
    _now_ist  = ist_now()
    _session  = classify_session()
    _exit_at  = "15:28 IST"
    _vix_txt  = vix_label()
    _pcr_txt  = pcr_label()
    send_telegram(
        f"🟢 *NSE Options Bot — STARTED*\n"
        f"━━━━━━━━━━━━━━━━━━━━━\n"
        f"🕐 Time      : {_now_ist.strftime('%d %b %Y  %I:%M %p')} IST\n"
        f"📊 Session   : {_session}\n"
        f"⏱ Scan every : {SCAN_INTERVAL_MIN} minutes\n"
        f"🎯 Min score  : {MIN_SCORE_TO_ALERT}/100\n"
        f"📈 India VIX  : {_vix_txt}\n"
        f"⚖️ Nifty PCR  : {_pcr_txt}\n"
        f"━━━━━━━━━━━━━━━━━━━━━\n"
        f"Bot will scan stocks until {_exit_at}.\n"
        f"_Commands: STATUS · WIN · LOSS · SKIP · TUNE_"
    )

    # Start background Telegram command listener
    start_telegram_listener()

    if is_market_open():
        print("📡 Market is open — running first scan now...")
        scan_job()

    while True:
        schedule.run_pending()
        time.sleep(20)
        # Auto-exit at 3:28 PM IST so GitHub Actions "commit" step can run
        # within the 6-hour job limit (job starts ~9:30 AM → 5h58m → safe)
        _t = ist_now()
        if _t.hour * 60 + _t.minute >= 15 * 60 + 28:
            print(f"[{_t.strftime('%H:%M')} IST] Market session complete — bot exiting.")
            break


# ═══════════════════════════════════════════════════════════════════
#  SETUP GUIDE
# ═══════════════════════════════════════════════════════════════════
#
#  STEP 1 — Create Telegram bot
#    @BotFather → /newbot → copy token → paste in TELEGRAM_TOKEN
#
#  STEP 2 — Get Chat ID
#    @userinfobot → /start → copy ID → paste in CHAT_ID
#
#  STEP 3 — Install dependencies
#
#    Minimum (Telegram only):
#      pip install requests schedule pytz
#
#    With desktop notifications (recommended):
#      pip install requests schedule pytz plyer
#
#    plyer works on Windows, Mac, and Linux automatically.
#    If plyer is not installed, the bot uses OS-native fallbacks:
#      Windows → PowerShell toast notification
#      Mac     → osascript (built-in, nothing to install)
#      Linux   → notify-send (install: sudo apt install libnotify-bin)
#
#  STEP 4 — Test
#    python nse_options_bot.py test
#    You should see:
#      • A desktop pop-up notification appear
#      • A Telegram message arrive
#
#  STEP 5 — Run
#    python nse_options_bot.py
#    Keep the terminal window open during market hours.
#
#  TOGGLE CHANNELS (in config at top of file):
#    ENABLE_TELEGRAM       = True / False
#    ENABLE_DESKTOP_NOTIFY = True / False
#    DESKTOP_NOTIFY_SOUND  = True / False
#
#  HOW SMART ALERTING WORKS:
#  ──────────────────────────
#  Every 5 min:
#  1. Fetches fresh OI gainers from Trendlyne
#  2. Scores each stock: Day move + Volume + OI build = 0–100
#  3. Measures market activity (breadth + OI + volume)
#  4. Classifies session: Opening/Morning/Midday/Afternoon/PowerHour
#  5. Alerts ONLY when:
#       • NEW stock today + score > 40
#       • OR score jumped 15+ pts (momentum accelerating)
#       • AND market is active (quiet market needs score > 60)
#  6. Fires desktop pop-up AND Telegram simultaneously
#  7. Stops after 2 alerts (your daily limit)
#  8. EOD summary at 3:35 PM IST (desktop + Telegram)
# ═══════════════════════════════════════════════════════════════════
