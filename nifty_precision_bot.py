#!/usr/bin/env python3
"""
Nifty Precision Bot — Multi-Indicator Confluence
═══════════════════════════════════════════════════
Instruments : Nifty50 | BankNifty | Sensex
Strategy    : ORB + VWAP + EMA 9/21 + SuperTrend(10,3) + RSI(14) + CPR + PDH/PDL + Fib Pullback + VIX/PCR
Risk:Reward : 1:3 (Strong/Ultra signals) | 1:2 (Medium)
Target      : >90% accuracy via 6/8+ indicator confluence

Scoring (0–100):
  ORB breakout    20 pts   EMA 9/21 aligned  15 pts
  VWAP aligned    15 pts   SuperTrend align  15 pts
  RSI zone        10 pts   CPR position      10 pts
  PDH/PDL break   10 pts   Fib Pullback     +3–12 pts
  VIX/PCR bonus  ±5 pts

Scan schedule: every 5 min | 9:35 AM – 3:30 PM IST (weekdays)
Alert threshold : score ≥ 75 (Strong) or ≥ 85 (Ultra)
"""

import os, time, platform, subprocess, threading
from datetime import datetime, date, timedelta
from typing import Optional
import pytz
import schedule
import requests

try:
    import yfinance as yf
    import numpy as np
    import pandas as pd
    _yf = True
except ImportError:
    _yf = False
    print("[WARN] yfinance/pandas/numpy not installed — pip install yfinance pandas numpy")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _openpyxl = True
except ImportError:
    _openpyxl = False

# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
CHAT_ID        = os.environ.get("CHAT_ID", "")
_CI            = os.environ.get("CI", "false").lower() == "true"
TEST_MODE      = os.environ.get("TEST_MODE", "false").lower() == "true"

ENABLE_TELEGRAM       = True
ENABLE_DESKTOP_NOTIFY = not _CI
ENABLE_EXCEL_LOG      = True
EXCEL_FOLDER          = os.environ.get("EXCEL_FOLDER", "./logs")

# Instruments: yfinance ticker, default SL in index points, lot size, strike step
INSTRUMENTS = {
    "NIFTY50": {
        "ticker":       "^NSEI",
        "sl_pts":       40,
        "lot_size":     25,
        "strike_step":  50,
        "min_score":    75,
    },
    "BANKNIFTY": {
        "ticker":       "^NSEBANK",
        "sl_pts":       120,
        "lot_size":     15,
        "strike_step":  100,
        "min_score":    75,
    },
    "SENSEX": {
        "ticker":       "^BSESN",
        "sl_pts":       200,
        "lot_size":     10,
        "strike_step":  100,
        "min_score":    75,
    },
}

# Scoring thresholds
SCORE_ULTRA     = 85    # 1:3 R:R — highest conviction
SCORE_STRONG    = 75    # 1:3 R:R — strong
SCORE_MEDIUM    = 65    # 1:2 R:R — moderate (alert only)

# Indicator settings
RSI_PERIOD      = 14
RSI_CALL_LOW    = 40    # RSI lower bound for CALL entry
RSI_CALL_HIGH   = 65    # RSI upper bound for CALL entry
RSI_PUT_LOW     = 35
RSI_PUT_HIGH    = 60
CPR_NARROW_PCT  = 0.20  # < 0.20% width = trending day
ORB_MINUTES     = 15    # Opening range = first 15 min (9:15–9:30)

# Fibonacci Pullback
FIB_TOLERANCE_PCT = 0.15  # price within 0.15% of a Fib level = "at" it
FIB_RATIOS        = [0.236, 0.382, 0.500, 0.618, 0.786]
FIB_SCORE_MAP     = {0.618: 12, 0.500: 8, 0.382: 6, 0.786: 4, 0.236: 3}
ST_PERIOD       = 10    # SuperTrend period
ST_MULT         = 3.0   # SuperTrend multiplier

# India VIX filters
VIX_EXTREME     = 25.0  # Block all intraday signals
VIX_HIGH_PENALT = 10    # Score penalty when VIX 20–25
VIX_CALM_BONUS  = 5     # Score bonus when VIX < 12
VIX_CALM_MAX    = 12.0
VIX_NORMAL_MAX  = 20.0

# Nifty PCR
PCR_VERY_BULL   = 1.5
PCR_BULL        = 1.2
PCR_BEAR        = 0.8
PCR_VERY_BEAR   = 0.6

# Timing
SCAN_INTERVAL   = 5         # Every N minutes
MARKET_OPEN     = (9, 15)
MARKET_CLOSE    = (15, 30)
BOT_START       = (9, 35)   # After ORB forms
LATE_SESSION    = (14, 30)  # Late trades need score >= SCORE_ULTRA

MAX_TRADES_DAY  = 6         # Max alerts across all instruments per day
RESIGAL_GAP_MIN = 30        # Don't re-alert same instrument within N min unless score rises 10+
HEARTBEAT_SCANS = 12        # Send alive ping every N scans (~60 min) even with no signal

IST = pytz.timezone("Asia/Kolkata")

# ═══════════════════════════════════════════════════════════════
#  STATE
# ═══════════════════════════════════════════════════════════════

state = {
    "scan_count":     0,
    "alerts_sent":    0,
    "trades_today":   0,
    "vix":            None,
    "pcr":            None,
    "last_vix_fetch": None,
    "last_signal":    {},   # name → {direction, score, time}
    "data_errors":    {},   # name → consecutive fetch error count
    "scores_seen":    [],   # last 10 scores (for heartbeat)
}

# ═══════════════════════════════════════════════════════════════
#  TECHNICAL INDICATORS
# ═══════════════════════════════════════════════════════════════

def calc_ema(series, period: int):
    return series.ewm(span=period, adjust=False).mean()

def calc_rsi(series, period: int = 14) -> float:
    delta  = series.diff()
    gain   = delta.clip(lower=0)
    loss   = -delta.clip(upper=0)
    avg_g  = gain.ewm(com=period - 1, adjust=False).mean()
    avg_l  = loss.ewm(com=period - 1, adjust=False).mean()
    rs     = avg_g / avg_l.replace(0, 1e-10)
    return float((100 - 100 / (1 + rs)).iloc[-1])

def calc_vwap(df) -> float:
    """Session VWAP. Falls back to typical price if volume missing."""
    tp  = (df["High"] + df["Low"] + df["Close"]) / 3
    vol = df["Volume"]
    if vol.sum() == 0:
        return float(tp.iloc[-1])
    cum_vol    = vol.cumsum()
    cum_tp_vol = (tp * vol).cumsum()
    return float((cum_tp_vol / cum_vol.replace(0, 1e-10)).iloc[-1])

def calc_atr(df, period: int = 14):
    hl  = df["High"] - df["Low"]
    hc  = (df["High"] - df["Close"].shift()).abs()
    lc  = (df["Low"]  - df["Close"].shift()).abs()
    tr  = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    return tr.ewm(com=period - 1, adjust=False).mean()

def calc_supertrend(df, period: int = 10, mult: float = 3.0) -> int:
    """Returns +1 (bullish) or -1 (bearish) based on SuperTrend."""
    if len(df) < period + 2:
        return 1 if float(df["Close"].iloc[-1]) > float(df["Close"].iloc[0]) else -1

    hl2  = (df["High"] + df["Low"]) / 2
    atr_ = calc_atr(df, period)
    upper_raw = hl2 + mult * atr_
    lower_raw = hl2 - mult * atr_

    upper  = upper_raw.copy()
    lower  = lower_raw.copy()
    trend  = pd.Series(1, index=df.index, dtype=float)

    for i in range(1, len(df)):
        # Upper band
        if upper_raw.iloc[i] < upper.iloc[i - 1] or df["Close"].iloc[i - 1] > upper.iloc[i - 1]:
            upper.iloc[i] = upper_raw.iloc[i]
        else:
            upper.iloc[i] = upper.iloc[i - 1]
        # Lower band
        if lower_raw.iloc[i] > lower.iloc[i - 1] or df["Close"].iloc[i - 1] < lower.iloc[i - 1]:
            lower.iloc[i] = lower_raw.iloc[i]
        else:
            lower.iloc[i] = lower.iloc[i - 1]
        # Trend direction
        if trend.iloc[i - 1] == -1 and df["Close"].iloc[i] > upper.iloc[i]:
            trend.iloc[i] = 1
        elif trend.iloc[i - 1] == 1 and df["Close"].iloc[i] < lower.iloc[i]:
            trend.iloc[i] = -1
        else:
            trend.iloc[i] = trend.iloc[i - 1]

    return int(trend.iloc[-1])

def calc_cpr(prev_h: float, prev_l: float, prev_c: float) -> dict:
    """Central Pivot Range from previous day OHLC."""
    pivot     = (prev_h + prev_l + prev_c) / 3
    bc        = (prev_h + prev_l) / 2
    tc        = 2 * pivot - bc
    width_pct = abs(tc - bc) / pivot * 100
    return {
        "pivot":     pivot,
        "bc":        bc,
        "tc":        tc,
        "top":       max(tc, bc),
        "bottom":    min(tc, bc),
        "width_pct": width_pct,
        "narrow":    width_pct < CPR_NARROW_PCT,
    }

def calc_fib_pullback(df_today, direction: str, orb_h: float, orb_l: float) -> dict:
    """
    Fibonacci retracement pullback entry detector.

    CALL setup — move: ORB low → session high.
      Fib levels = support zones where price may bounce back up.
      Entry signal when price pulls back to 38.2%, 50%, or 61.8%.

    PUT setup — move: ORB high → session low.
      Fib levels = resistance zones where price may resume down.
      Entry signal when price bounces to 38.2%, 50%, or 61.8%.

    Returns a dict with the nearest Fib level, score bonus, and all levels.
    """
    empty = {"level": None, "price": None, "dist_pct": None,
             "at_fib": False, "score_bonus": 0,
             "levels": {}, "swing_high": None, "swing_low": None}

    if df_today is None or len(df_today) < 3:
        return empty

    ltp = float(df_today["Close"].iloc[-1])

    if direction == "CALL":
        # Anchor swing: ORB low as base, session high as peak
        swing_low  = orb_l
        swing_high = float(df_today["High"].max())
        move = swing_high - swing_low
        if move < 1:
            return empty
        # Pullback levels count DOWN from the high
        raw = {r: swing_high - move * r for r in FIB_RATIOS}
    else:
        # Anchor swing: ORB high as base, session low as trough
        swing_high = orb_h
        swing_low  = float(df_today["Low"].min())
        move = swing_high - swing_low
        if move < 1:
            return empty
        # Bounce levels count UP from the low
        raw = {r: swing_low + move * r for r in FIB_RATIOS}

    # Nearest level to current price
    nearest_ratio = min(raw, key=lambda r: abs(raw[r] - ltp))
    nearest_price = raw[nearest_ratio]
    dist_abs      = abs(ltp - nearest_price)
    dist_pct      = dist_abs / ltp * 100

    at_fib      = dist_pct <= FIB_TOLERANCE_PCT
    score_bonus = FIB_SCORE_MAP.get(nearest_ratio, 0) if at_fib else 0

    return {
        "level":      f"{nearest_ratio:.3f}",
        "price":      round(nearest_price, 1),
        "dist_pct":   round(dist_pct, 3),
        "at_fib":     at_fib,
        "score_bonus":score_bonus,
        "levels":     {f"{r:.3f}": round(p, 1) for r, p in raw.items()},
        "swing_high": round(swing_high, 1),
        "swing_low":  round(swing_low, 1),
    }

# ═══════════════════════════════════════════════════════════════
#  DATA FETCHING
# ═══════════════════════════════════════════════════════════════

_NSE_HDR = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept":          "application/json",
    "Referer":         "https://www.nseindia.com/",
}
_nse_sess = None

def _nse_session():
    global _nse_sess
    if _nse_sess is None:
        _nse_sess = requests.Session()
        _nse_sess.headers.update(_NSE_HDR)
        try:
            _nse_sess.get("https://www.nseindia.com", timeout=10)
        except Exception:
            pass
    return _nse_sess

def fetch_vix_pcr():
    vix_val = pcr_val = None
    try:
        r = _nse_session().get("https://www.nseindia.com/api/allIndices", timeout=10)
        if r.status_code == 200:
            for item in r.json().get("data", []):
                if item.get("index") == "INDIA VIX":
                    vix_val = float(item["last"])
                    break
    except Exception as e:
        print(f"[VIX] {e}")
    try:
        r = _nse_session().get(
            "https://www.nseindia.com/api/option-chain-indices?symbol=NIFTY",
            timeout=15
        )
        if r.status_code == 200:
            data   = r.json()["records"]["data"]
            ce_oi  = sum(d["CE"]["openInterest"] for d in data if "CE" in d)
            pe_oi  = sum(d["PE"]["openInterest"] for d in data if "PE" in d)
            pcr_val = round(pe_oi / ce_oi, 3) if ce_oi else None
    except Exception as e:
        print(f"[PCR] {e}")
    return vix_val, pcr_val

def fetch_ohlcv(ticker: str) -> Optional["pd.DataFrame"]:
    """Download 5-min OHLCV from yfinance with fallback periods, IST-converted."""
    if not _yf:
        return None
    # Try progressively broader periods in case Yahoo throttles short periods
    for period in ("5d", "7d", "1mo"):
        try:
            df = yf.download(ticker, period=period, interval="5m",
                             progress=False, auto_adjust=True)
            if df is None or df.empty:
                continue
            # Flatten MultiIndex columns (yfinance sometimes returns them)
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
            # Convert to IST
            if df.index.tz is None:
                df.index = df.index.tz_localize("UTC").tz_convert(IST)
            else:
                df.index = df.index.tz_convert(IST)
            df = df.dropna(subset=["Close", "High", "Low"])
            if len(df) >= 20:
                return df
        except Exception as e:
            print(f"[DATA] {ticker} (period={period}): {e}")
    return None

def split_sessions(df) -> tuple:
    """Split full OHLCV into today's and most recent previous trading day."""
    today_date = datetime.now(IST).date()
    today_mask = df.index.date == today_date

    df_today = df[today_mask].copy()
    df_prev  = df[~today_mask].copy()

    # Most recent previous trading day
    if not df_prev.empty:
        last_prev_date = max(df_prev.index.date)
        df_prev = df_prev[df_prev.index.date == last_prev_date].copy()

    return df_today, df_prev

def get_orb(df_today) -> tuple:
    """Opening Range High/Low: first ORB_MINUTES minutes of session."""
    if df_today.empty:
        return 0.0, 0.0
    orb_end = df_today.index[0].replace(
        hour=MARKET_OPEN[0],
        minute=MARKET_OPEN[1] + ORB_MINUTES,
        second=0
    )
    orb_df = df_today[df_today.index < orb_end]
    if orb_df.empty:
        orb_df = df_today.iloc[:3]
    return float(orb_df["High"].max()), float(orb_df["Low"].min())

# ═══════════════════════════════════════════════════════════════
#  SIGNAL ANALYSIS
# ═══════════════════════════════════════════════════════════════

def analyze(name: str, cfg: dict) -> Optional[dict]:
    df = fetch_ohlcv(cfg["ticker"])
    if df is None or len(df) < 20:
        data_error_ping(name, f"yfinance returned no/empty data for {cfg['ticker']}")
        return None

    df_today, df_prev = split_sessions(df)
    if df_today.empty or len(df_today) < 5:
        data_error_ping(name, f"No today's candles found (today date not in data)")
        return None
    if df_prev.empty:
        data_error_ping(name, "No previous-day data — can't compute CPR/PDH/PDL")
        return None

    reset_data_error(name)

    # ── Previous-day levels ────────────────────────────────────
    pdh = float(df_prev["High"].max())
    pdl = float(df_prev["Low"].min())
    pdc = float(df_prev["Close"].iloc[-1])
    cpr = calc_cpr(pdh, pdl, pdc)

    # ── Today indicators ───────────────────────────────────────
    close = df_today["Close"]
    ltp   = float(close.iloc[-1])

    e9    = float(calc_ema(close, 9).iloc[-1])
    e21   = float(calc_ema(close, 21).iloc[-1])
    vwap_ = calc_vwap(df_today)
    rsi_  = calc_rsi(close, RSI_PERIOD)
    st_   = calc_supertrend(df_today, ST_PERIOD, ST_MULT)

    orb_h, orb_l = get_orb(df_today)
    fib  = calc_fib_pullback(df_today, "CALL", orb_h, orb_l)  # recomputed per direction below

    # ── Determine direction (majority vote of 4 primary indicators) ─
    bull = sum([e9 > e21, ltp > vwap_, st_ == 1, ltp > orb_h])
    bear = sum([e9 < e21, ltp < vwap_, st_ == -1, ltp < orb_l])
    direction   = "CALL" if bull >= bear else "PUT"
    option_type = "CE"   if direction == "CALL" else "PE"

    # Recompute Fib now that direction is known
    fib = calc_fib_pullback(df_today, direction, orb_h, orb_l)

    # ── Scoring ────────────────────────────────────────────────
    score = 0
    cond  = {}   # condition name → True/False/str

    # 1. EMA 9/21 (15 pts)
    ema_ok = (e9 > e21 and direction == "CALL") or (e9 < e21 and direction == "PUT")
    cond["EMA 9/21"]  = ema_ok
    score += 15 if ema_ok else 0

    # 2. VWAP (15 pts)
    vwap_ok = (ltp > vwap_ and direction == "CALL") or (ltp < vwap_ and direction == "PUT")
    cond["VWAP"]  = vwap_ok
    score += 15 if vwap_ok else 0

    # 3. SuperTrend (15 pts)
    st_ok = (st_ == 1 and direction == "CALL") or (st_ == -1 and direction == "PUT")
    cond["SuperTrend"] = st_ok
    score += 15 if st_ok else 0

    # 4. ORB breakout (20 pts)
    orb_ok = (ltp > orb_h and direction == "CALL") or (ltp < orb_l and direction == "PUT")
    cond["ORB"] = orb_ok
    # Partial credit if price is within 0.05% of ORB level
    near_orb = abs(ltp - (orb_h if direction == "CALL" else orb_l)) / ltp < 0.0005
    score += 20 if orb_ok else (8 if near_orb else 0)

    # 5. RSI optimal zone (10 pts)
    if direction == "CALL":
        rsi_ok = RSI_CALL_LOW <= rsi_ <= RSI_CALL_HIGH
    else:
        rsi_ok = RSI_PUT_LOW <= rsi_ <= RSI_PUT_HIGH
    cond["RSI"] = rsi_ok
    score += 10 if rsi_ok else 0

    # 6. CPR position (10 pts)
    if direction == "CALL":
        cpr_ok = ltp > cpr["top"] and cpr["narrow"]
    else:
        cpr_ok = ltp < cpr["bottom"] and cpr["narrow"]
    cond["CPR"] = cpr_ok
    # Partial: narrow CPR even if price not yet through it
    score += 10 if cpr_ok else (4 if cpr["narrow"] else 0)

    # 7. PDH/PDL breakout (10 pts)
    pdhl_ok = (ltp > pdh and direction == "CALL") or (ltp < pdl and direction == "PUT")
    cond["PDH/PDL"] = pdhl_ok
    score += 10 if pdhl_ok else 0

    # 8. Fibonacci Pullback entry (+3 to +12 pts)
    if fib["at_fib"]:
        fib_label = f"0.{fib['level'].split('.')[1]} ({fib['price']:,.1f}) ✓ +{fib['score_bonus']}pts"
        cond["Fib Pullback"] = True
        score += fib["score_bonus"]
    elif fib["level"]:
        fib_label = f"nearest {fib['level']} ({fib['price']:,.1f})  dist {fib['dist_pct']:.2f}%"
        cond["Fib Pullback"] = False
    else:
        fib_label = "N/A"
        cond["Fib Pullback"] = False

    # 9. VIX adjustment (±10)
    vix_v = state["vix"]
    if vix_v:
        if vix_v > VIX_EXTREME:
            score -= 30
            cond["VIX"] = f"EXTREME {vix_v:.1f} ⛔"
        elif vix_v > VIX_NORMAL_MAX:
            score -= VIX_HIGH_PENALT
            cond["VIX"] = f"HIGH {vix_v:.1f} ⚠"
        elif vix_v < VIX_CALM_MAX:
            score += VIX_CALM_BONUS
            cond["VIX"] = f"CALM {vix_v:.1f} ✓"
        else:
            cond["VIX"] = f"NORMAL {vix_v:.1f}"

    # 9. PCR adjustment (±5)
    pcr_v = state["pcr"]
    if pcr_v:
        if direction == "CALL":
            bonus = 5 if pcr_v >= PCR_VERY_BULL else (3 if pcr_v >= PCR_BULL else
                   (-5 if pcr_v <= PCR_VERY_BEAR else 0))
        else:
            bonus = 5 if pcr_v <= PCR_VERY_BEAR else (3 if pcr_v <= PCR_BEAR else
                   (-5 if pcr_v >= PCR_VERY_BULL else 0))
        score += bonus
        cond["PCR"] = f"{pcr_v:.2f}"

    score = max(0, min(100, score))

    # ── Late session penalty ───────────────────────────────────
    now_ist = datetime.now(IST)
    late    = now_ist.hour > LATE_SESSION[0] or (
              now_ist.hour == LATE_SESSION[0] and now_ist.minute >= LATE_SESSION[1])
    if late and score < SCORE_ULTRA:
        score = max(0, score - 10)
        cond["Session"] = "LATE ⚠ (-10)"

    # ── Rating & R:R ───────────────────────────────────────────
    # Count all boolean True conditions (7 core + Fib = 8 max)
    conds_met = sum(1 for k, v in cond.items()
                    if isinstance(v, bool) and v)
    if score >= SCORE_ULTRA:
        rating, rr = "ULTRA STRONG", "1:3"
    elif score >= SCORE_STRONG:
        rating, rr = "STRONG", "1:3"
    elif score >= SCORE_MEDIUM:
        rating, rr = "MEDIUM", "1:2"
    else:
        rating, rr = "WEAK", "SKIP"

    # ── Trade levels ───────────────────────────────────────────
    sl_pts     = cfg["sl_pts"]
    target_pts = sl_pts * 3

    if direction == "CALL":
        sl_level     = round(ltp - sl_pts, 1)
        target_level = round(ltp + target_pts, 1)
        strike       = round(ltp / cfg["strike_step"]) * cfg["strike_step"]
    else:
        sl_level     = round(ltp + sl_pts, 1)
        target_level = round(ltp - target_pts, 1)
        strike       = round(ltp / cfg["strike_step"]) * cfg["strike_step"]

    return {
        "instrument":  name,
        "ticker":      cfg["ticker"],
        "direction":   direction,
        "option_type": option_type,
        "ltp":         ltp,
        "strike":      strike,
        "sl_level":    sl_level,
        "sl_pts":      sl_pts,
        "target_level":target_level,
        "target_pts":  target_pts,
        "score":       score,
        "rating":      rating,
        "rr":          rr,
        "conds_met":   conds_met,
        "conditions":  cond,
        "e9":          e9,
        "e21":         e21,
        "vwap":        vwap_,
        "rsi":         rsi_,
        "st_dir":      st_,
        "orb_high":    orb_h,
        "orb_low":     orb_l,
        "pdh":         pdh,
        "pdl":         pdl,
        "cpr":         cpr,
        "fib":         fib,
        "fib_label":   fib_label,
        "time":        now_ist.strftime("%H:%M"),
    }

# ═══════════════════════════════════════════════════════════════
#  NOTIFICATIONS
# ═══════════════════════════════════════════════════════════════

_OS = platform.system()

def _beep():
    try:
        if _OS == "Windows":
            import winsound
            winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        elif _OS == "Darwin":
            subprocess.Popen(["afplay", "/System/Library/Sounds/Glass.aiff"],
                             stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass

def _tkinter_popup(title: str, msg: str, urgent: bool = False):
    import tkinter as tk
    bg, accent = "#1e1e2e", "#e05c5c" if urgent else "#1D9E75"

    def _show():
        root = tk.Tk()
        root.overrideredirect(True)
        root.attributes("-topmost", True)
        root.configure(bg=bg)
        w, h = 420, 120
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        root.geometry(f"{w}x{h}+{sw-w-20}+{sh-h-60}")
        tk.Frame(root, bg=accent, height=4).pack(fill="x")
        tk.Label(root, text=title, font=("Segoe UI", 11, "bold"),
                 fg="#ffffff", bg=bg, anchor="w", padx=12).pack(fill="x", pady=(6, 0))
        tk.Label(root, text=msg, font=("Segoe UI", 9), fg="#aaaaaa", bg=bg,
                 anchor="w", padx=12, wraplength=400, justify="left").pack(fill="x")
        tk.Button(root, text="✕", font=("Segoe UI", 8), fg="#666", bg=bg,
                  bd=0, activebackground=bg, cursor="hand2",
                  command=root.destroy).place(x=w-24, y=6)
        root.after(12000, root.destroy)
        root.mainloop()

    threading.Thread(target=_show, daemon=True).start()

def send_desktop(title: str, msg: str, urgent: bool = False):
    if not ENABLE_DESKTOP_NOTIFY:
        return
    _beep()
    if _OS == "Windows":
        try:
            _tkinter_popup(title, msg, urgent)
            return
        except Exception:
            pass
    print(f"\n{'='*55}\n  {title}\n  {msg}\n{'='*55}\n")

def send_telegram(msg: str):
    if not TELEGRAM_TOKEN or not CHAT_ID:
        return
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            data={"chat_id": CHAT_ID, "text": msg,
                  "parse_mode": "Markdown", "disable_web_page_preview": True},
            timeout=10
        )
    except Exception as e:
        print(f"[TG] {e}")

def startup_ping():
    """Send Telegram the moment the bot starts — confirms it's alive."""
    now_str = datetime.now(IST).strftime("%d %b %Y %H:%M IST")
    mode    = "🧪 TEST MODE" if TEST_MODE else "🟢 LIVE"
    msg = (
        f"🚀 *Nifty Precision Bot — Started*\n"
        f"⏰ {now_str}\n"
        f"📊 Mode: {mode}\n"
        f"🎯 Scanning: {', '.join(INSTRUMENTS.keys())}\n"
        f"📈 Strategy: ORB + EMA + VWAP + ST + RSI + CPR + PDH/PDL + Fib\n"
        f"🔔 Alert threshold: score ≥ {SCORE_STRONG}\n"
        f"⏱ Scan every {SCAN_INTERVAL} min | Stops at 3:32 PM IST"
    )
    send_telegram(msg)
    print("[STARTUP] Telegram ping sent.")

def heartbeat_ping():
    """Periodic alive check — sent every HEARTBEAT_SCANS scans."""
    now_str = datetime.now(IST).strftime("%H:%M IST")
    recent  = state["scores_seen"][-6:] if state["scores_seen"] else []
    avg_sc  = round(sum(recent) / len(recent)) if recent else 0
    msg = (
        f"💓 *Nifty Precision Bot — Alive*\n"
        f"⏰ {now_str}\n"
        f"🔍 Scans: {state['scan_count']}  |  Alerts: {state['alerts_sent']}\n"
        f"📊 Avg score (recent): {avg_sc}/100\n"
        f"📈 VIX: {state['vix'] or '—'}  |  PCR: {state['pcr'] or '—'}\n"
        f"_No signal yet — watching for confluence…_"
    )
    send_telegram(msg)
    print(f"[HEARTBEAT] Sent at scan #{state['scan_count']}")

def data_error_ping(name: str, reason: str):
    """Send Telegram when a data fetch fails — 1st time and every 6th repeat."""
    err_count = state["data_errors"].get(name, 0) + 1
    state["data_errors"][name] = err_count
    if err_count == 1 or err_count % 6 == 0:
        send_telegram(
            f"⚠️ *Data Fetch Issue — {name}*\n"
            f"Reason: {reason}\n"
            f"Failures: {err_count}\n"
            f"_Bot still running, retrying next scan_"
        )
        print(f"[DATA ERROR] {name}: {reason} (failure #{err_count})")

def reset_data_error(name: str):
    state["data_errors"][name] = 0

def format_signal(sig: dict) -> str:
    d_icon   = "▲ CALL" if sig["direction"] == "CALL" else "▼ PUT"
    now_str  = datetime.now(IST).strftime("%d %b %Y %H:%M IST")
    rating_icon = {"ULTRA STRONG": "🔥", "STRONG": "💪", "MEDIUM": "⚡", "WEAK": "⚠"}.get(sig["rating"], "")

    cond_lines = []
    for k, v in sig["conditions"].items():
        if isinstance(v, bool):
            mark = "✅" if v else "❌"
            cond_lines.append(f"  {mark} {k}")
        else:
            cond_lines.append(f"  ℹ️ {k}: {v}")
    conds_str = "\n".join(cond_lines)

    cpr = sig["cpr"]
    return f"""
{rating_icon} *{sig['instrument']} — {d_icon} | {sig['rating']}*
📅 {now_str}

💰 *LTP:* {sig['ltp']:,.1f}
🎫 *Strike:* {sig['strike']} {sig['option_type']}
📊 *Score:* {sig['score']}/100 ({sig['conds_met']}/8 conditions)
⚖️ *R:R:* {sig['rr']}

📍 *Trade Levels:*
  Entry  : {sig['ltp']:,.1f}
  SL     : {sig['sl_level']:,.1f}  (−{sig['sl_pts']} pts)
  Target : {sig['target_level']:,.1f}  (+{sig['target_pts']} pts)

📈 *Indicators:*
  EMA9/21 : {sig['e9']:,.0f} / {sig['e21']:,.0f}
  VWAP    : {sig['vwap']:,.1f}
  RSI(14) : {sig['rsi']:.1f}
  ST      : {'↑ Bullish' if sig['st_dir'] == 1 else '↓ Bearish'}
  ORB     : {sig['orb_high']:,.1f} H / {sig['orb_low']:,.1f} L
  PDH/PDL : {sig['pdh']:,.1f} / {sig['pdl']:,.1f}
  CPR     : {cpr['bottom']:,.1f}–{cpr['top']:,.1f} ({'Narrow 📈' if cpr['narrow'] else 'Wide 📉'})

🔢 *Fib Pullback:*
  Swing   : {sig['fib'].get('swing_low', 0):,.1f} → {sig['fib'].get('swing_high', 0):,.1f}
  {sig['fib_label']}
  Levels  : 0.382={sig['fib']['levels'].get('0.382','—')}  0.500={sig['fib']['levels'].get('0.500','—')}  0.618={sig['fib']['levels'].get('0.618','—')}

📋 *Conditions:*
{conds_str}

⚠️ _For informational purposes only. Trade at your own risk._
""".strip()

# ═══════════════════════════════════════════════════════════════
#  EXCEL LOGGING
# ═══════════════════════════════════════════════════════════════

os.makedirs(EXCEL_FOLDER, exist_ok=True)
EXCEL_FILE = os.path.join(EXCEL_FOLDER, "Nifty_Precision_Bot.xlsx")

_HEADERS = [
    "Date", "Time", "Instrument", "Direction", "Strike", "Option",
    "LTP", "Score /100", "Rating", "R:R", "Conds Met /8",
    "EMA9", "EMA21", "VWAP", "RSI(14)", "SuperTrend",
    "ORB High", "ORB Low", "PDH", "PDL",
    "CPR Top", "CPR Bottom", "CPR Narrow?",
    "Fib Level", "Fib Price", "Fib Dist %", "Fib At Level?",
    "SL Level", "SL Pts", "Target Level", "Target Pts",
    "VIX", "PCR",
    "Entry Premium", "Exit Premium", "P&L (Rs)", "Result", "Notes",
]

_HDR_BG = "1F3864"
_CALL_FG = "276221"
_PUT_FG  = "9C0006"

def _init_excel():
    if not _openpyxl or os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Signals"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20
    for i, h in enumerate(_HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
    # Column widths
    widths = [12, 8, 12, 10, 9, 7, 10, 10, 14, 6, 12,
              10, 10, 10, 9, 11, 10, 10, 10, 10,
              10, 10, 12,
              10, 10, 10, 12,        # Fib Level, Price, Dist%, At Level?
              10, 8, 12, 10, 8, 8,
              14, 14, 12, 10, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    wb.save(EXCEL_FILE)

def log_excel(sig: dict):
    if not _openpyxl or not ENABLE_EXCEL_LOG:
        return
    try:
        _init_excel()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Signals"]
        now = datetime.now(IST)
        cpr = sig["cpr"]
        row = [
            now.strftime("%Y-%m-%d"), now.strftime("%H:%M"),
            sig["instrument"], sig["direction"], sig["strike"], sig["option_type"],
            round(sig["ltp"], 1), sig["score"], sig["rating"], sig["rr"], sig["conds_met"],
            round(sig["e9"], 1), round(sig["e21"], 1), round(sig["vwap"], 1),
            round(sig["rsi"], 1), "Bull" if sig["st_dir"] == 1 else "Bear",
            round(sig["orb_high"], 1), round(sig["orb_low"], 1),
            round(sig["pdh"], 1), round(sig["pdl"], 1),
            round(cpr["top"], 1), round(cpr["bottom"], 1),
            "Yes" if cpr["narrow"] else "No",
            sig["fib"].get("level") or "",
            sig["fib"].get("price") or "",
            sig["fib"].get("dist_pct") or "",
            "Yes" if sig["fib"].get("at_fib") else "No",
            round(sig["sl_level"], 1), sig["sl_pts"],
            round(sig["target_level"], 1), sig["target_pts"],
            state.get("vix") or "", state.get("pcr") or "",
            "", "", "", "", "",  # Entry/Exit Premium, P&L, Result, Notes
        ]
        ws.append(row)
        # Color direction column
        last_row = ws.max_row
        dc = ws.cell(row=last_row, column=4)
        dc.font = Font(
            bold=True,
            color=_CALL_FG if sig["direction"] == "CALL" else _PUT_FG
        )
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[EXCEL] {e}")

# ═══════════════════════════════════════════════════════════════
#  MAIN SCAN LOOP
# ═══════════════════════════════════════════════════════════════

def in_market_hours() -> bool:
    now  = datetime.now(IST)
    if now.weekday() >= 5:
        return False
    mins = now.hour * 60 + now.minute
    bot_start  = BOT_START[0]    * 60 + BOT_START[1]
    mkt_close  = MARKET_CLOSE[0] * 60 + MARKET_CLOSE[1]
    return bot_start <= mins <= mkt_close

def refresh_vix_pcr():
    last = state.get("last_vix_fetch")
    now  = datetime.now(IST)
    if last is None or (now - last).total_seconds() > 900:
        vix_v, pcr_v = fetch_vix_pcr()
        if vix_v is not None:
            state["vix"] = vix_v
        if pcr_v is not None:
            state["pcr"] = pcr_v
        state["last_vix_fetch"] = now
        print(f"[VIX/PCR] VIX={state['vix']}  PCR={state['pcr']}")

def scan():
    if not in_market_hours():
        return
    if state["trades_today"] >= MAX_TRADES_DAY:
        print("[LIMIT] Max trades for today reached.")
        return

    state["scan_count"] += 1
    print(f"\n{'─'*55}")
    print(f"[SCAN #{state['scan_count']}] {datetime.now(IST).strftime('%H:%M IST')}")

    # Heartbeat every N scans so user knows bot is alive even with no signals
    if state["scan_count"] % HEARTBEAT_SCANS == 0:
        heartbeat_ping()

    refresh_vix_pcr()

    if state["vix"] and state["vix"] > VIX_EXTREME:
        print(f"[BLOCK] VIX={state['vix']:.1f} — extreme volatility, skipping all signals")
        return

    for name, cfg in INSTRUMENTS.items():
        try:
            sig = analyze(name, cfg)
            if sig is None:
                continue

            state["scores_seen"].append(sig["score"])
            if len(state["scores_seen"]) > 30:
                state["scores_seen"] = state["scores_seen"][-30:]

            print(f"  [{name}] {sig['direction']} | Score={sig['score']} | {sig['rating']} | "
                  f"RSI={sig['rsi']:.1f} | EMA9/21={sig['e9']:.0f}/{sig['e21']:.0f} | "
                  f"ST={'↑' if sig['st_dir']==1 else '↓'}")

            for k, v in sig["conditions"].items():
                mark = "✓" if v is True else ("✗" if v is False else "·")
                print(f"      {mark} {k}: {v}")

            if sig["rating"] in ("WEAK", "MEDIUM") and sig["score"] < cfg["min_score"]:
                print(f"      [SKIP] Score {sig['score']} < threshold {cfg['min_score']}")
                continue

            # Suppress duplicate signal
            now = datetime.now(IST)
            last = state["last_signal"].get(name, {})
            if last:
                elapsed = (now - last["time"]).total_seconds() / 60
                same_dir = last["direction"] == sig["direction"]
                if same_dir and elapsed < RESIGAL_GAP_MIN and sig["score"] < last["score"] + 10:
                    print(f"      [SKIP] Same dir, {elapsed:.0f}m ago, score not improved")
                    continue

            # Send alert
            msg = format_signal(sig)
            send_telegram(msg)
            log_excel(sig)

            desk_title = f"{name} {sig['direction']} {sig['rating']}"
            desk_body  = (f"LTP {sig['ltp']:,.0f} | Score {sig['score']} | "
                          f"SL {sig['sl_level']:,.0f} → T {sig['target_level']:,.0f} ({sig['rr']})")
            send_desktop(desk_title, desk_body, urgent=(sig["score"] >= SCORE_ULTRA))

            state["last_signal"][name] = {
                "direction": sig["direction"],
                "score":     sig["score"],
                "time":      now,
            }
            state["alerts_sent"]  += 1
            state["trades_today"] += 1
            print(f"      [ALERTED] {name} {sig['direction']} {sig['score']}/100")

        except Exception as e:
            import traceback
            print(f"  [{name}] error: {e}")
            traceback.print_exc()

def eod_summary():
    now_str = datetime.now(IST).strftime("%d %b %Y")
    msg = (
        f"📊 *Nifty Precision Bot — EOD {now_str}*\n\n"
        f"🔍 Scans run    : {state['scan_count']}\n"
        f"📢 Signals sent : {state['alerts_sent']}\n"
        f"📈 VIX (last)   : {state['vix']}\n"
        f"📊 PCR (last)   : {state['pcr']}\n"
    )
    send_telegram(msg)
    send_desktop("Nifty Precision Bot — Market Closed",
                 f"Scans: {state['scan_count']}  Signals: {state['alerts_sent']}")
    print(f"\n{'='*55}")
    print(f"EOD | Scans: {state['scan_count']} | Alerts: {state['alerts_sent']}")
    print("=" * 55)

def main():
    print("=" * 55)
    print("  Nifty Precision Bot — Multi-Indicator Confluence")
    print(f"  Instruments : {', '.join(INSTRUMENTS)}")
    print(f"  Strategy    : ORB + EMA9/21 + VWAP + SuperTrend + RSI + CPR + PDH/PDL + Fib")
    print(f"  Alert score : ≥{SCORE_STRONG} (Strong) / ≥{SCORE_ULTRA} (Ultra)")
    print(f"  R:R target  : 1:3")
    print(f"  Test mode   : {TEST_MODE}")
    print("=" * 55)

    if not _yf:
        msg = "[FATAL] yfinance not installed. Run: pip install yfinance pandas numpy"
        print(msg)
        send_telegram(f"❌ *Nifty Precision Bot — FATAL*\n{msg}")
        return

    # Always ping Telegram on startup so user knows bot is alive
    startup_ping()

    # TEST MODE: verify Telegram works + data fetches, then exit
    if TEST_MODE:
        print("[TEST] Running quick data check for all instruments…")
        results = []
        for name, cfg in INSTRUMENTS.items():
            df = fetch_ohlcv(cfg["ticker"])
            if df is not None and len(df) >= 20:
                ltp = float(df["Close"].iloc[-1])
                results.append(f"  ✅ {name}: {len(df)} candles, LTP={ltp:,.1f}")
            else:
                results.append(f"  ❌ {name}: NO DATA from yfinance ({cfg['ticker']})")
        vix_v, pcr_v = fetch_vix_pcr()
        result_str = "\n".join(results)
        send_telegram(
            f"🧪 *Nifty Precision Bot — TEST RESULT*\n\n"
            f"{result_str}\n\n"
            f"📈 VIX: {vix_v or 'fetch failed'}  |  PCR: {pcr_v or 'fetch failed'}\n\n"
            f"_Test complete. Bot will run normally on schedule._"
        )
        print("[TEST] Done. Exiting.")
        return

    _init_excel()
    schedule.every(SCAN_INTERVAL).minutes.do(scan)
    schedule.every().day.at("15:35").do(eod_summary)

    if in_market_hours():
        scan()

    while True:
        schedule.run_pending()
        now = datetime.now(IST)
        # Auto-exit by 3:32 PM IST (~5h57m total from 9:35 AM start on GitHub Actions)
        if now.hour > 15 or (now.hour == 15 and now.minute >= 32):
            eod_summary()
            break
        time.sleep(30)

if __name__ == "__main__":
    main()
